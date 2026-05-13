#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import glob
import json
import fnmatch
import ftplib
import smtplib
import os
import re
import shutil
import requests
import sys
import time
import warnings
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from email.message import EmailMessage
from typing import Callable

from env_config import load_env_file

load_env_file()

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - handled at runtime
    load_workbook = None
    Workbook = None
else:
    warnings.filterwarnings(
        "ignore",
        message="Workbook contains no default style, apply openpyxl's default",
        module="openpyxl.styles.stylesheet",
    )

FTP_HOST = os.getenv("FTP_HOST", "")
FTP_USER = os.getenv("FTP_USER", "")
FTP_PASS = os.getenv("FTP_PASS", "")

LEGALSUITE_API_BASE = "https://api.legalsuite.net"
LEGALSUITE_API_KEY = os.getenv("LEGALSUITE_API_KEY", "")
LEGALSUITE_EMPLOYEE_ID = "1"
LEGALSUITE_ARCHIVE_STATUS = "2"
LEGALSUITE_OFFSET = 36161
EXCEL_BASE = dt.datetime(1899, 12, 30)

TARGETS = [
    ("SBSA/Debt Review/Debt_Review_Close_APT_LSW", "Standard_Bank_Panel_L_Close_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_PTP_APT_LSW", "Standard_Bank_Panel_L_PTP_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_Feedback_APT_LSW", "Standard_Bank_Panel_L_Update_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_Reopen_APT_LSW", "Standard_Bank_Panel_L_Reopen_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_Handover_APT_LWS", "Standard_Bank_Panel_L_Handover_{date}_DR.xlsx"),
    ("SBSA/Panel L/PTP_APT_LSW/{month_year}", "*_{date}.xlsx"),
    ("SBSA/Panel L/Feedback_APT_LSW/{month_year}", "*_{date}.xlsx"),
    ("SBSA/Panel L/Handover_APT_LSW/{month_year}", "*_{date}.xlsx"),
    ("SBSA/Panel L/Closed_APT_LSW/{month_year}", "*_{date}.xlsx"),
    ("SBSA/Panel L/Reopen_APT_LSW/{month_year}", "*_{date}.xlsx"),
    ("Standard Bank_ClaimsAmount", "Standard Bank Legal Claim Amount_Panel_L{year}_{month}_{day}_*.xlsx"),
    ("SBSA POC AND SUMMONS", "{day}{month}{year}.csv"),
]

REMOTE_DIR_ALIASES = {
    "SBSA/Debt Review/Debt_Review_Handover_APT_LWS": [
        "SBSA/Debt Review/Debt_Review_ Handover_APT_LWS",
    ],
}

CLAIM_AMOUNT_PATTERN = "Standard Bank Legal Claim Amount_Panel_L*.xlsx"
HANDOVER_PREFIXES = [
    "SBSA/Debt Review/Debt_Review_Handover_APT_LWS/",
    "SBSA/Debt Review/Debt_Review_ Handover_APT_LWS/",
    "SBSA/Panel L/Handover_APT_LSW/",
]

FEEDBACK_FIELD_MAP = [
    ("accountnumber", "field1", False),
    ("ptpcapturedate", "field2", True),
    ("ptpduedate", "field3", True),
    ("ptpamount", "field4", False),
    ("lastpaymentdate", "field5", True),
    ("lastpaymentamount", "field6", False),
    ("lastquickcomment", "field7", False),
    ("lastquickcommentdate", "field8", True),
    ("lastmemo", "field9", False),
    ("lastmemodate", "field10", True),
    ("accountcloseddate", "field11", True),
    ("reasonforclosure", "field12", False),
    ("branchid", "field13", False),
]

PTP_FIELD_MAP = [
    ("accountnumber", "field1", False),
    ("ptpcapturedate", "field2", True),
    ("ptpduedate", "field3", True),
    ("ptpamount", "field4", False),
    ("lastpaymentdate", "field5", True),
    ("branchid", "field6", False),
    ("lastquickcomment", "field7", False),
    ("ptpamount2", "field8", False),
    ("ptpduedate2", "field9", True),
    ("ptpamount3", "field10", False),
    ("ptpduedate3", "field11", True),
    ("ptpamount4", "field12", False),
    ("ptpduedate4", "field13", True),
    ("ptpamount5", "field14", False),
    ("ptpduedate5", "field15", True),
    ("ptpamount6", "field16", False),
    ("ptpduedate6", "field17", True),
    ("ptpamount7", "field18", False),
    ("ptpduedate7", "field19", True),
    ("ptpamount8", "field20", False),
    ("ptpduedate8", "field21", True),
    ("ptpamount9", "field22", False),
    ("ptpduedate9", "field23", True),
    ("ptpamount10", "field24", False),
    ("ptpduedate10", "field25", True),
    ("ptpamount11", "field26", False),
    ("ptpduedate11", "field27", True),
    ("ptpamount12", "field28", False),
    ("ptpduedate12", "field29", True),
]

POC_SUMMONS_FIELD_MAP = [
    ("noofcallattempts", "field2", False),
    ("noofdispatchedsmss", "field3", False),
    ("noofdispatchedemails", "field4", False),
    ("noofbrokenptps", "field5", False),
]

CLIENT_CODE_MAP = {
    "STA387": "150307",
    "DR387": "334695",
    "STD9": "155128",
    "DRR9": "334565",
    "STA482": "209250",
    "DR482": "334568",
    "STA822": "283850",
    "DR822": "334567",
    "STA614": "267742",
    "DR614": "334569",
}

CREATE_FIELD_COLUMN_MAP = {
    "theirref": "Reference",
    "claimamount": "Claim Amount",
    "interestrate": "Interest Rate",
    "employerid": "EmployerID",
    "tracingagentid": "TracingAgentID",
    "sheriffareaid": "SheriffAreaID",
    "sheriffid": "SheriffID",
    "branchid": "BranchID",
    "employeeid": "EmployeeID",
    "stagegroupid": "StageGroupID",
    "mattertypeid": "MatterTypeID",
    "debtorfeesheetid": "DebtorFeeSheetID",
    "clientfeesheetid": "ClientFeeSheetID",
    "debtorcollcommoption": "DebtorCollCommOption",
    "debtorcollcommpercent": "DebtorCollCommPercent",
    "collcommoption": "CollCommOption",
    "clientcollcommpercent": "ClientCollCommPercent",
    "costcentreid": "CostCentreID",
    "defendantemail": "DefendantEmail",
    "magcourtdistrict": "MagCourtDistrict",
    "magcourtheldat": "MagCourtHeldAt",
    "extrascreenid": "ExtraScreenID",
    "induplumamount": "In Duplum Amount",
    "maximuminterestamount": "Maximum Interest Amount",
    "alternateref": "Alternate Reference",
}

CREATE_DEFAULTS = {
    "mattertypeid": 4,
    "clientfeesheetid": 1,
    "docgenid": 5,
    "costcentreid": 193,
    "todogroupid": 1,
    "stagegroupid": 59,
    "debtorfeesheetid": 45,
    "businessbankid": 1103,
    "trustbankid": 1198,
    "employeeid": 174,
    "loggedinemployeeid": 174,
    "archivestatusdescription": "LIVE",
    "archivestatus": 2,
    "partyname": "STANDARD BANK",
    "employeename": "Angie Reddy",
    "mattypedescription": "Collections",
    "docgendescription": "Magistrate Court",
    "branchdescription": "STRAUSS DALY UMHLANGA",
    "docgencode": "MAG",
    "docgentype": "LIT",
    "costcentredescription": "A Reddy",
    "planofactiondescription": "Collections",
    "stagegroupdescription": "Collections APT - E4",
    "clientfeesheetdescription": "Conveyancing",
    "debtorfeesheetdescription": "Mag Court Tariff",
    "backgroundcolour": "FF0000",
}

PARTY_DEFAULTS = {
    "party[partytypeid]": 1,
    "party[defaultlanguageid]": 1,
    "party[createdid]": 1,
    "party[parlang][languageid]": 1,
}

MATPARTY_DEFAULTS = {
    "roleid": 103,
    "sorter": 1,
    "languageid": 0,
    "reference": "",
}

HANDOVER_REPORT_TO = [
    "tnxumalo@straussdaly.co.za",
    "areddy@straussdaly.co.za",
    "gharris@straussdaly.co.za",
    "defbloem@straussdaly.co.za",
]

HANDOVER_REPORT_CC = [
    "agashnee.pillay@iconis.co.za",
    "thileshnee.chinnasamy@iconis.co.za",
]

HANDOVER_REPORT_TEST_TO = [
    "dev@iconis.co.za",
]

HANDOVER_REPORT_TEST_CC = [
    "boyiajas@gmail.com",
]

LEGALSUITE_MAX_ATTEMPTS = 3
LEGALSUITE_RETRY_DELAYS = (2, 5)

XLSX_NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
}

@dataclass(frozen=True)
class DateContext:
    date_str: str
    month_year: str
    year: str
    month: str
    day: str


@dataclass
class HandoverRow:
    source_path: str
    row_number: int
    headers: list[str]
    row_values: tuple[object, ...]
    values_by_header: dict[str, object]
    client_code: str
    client_id: str
    reference: str | None


@dataclass(frozen=True)
class HandoverCreatedMatter:
    file_ref: str
    their_reference: str
    description: str


@dataclass
class VerificationWorkbookState:
    source_path: str
    verification_path: str
    workbook: object
    header_indexes: dict[str, dict[str, int]]


class VerificationWorkbookRecorder:
    def __init__(self, verification_dir: str, path_roots: list[str]) -> None:
        self._verification_dir = os.path.abspath(verification_dir)
        self._path_roots = [os.path.abspath(path) for path in path_roots if path]
        self._states: dict[str, VerificationWorkbookState] = {}

    def record_row(
        self,
        source_path: str,
        row_number: int,
        status: str,
        notes: str,
        get_response: object | None,
        verified_values: dict[str, object] | None = None,
        worksheet_name: str | None = None,
    ) -> str:
        state = self._ensure_state(source_path)
        worksheet = self._resolve_worksheet(state, worksheet_name)
        values = {
            "Verification Status": status,
            "Verification Timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Verification Notes": notes,
            "Verification GET Response": self._serialize_response(get_response),
        }
        if verified_values:
            values.update(verified_values)

        for header_name, value in values.items():
            column_idx = self._ensure_column(state, worksheet.title, header_name)
            worksheet.cell(row=row_number, column=column_idx).value = value

        return state.verification_path

    def finalize(self) -> list[str]:
        saved_paths: list[str] = []
        for source_path, state in list(self._states.items()):
            state.workbook.save(state.verification_path)
            state.workbook.close()
            saved_paths.append(state.verification_path)
            del self._states[source_path]
        return sorted(saved_paths)

    def _ensure_state(self, source_path: str) -> VerificationWorkbookState:
        source_abs = os.path.abspath(source_path)
        state = self._states.get(source_abs)
        if state is not None:
            return state

        verification_path = self._verification_path(source_abs)
        os.makedirs(os.path.dirname(verification_path), exist_ok=True)
        shutil.copy2(source_abs, verification_path)
        workbook = load_workbook(verification_path, read_only=False, data_only=False)
        state = VerificationWorkbookState(
            source_path=source_abs,
            verification_path=verification_path,
            workbook=workbook,
            header_indexes={},
        )
        self._states[source_abs] = state
        return state

    def _verification_path(self, source_path: str) -> str:
        for root in self._path_roots:
            try:
                rel_path = os.path.relpath(source_path, root)
            except ValueError:
                continue
            if rel_path == ".":
                return os.path.join(self._verification_dir, os.path.basename(source_path))
            if not rel_path.startswith(f"..{os.sep}") and rel_path != "..":
                return os.path.join(self._verification_dir, rel_path)
        return os.path.join(self._verification_dir, os.path.basename(source_path))

    @staticmethod
    def _normalize_header(value: object) -> str:
        if value is None:
            return ""
        return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())

    @staticmethod
    def _serialize_response(response: object | None) -> str:
        if response in (None, ""):
            return ""
        text = json.dumps(response, default=str, ensure_ascii=True)
        if len(text) > 32000:
            return text[:31997] + "..."
        return text

    @staticmethod
    def _resolve_worksheet(state: VerificationWorkbookState, worksheet_name: str | None):
        if worksheet_name and worksheet_name in state.workbook.sheetnames:
            return state.workbook[worksheet_name]
        return state.workbook.active

    def _ensure_column(self, state: VerificationWorkbookState, worksheet_name: str, header_name: str) -> int:
        header_index = state.header_indexes.get(worksheet_name)
        worksheet = state.workbook[worksheet_name]
        if header_index is None:
            header_index = {}
            max_col = worksheet.max_column or 1
            for idx in range(1, max_col + 1):
                key = self._normalize_header(worksheet.cell(row=1, column=idx).value)
                if key and key not in header_index:
                    header_index[key] = idx
            state.header_indexes[worksheet_name] = header_index

        normalized_name = self._normalize_header(header_name)
        existing_idx = header_index.get(normalized_name)
        if existing_idx is not None:
            return existing_idx

        column_idx = (worksheet.max_column or 0) + 1
        worksheet.cell(row=1, column=column_idx).value = header_name
        header_index[normalized_name] = column_idx
        return column_idx


class FTPClient:
    def __init__(self, host: str, user: str, password: str, timeout: int) -> None:
        self._host = host
        self._user = user
        self._password = password
        self._timeout = timeout
        self._ftp: ftplib.FTP | None = None

    def connect(self) -> None:
        missing = [
            name
            for name, value in (
                ("FTP_HOST", self._host),
                ("FTP_USER", self._user),
                ("FTP_PASS", self._password),
            )
            if not value
        ]
        if missing:
            raise ValueError(f"Missing FTP credentials: {', '.join(missing)}")
        self._ftp = ftplib.FTP(self._host, timeout=self._timeout)
        self._ftp.login(self._user, self._password)
        self._ftp.set_pasv(True)

    def close(self) -> None:
        if not self._ftp:
            return
        try:
            self._ftp.quit()
        except ftplib.all_errors:
            self._ftp.close()

    def list_dir(self, remote_dir: str) -> list[str] | None:
        ftp = self._require()
        try:
            return ftp.nlst(remote_dir)
        except ftplib.error_perm:
            pass
        current_dir = None
        try:
            current_dir = ftp.pwd()
            ftp.cwd(remote_dir)
            listing = ftp.nlst()
            return [f"{remote_dir}/{name}" for name in listing]
        except ftplib.error_perm:
            return None
        finally:
            if current_dir:
                try:
                    ftp.cwd(current_dir)
                except ftplib.all_errors:
                    pass

    def mdtm_timestamp(self, remote_path: str) -> dt.datetime | None:
        ftp = self._require()
        try:
            response = ftp.sendcmd(f"MDTM {remote_path}")
        except ftplib.all_errors:
            return None
        parts = response.split()
        if len(parts) != 2 or parts[0] != "213":
            return None
        try:
            return dt.datetime.strptime(parts[1], "%Y%m%d%H%M%S")
        except ValueError:
            return None

    def select_newest_by_mdtm(self, remote_dir: str, names: list[str]) -> str:
        candidates: list[tuple[dt.datetime, str]] = []
        for name in names:
            ts = self.mdtm_timestamp(f"{remote_dir}/{name}")
            if ts:
                candidates.append((ts, name))
        if candidates:
            candidates.sort()
            return candidates[-1][1]
        names.sort()
        return names[-1]

    def resolve_remote_file(self, remote_dir: str, filename_tmpl: str) -> tuple[str | None, str | None]:
        listing = self.list_dir(remote_dir)
        if listing is None:
            return None, "missing_dir"
        names = [os.path.basename(item) for item in listing]
        if any(ch in filename_tmpl for ch in ["*", "?", "["]):
            matches = [name for name in names if fnmatch.fnmatch(name, filename_tmpl)]
            if not matches:
                return None, "missing_file"
            return self.select_newest_by_mdtm(remote_dir, matches), None
        if filename_tmpl in names:
            return filename_tmpl, None
        ftp = self._require()
        try:
            ftp.size(f"{remote_dir}/{filename_tmpl}")
            return filename_tmpl, None
        except ftplib.all_errors:
            return None, "missing_file"

    def download_file(self, remote_path: str, local_path: str) -> None:
        ftp = self._require()
        with open(local_path, "wb") as handle:
            ftp.retrbinary(f"RETR {remote_path}", handle.write)

    def _require(self) -> ftplib.FTP:
        if not self._ftp:
            raise RuntimeError("FTP client not connected")
        return self._ftp


def post_with_retry(
    url: str,
    headers: dict[str, str],
    data,
    timeout: int,
) -> requests.Response:
    last_exc: Exception | None = None
    for attempt in range(1, LEGALSUITE_MAX_ATTEMPTS + 1):
        try:
            response = requests.post(url, headers=headers, data=data, timeout=timeout)
            if response.status_code >= 500:
                response.raise_for_status()
            return response
        except (requests.Timeout, requests.ConnectionError, requests.HTTPError) as exc:
            retryable = isinstance(exc, (requests.Timeout, requests.ConnectionError))
            if isinstance(exc, requests.HTTPError):
                status_code = exc.response.status_code if exc.response is not None else None
                retryable = status_code is not None and status_code >= 500

            last_exc = exc
            if not retryable or attempt >= LEGALSUITE_MAX_ATTEMPTS:
                raise

            delay = LEGALSUITE_RETRY_DELAYS[min(attempt - 1, len(LEGALSUITE_RETRY_DELAYS) - 1)]
            print(
                "LegalSuite request failed "
                f"(attempt {attempt}/{LEGALSUITE_MAX_ATTEMPTS}): {exc}. "
                f"Retrying in {delay}s..."
            )
            time.sleep(delay)

    if last_exc:
        raise last_exc
    raise RuntimeError("LegalSuite request failed without an exception")


class LegalSuiteClient:
    def __init__(self, api_base: str, api_key: str) -> None:
        self._api_base = api_base.rstrip("/")
        self._api_key = api_key

    def get_matter_by_fileref(self, file_ref: str) -> dict:
        url = f"{self._api_base}/matter/get"
        data = {
            "where[]": f"Matter.FileRef,=,{file_ref}",
        }
        resp = post_with_retry(url, headers=self._headers(), data=data, timeout=60)
        resp.raise_for_status()
        payload = resp.json()
        items = payload.get("data", [])
        if not items:
            raise ValueError(f"No matter found for FileRef: {file_ref}")
        return items[0]

    def update_matter(self, payload: dict) -> dict:
        url = f"{self._api_base}/matter/update"
        resp = post_with_retry(url, headers=self._headers(), data=payload, timeout=60)
        resp.raise_for_status()
        try:
            return resp.json()
        except ValueError:
            return {"raw_response": resp.text}

    def update_matter_extrascreen(self, payload: dict) -> dict:
        url = f"{self._api_base}/matdocsc/update"
        resp = post_with_retry(url, headers=self._headers(), data=payload, timeout=60)
        resp.raise_for_status()
        try:
            return resp.json()
        except ValueError:
            return {"raw_response": resp.text}

    def get_matter_extrascreen(self, matter_id: int | str, docscreenid: int | str) -> list[dict]:
        url = f"{self._api_base}/matdocsc/get"
        data = [
            ("where[]", f"MatDocSc.MatterID,=,{matter_id}"),
            ("where[]", f"MatDocSc.DocScreenID,=,{docscreenid}"),
        ]
        resp = post_with_retry(url, headers=self._headers(), data=data, timeout=60)
        resp.raise_for_status()
        payload = resp.json()
        return payload.get("data", [])

    @staticmethod
    def build_archive_payload(
        matter: dict,
        logged_in_employee_id: str,
        archive_no: str | None = None,
        archive_status: str | None = None,
    ) -> dict:
        now = dt.datetime.now()
        payload = {}
        for key, value in matter.items():
            if isinstance(value, (dict, list, tuple, set)):
                continue
            payload[key] = value
        payload.update(
            {
                "loggedinemployeeid": str(logged_in_employee_id),
                "archiveflag": "1",
                "archivestatusdescription": "Archived",
                "formattedupdatedbydate": now.strftime("%d %b %Y"),
                "formattedupdatedbytime": now.strftime("%H:%M:%S"),
            }
        )
        for field_name in ("actual", "reserved", "invested", "transfer", "batchednormal"):
            payload[field_name] = matter.get(field_name)
        if archive_status is not None:
            payload["archivestatus"] = str(archive_status)
        if archive_no is not None:
            payload["archiveno"] = str(archive_no)
        return {k: v for k, v in payload.items() if v not in ("", None)}

    @staticmethod
    def build_pending_deletion_payload(
        matter: dict,
        logged_in_employee_id: str,
        archive_no: str | None = None,
    ) -> dict:
        now = dt.datetime.now()
        payload = {}
        for key, value in matter.items():
            if isinstance(value, (dict, list, tuple, set)):
                continue
            payload[key] = value
        payload.update(
            {
                "loggedinemployeeid": str(logged_in_employee_id),
                "archiveflag": "0",
                "archivestatus": "1",
                "archivestatusdescription": "Pending Deletion",
                "formattedupdatedbydate": now.strftime("%d %b %Y"),
                "formattedupdatedbytime": now.strftime("%H:%M:%S"),
            }
        )
        if archive_no is not None:
            payload["archiveno"] = str(archive_no)
        return {k: v for k, v in payload.items() if v not in ("", None)}

    @staticmethod
    def build_reopen_payload(
        matter: dict,
        logged_in_employee_id: str,
    ) -> dict:
        now = dt.datetime.now()
        payload = {}
        for key, value in matter.items():
            if isinstance(value, (dict, list, tuple, set)):
                continue
            payload[key] = value
        payload.update(
            {
                "loggedinemployeeid": str(logged_in_employee_id),
                "archiveflag": "0",
                "archivestatus": "0",
                "archiveno": "0",
                "archivestatusdescription": "Live",
                "archivedate": "",
                "formattedupdatedbydate": now.strftime("%d %b %Y"),
                "formattedupdatedbytime": now.strftime("%H:%M:%S"),
            }
        )
        return {
            key: value
            for key, value in payload.items()
            if value is not None and (value != "" or key == "archivedate")
        }

    @staticmethod
    def build_claim_amount_payload(
        matter: dict,
        logged_in_employee_id: str,
        claim_amount: object,
    ) -> dict:
        payload = {}
        for key, value in matter.items():
            if isinstance(value, (dict, list, tuple, set)):
                continue
            payload[key] = value
        payload["loggedinemployeeid"] = str(logged_in_employee_id)
        payload["claimamount"] = claim_amount
        return {k: v for k, v in payload.items() if v not in ("", None)}

    @staticmethod
    def build_claim_amount_fileref_only_payload(
        matter: dict,
        file_ref: str,
        logged_in_employee_id: str,
        claim_amount: object,
    ) -> dict:
        payload = {}
        for key, value in matter.items():
            if key == "oldcode" or isinstance(value, (dict, list, tuple, set)):
                continue
            payload[key] = value
        payload["fileref"] = file_ref
        payload["loggedinemployeeid"] = str(logged_in_employee_id)
        payload["claimamount"] = claim_amount
        return {k: v for k, v in payload.items() if v not in ("", None)}

    def _headers(self) -> dict:
        return {
            "Authorization": f"Bearer {self._api_key}",
            "Content-Type": "application/x-www-form-urlencoded",
        }


def extract_update_error_text(result: dict | str) -> str:
    if isinstance(result, dict):
        errors = result.get("errors")
        if errors:
            return str(errors)
        raw_response = result.get("raw_response")
        if raw_response:
            return str(raw_response)
    return str(result or "")


def is_archive_rejected_error(result: dict | str) -> bool:
    error_text = extract_update_error_text(result).lower()
    return (
        "you cannot archive a matter" in error_text
        or "you cannot archieve a matter" in error_text
    )


def is_old_code_unique_error(result: dict | str) -> bool:
    error_text = extract_update_error_text(result).lower()
    return "already has this old code" in error_text and "old code must be unique" in error_text


class LegalSuiteLookupClient:
    def __init__(self, api_base: str, api_key: str) -> None:
        self._api_base = api_base.rstrip("/")
        self._api_key = api_key

    def get_matters_by_clientid_and_prefix(self, client_id: str, prefix: str) -> list[dict]:
        url = f"{self._api_base}/matter/get"
        data = [
            ("where[]", f"Matter.ClientID,=,{client_id}"),
            ("where[]", f"Matter.FileRef,like,{prefix}/%"),
        ]
        response = post_with_retry(url, headers=self._headers(), data=data, timeout=120)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def get_matters_by_fileref(self, file_ref: str) -> list[dict]:
        url = f"{self._api_base}/matter/get"
        data = {
            "where[]": f"Matter.FileRef,=,{file_ref}",
        }
        response = post_with_retry(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def get_matters_by_clientid_and_reference(self, client_id: str, reference: str) -> list[dict]:
        url = f"{self._api_base}/matter/get"
        data = [
            ("where[]", f"Matter.ClientID,=,{client_id}"),
            ("where[]", f"Matter.TheirRef,=,{reference}"),
        ]
        response = post_with_retry(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def create_matter(self, data: dict) -> dict | str:
        url = f"{self._api_base}/matter/store"
        payload = {key: str(value) for key, value in data.items()}
        response = post_with_retry(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def get_matter_by_recordid(self, recordid: int | str) -> dict | str:
        url = f"{self._api_base}/matter/get"
        data = {
            "where[]": f"Matter.RecordID,=,{recordid}",
        }
        response = post_with_retry(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def update_matter(self, recordid: int | str, updates: dict) -> dict | str:
        url = f"{self._api_base}/matter/update"
        payload = {
            "recordid": str(recordid),
        }
        for key, value in updates.items():
            payload[key] = str(value)
        response = post_with_retry(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def create_party(self, data: dict) -> dict | str:
        url = f"{self._api_base}/party/store"
        payload = {key: str(value) for key, value in data.items()}
        response = post_with_retry(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def get_party_by_identitynumber(self, identity_number: str) -> list[dict]:
        url = f"{self._api_base}/party/get"
        data = {
            "where[]": f"Party.IdentityNumber,=,{identity_number}",
        }
        response = post_with_retry(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def get_matparty_by_matter_and_party(self, matter_id: int | str, party_id: int | str) -> list[dict]:
        url = f"{self._api_base}/matparty/get"
        data = [
            ("where[]", f"MatParty.MatterID,=,{matter_id}"),
            ("where[]", f"MatParty.PartyID,=,{party_id}"),
        ]
        response = post_with_retry(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def create_matparty(self, data: dict) -> dict | str:
        url = f"{self._api_base}/matparty/store"
        payload = {key: str(value) for key, value in data.items()}
        response = post_with_retry(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def update_matter_extrascreen(self, data: dict) -> dict | str:
        url = f"{self._api_base}/matdocsc/update"
        payload = {key: str(value) for key, value in data.items()}
        response = post_with_retry(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def get_matter_extrascreen(self, matter_id: int | str, docscreenid: int | str) -> list[dict]:
        url = f"{self._api_base}/matdocsc/get"
        data = [
            ("where[]", f"MatDocSc.MatterID,=,{matter_id}"),
            ("where[]", f"MatDocSc.DocScreenID,=,{docscreenid}"),
        ]
        response = post_with_retry(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def _headers(self) -> dict[str, str]:
        return {
            "Authorization": f"Bearer {self._api_key}",
            "Content-Type": "application/x-www-form-urlencoded",
        }


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def cell_value(row: tuple[object, ...], col_idx: int) -> object | None:
    if col_idx >= len(row):
        return None
    return row[col_idx]


def normalize_reference(value: object) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("'"):
        text = text[1:].strip()
    return text or None


def normalize_cell_value(value: object) -> object | None:
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("'"):
            text = text[1:].strip()
        return text or None
    return value


def normalize_money(value: object) -> object | None:
    value = normalize_cell_value(value)
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if not cleaned or cleaned in {"-", ".", "-."}:
        return None
    try:
        return float(cleaned) if "." in cleaned else int(cleaned)
    except ValueError:
        return None


def digits_only(value: object) -> str | None:
    if value in (None, ""):
        return None
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    return digits or None


def encode_legalsuite_date(value: object) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return int(value) + LEGALSUITE_OFFSET
    if isinstance(value, dt.datetime):
        date_value = value
    elif isinstance(value, dt.date):
        date_value = dt.datetime.combine(value, dt.time())
    else:
        text = str(value).strip()
        if not text:
            return None
        if re.fullmatch(r"-?\d+(?:\.0+)?", text):
            return int(float(text)) + LEGALSUITE_OFFSET
        date_text = text.split()[0].split("T")[0].replace("/", "-")
        parts = date_text.split("-")
        if len(parts) == 3 and all(parts):
            if len(parts[0]) == 4:
                date_text = f"{parts[0]}-{parts[1]}-{parts[2]}"
            elif len(parts[2]) == 4:
                date_text = f"{parts[2]}-{parts[1]}-{parts[0]}"
        try:
            date_value = dt.datetime.strptime(date_text, "%Y-%m-%d")
        except ValueError:
            return None
    excel_serial = (date_value - EXCEL_BASE).days
    return excel_serial + LEGALSUITE_OFFSET


def encode_legalsuite_time(value: dt.datetime | dt.time | None = None) -> int:
    if value is None:
        value = dt.datetime.now()
    if isinstance(value, dt.datetime):
        time_value = value.time()
    else:
        time_value = value
    return int(time_value.strftime("%H%M%S"))


def find_column_index(header_row: tuple[object, ...], expected_name: str) -> int | None:
    expected_key = normalize_header(expected_name)
    for idx, value in enumerate(header_row):
        if normalize_header(value) == expected_key:
            return idx
    return None


def find_header_row(rows, expected_name: str, max_scan_rows: int = 10) -> tuple[tuple[object, ...], int, int]:
    scanned_rows: list[tuple[object, ...]] = []
    for row_number, row in enumerate(rows, start=1):
        row_tuple = tuple(row)
        scanned_rows.append(row_tuple)
        column_idx = find_column_index(row_tuple, expected_name)
        if column_idx is not None:
            return row_tuple, column_idx, row_number
        if row_number >= max_scan_rows:
            break

    sample_rows = [", ".join(str(value or "") for value in row[:8]) for row in scanned_rows[:3]]
    sample_text = " | ".join(sample_rows) if sample_rows else "<no rows scanned>"
    raise ValueError(
        f"{expected_name} column not found in first {max_scan_rows} row(s). "
        f"Header sample: {sample_text}"
    )


def iter_excel_rows(path: str):
    if path.lower().endswith(".xlsx"):
        try:
            yield from iter_xlsx_rows_stdlib(path)
            return
        except (zipfile.BadZipFile, KeyError, ET.ParseError, ValueError):
            pass

    if load_workbook is None:
        yield from iter_xlsx_rows_stdlib(path)
        return

    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        if not workbook.worksheets:
            return
        worksheet = workbook.worksheets[0]
        for row in worksheet.iter_rows(values_only=True):
            yield tuple(row)
    finally:
        workbook.close()


def iter_xlsx_rows_stdlib(path: str):
    with zipfile.ZipFile(path) as workbook_zip:
        shared_strings = read_shared_strings(workbook_zip)
        sheet_path = first_sheet_path(workbook_zip)
        sheet_root = ET.fromstring(workbook_zip.read(sheet_path))
        for row in sheet_root.findall("a:sheetData/a:row", XLSX_NS):
            yield read_xlsx_row(row, shared_strings)


def read_shared_strings(workbook_zip: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in workbook_zip.namelist():
        return []

    root = ET.fromstring(workbook_zip.read("xl/sharedStrings.xml"))
    values: list[str] = []
    for item in root.findall("a:si", XLSX_NS):
        text = "".join(node.text or "" for node in item.findall(".//a:t", XLSX_NS))
        values.append(text)
    return values


def first_sheet_path(workbook_zip: zipfile.ZipFile) -> str:
    workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
    first_sheet = workbook_root.find("a:sheets/a:sheet", XLSX_NS)
    if first_sheet is None:
        raise ValueError("Workbook contains no sheets")

    rel_id = first_sheet.attrib.get(
        "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    )
    if not rel_id:
        raise ValueError("Workbook sheet relationship missing")

    rel_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
    target = None
    for rel in rel_root:
        if rel.attrib.get("Id") == rel_id:
            target = rel.attrib.get("Target")
            break
    if not target:
        raise ValueError("Workbook sheet target missing")
    if target.startswith("xl/"):
        return target
    return f"xl/{target.lstrip('/')}"


def read_xlsx_row(row_node: ET.Element, shared_strings: list[str]) -> tuple[object, ...]:
    values: list[object] = []
    last_col = 0
    for cell in row_node.findall("a:c", XLSX_NS):
        ref = cell.attrib.get("r", "A1")
        col_idx = column_index(ref)
        while last_col + 1 < col_idx:
            values.append(None)
            last_col += 1
        values.append(read_xlsx_cell(cell, shared_strings))
        last_col = col_idx
    return tuple(values)


def read_xlsx_cell(cell_node: ET.Element, shared_strings: list[str]) -> object | None:
    cell_type = cell_node.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell_node.findall(".//a:t", XLSX_NS))

    value_node = cell_node.find("a:v", XLSX_NS)
    if value_node is None:
        return None

    value = value_node.text
    if cell_type == "s" and value is not None:
        return shared_strings[int(value)]
    if cell_type == "b":
        return value == "1"
    return value


def column_index(cell_ref: str) -> int:
    letters = ""
    for char in cell_ref:
        if char.isalpha():
            letters += char
        else:
            break

    result = 0
    for char in letters.upper():
        result = (result * 26) + (ord(char) - 64)
    return result


def read_client_codes_from_file(path: str) -> tuple[dict[str, int], list[str], dict[str, list[str]]]:
    rows = iter_excel_rows(path)
    header_row, client_code_idx, header_row_number = find_header_row(rows, "Client Code")
    reference_idx = find_column_index(header_row, "Reference")
    if reference_idx is None:
        raise ValueError(f"Reference column not found in {path}")

    counts: dict[str, int] = {}
    unknown_codes: list[str] = []
    references_by_code: dict[str, list[str]] = {}
    for row in rows:
        raw_value = cell_value(row, client_code_idx)
        if raw_value in (None, ""):
            continue
        client_code = str(raw_value).strip().upper()
        if not client_code or normalize_header(client_code) == "clientcode":
            continue
        counts[client_code] = counts.get(client_code, 0) + 1
        reference = normalize_reference(cell_value(row, reference_idx))
        if reference:
            references = references_by_code.setdefault(client_code, [])
            if reference not in references:
                references.append(reference)
        if client_code not in CLIENT_CODE_MAP and client_code not in unknown_codes:
            unknown_codes.append(client_code)

    if header_row_number > 1:
        print(f"Found header row at line {header_row_number} in {path}")

    return counts, unknown_codes, references_by_code


def read_handover_rows_from_file(path: str) -> tuple[list[HandoverRow], list[str]]:
    rows = iter_excel_rows(path)
    header_row, client_code_idx, header_row_number = find_header_row(rows, "Client Code")
    reference_idx = find_column_index(header_row, "Reference")
    if reference_idx is None:
        raise ValueError(f"Reference column not found in {path}")

    headers = [str(value).strip() if value not in (None, "") else "" for value in header_row]
    items: list[HandoverRow] = []
    unknown_codes: list[str] = []

    for offset, row in enumerate(rows, start=1):
        row_number = header_row_number + offset
        raw_client_code = cell_value(row, client_code_idx)
        if raw_client_code in (None, ""):
            continue

        client_code = str(raw_client_code).strip().upper()
        if not client_code or normalize_header(client_code) == "clientcode":
            continue

        client_id = CLIENT_CODE_MAP.get(client_code)
        if not client_id:
            if client_code not in unknown_codes:
                unknown_codes.append(client_code)
            continue

        values_by_header: dict[str, object] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            values_by_header[header] = cell_value(row, idx)

        items.append(
            HandoverRow(
                source_path=path,
                row_number=row_number,
                headers=headers,
                row_values=tuple(row),
                values_by_header=values_by_header,
                client_code=client_code,
                client_id=client_id,
                reference=normalize_reference(cell_value(row, reference_idx)),
            )
        )

    return items, unknown_codes


def read_handover_rows(paths: list[str]) -> tuple[list[HandoverRow], list[str]]:
    all_rows: list[HandoverRow] = []
    unknown_codes: list[str] = []
    for path in paths:
        file_rows, file_unknown_codes = read_handover_rows_from_file(path)
        all_rows.extend(file_rows)
        for code in file_unknown_codes:
            if code not in unknown_codes:
                unknown_codes.append(code)
    return all_rows, unknown_codes


def find_latest_fileref(matters: list[dict], prefix: str) -> tuple[str | None, str]:
    best_number = 0
    best_ref = None
    best_width = 4
    pattern = re.compile(rf"^{re.escape(prefix)}/(\d+)$", re.IGNORECASE)

    for matter in matters:
        file_ref = str(matter.get("fileref") or "").strip()
        match = pattern.match(file_ref)
        if not match:
            continue
        digits = match.group(1)
        number = int(digits)
        if number >= best_number:
            best_number = number
            best_ref = file_ref
            best_width = max(4, len(digits))

    next_ref = f"{prefix}/{best_number + 1:0{best_width}d}"
    return best_ref, next_ref


def next_ref_sequence_start(next_ref: str) -> tuple[int, int]:
    match = re.search(r"/(\d+)$", next_ref)
    if not match:
        return 1, 4
    digits = match.group(1)
    return int(digits), max(4, len(digits))


def get_row_value(row: HandoverRow, header_name: str) -> object | None:
    expected_key = normalize_header(header_name)
    for key, value in row.values_by_header.items():
        if normalize_header(key) == expected_key:
            return value
    return None


def build_debtor_name(row: HandoverRow) -> str | None:
    surname = normalize_cell_value(get_row_value(row, "Debtor Surname"))
    first_name = normalize_cell_value(get_row_value(row, "Debtor First Name"))
    name_parts = [str(value) for value in (surname, first_name) if value not in (None, "")]
    if name_parts:
        return " ".join(name_parts)
    return None


def build_description(row: HandoverRow) -> str:
    debtor_name = build_debtor_name(row)
    if debtor_name:
        return debtor_name
    matter_description = normalize_cell_value(get_row_value(row, "Matter Description"))
    if matter_description:
        return str(matter_description)
    if row.reference:
        return row.reference
    return "New Matter"


def build_party_prefix(row: HandoverRow) -> str:
    surname = normalize_cell_value(get_row_value(row, "Debtor Surname"))
    base_text = str(surname or build_description(row) or "PTY")
    letters = "".join(ch for ch in base_text.upper() if ch.isalpha())
    prefix = (letters[:3] or "PTY").ljust(3, "X")
    suffix_source = digits_only(row.reference) or digits_only(get_row_value(row, "ID Number")) or str(row.row_number)
    suffix = suffix_source[-3:].rjust(3, "0")
    return f"{prefix}{suffix}"


def build_matter_create_payload(
    row: HandoverRow,
    file_ref: str,
    date_ctx: DateContext,
    logged_in_employee_id: str,
) -> dict[str, object]:
    now = dt.datetime.now()
    payload: dict[str, object] = {}
    payload.update(CREATE_DEFAULTS)
    payload["clientid"] = row.client_id
    payload["fileref"] = file_ref
    payload["description"] = build_description(row)
    payload["dateinstructed"] = encode_legalsuite_date(dt.datetime.strptime(date_ctx.date_str, "%Y%m%d"))
    payload["updatedbydate"] = encode_legalsuite_date(now)
    payload["updatedbytime"] = encode_legalsuite_time(now)
    payload["partymatterprefix"] = row.client_code
    payload["internalcomment"] = (
        f"Imported from handover file on "
        f"{dt.datetime.strptime(date_ctx.date_str, '%Y%m%d').strftime('%d %B %Y')}"
    )
    payload["loggedinemployeeid"] = logged_in_employee_id

    for field_name, header_name in CREATE_FIELD_COLUMN_MAP.items():
        raw_value = get_row_value(row, header_name)
        if raw_value in (None, ""):
            continue
        value = normalize_reference(raw_value) if field_name == "theirref" else normalize_cell_value(raw_value)
        if value not in (None, ""):
            payload[field_name] = value

    if row.reference:
        payload["theirref"] = row.reference

    claim_amount = normalize_money(get_row_value(row, "Claim Amount"))
    if claim_amount is not None:
        payload["claimamount"] = claim_amount
        payload["debtorsbalance"] = claim_amount
        payload["debtorsopeningbalance"] = claim_amount
        payload["interestonamount"] = claim_amount
        payload["debtorscapitalbalance"] = claim_amount

    return {key: value for key, value in payload.items() if value not in (None, "")}


def build_matter_description_update_payload(
    create_payload: dict[str, object],
    logged_in_employee_id: str,
) -> dict[str, object]:
    payload = {
        "clientid": create_payload.get("clientid"),
        "archivestatusdescription": "Live",
        "loggedinemployeeid": logged_in_employee_id,
        "mattertypeid": create_payload.get("mattertypeid", CREATE_DEFAULTS["mattertypeid"]),
        "clientfeesheetid": create_payload.get("clientfeesheetid", CREATE_DEFAULTS["clientfeesheetid"]),
        "docgenid": create_payload.get("docgenid", CREATE_DEFAULTS["docgenid"]),
        "archiveflag": 0,
        "archivestatus": 0,
        "archiveno": 0,
        "description": create_payload.get("description"),
        "claimamount": create_payload.get("claimamount"),
        "debtorsbalance": create_payload.get("debtorsbalance"),
        "debtorsopeningbalance": create_payload.get("debtorsopeningbalance"),
        "interestonamount": create_payload.get("interestonamount"),
        "debtorscapitalbalance": create_payload.get("debtorscapitalbalance"),
    }
    return {key: value for key, value in payload.items() if value not in (None, "")}


def add_payload_value(payload: dict[str, object], field_name: str, value: object) -> None:
    normalized = normalize_cell_value(value)
    if normalized not in (None, ""):
        payload[field_name] = normalized


def build_party_create_payload(
    row: HandoverRow,
    date_ctx: DateContext,
    logged_in_employee_id: str,
) -> dict[str, object]:
    now = dt.datetime.now()
    name = build_debtor_name(row) or build_description(row)
    imported_date = dt.datetime.strptime(date_ctx.date_str, "%Y%m%d")
    notes = f"Imported on {imported_date.strftime('%d %B %Y')}"

    payload: dict[str, object] = {}
    payload.update(PARTY_DEFAULTS)
    payload["party[name]"] = name
    payload["party[matterprefix]"] = build_party_prefix(row)
    payload["party[updatedbydate]"] = encode_legalsuite_date(now)
    payload["party[updatedbytime]"] = encode_legalsuite_time(now)
    payload["party[createdid]"] = logged_in_employee_id
    payload["party[notes]"] = notes

    identity_number = digits_only(get_row_value(row, "ID Number"))
    if identity_number:
        payload["party[identitynumber]"] = identity_number

    add_payload_value(payload, "party[parlang][salutation]", get_row_value(row, "Debtor Title"))
    add_payload_value(payload, "party[parlang][physicalline1]", get_row_value(row, "Physical Address Line 1"))
    add_payload_value(payload, "party[parlang][physicalline2]", get_row_value(row, "Physical Address Line 2"))
    add_payload_value(payload, "party[parlang][physicalline3]", get_row_value(row, "Physical Address Line 3"))
    add_payload_value(payload, "party[parlang][physicalcode]", get_row_value(row, "Physical Postal Code"))
    add_payload_value(payload, "party[parlang][postalline1]", get_row_value(row, "Postal Address Line 1"))
    add_payload_value(payload, "party[parlang][postalline2]", get_row_value(row, "Postal Address Line 2"))
    add_payload_value(payload, "party[parlang][postalline3]", get_row_value(row, "Postal Address Line 3"))
    add_payload_value(payload, "party[parlang][postalcode]", get_row_value(row, "Postal Code"))
    add_payload_value(payload, "party[parlang][homenumber]", get_row_value(row, "Telephone (Home)"))
    add_payload_value(payload, "party[parlang][worknumber]", get_row_value(row, "Telephone (Work)"))
    add_payload_value(payload, "party[parlang][cellnumber]", get_row_value(row, "Cell Phone"))
    add_payload_value(payload, "party[parlang][emailaddress]", get_row_value(row, "DefendantEmail"))
    return {key: value for key, value in payload.items() if value not in (None, "")}


def build_matparty_create_payload(matterid: int | str, partyid: int | str) -> dict[str, object]:
    return {
        "matterid": matterid,
        "partyid": partyid,
        **MATPARTY_DEFAULTS,
    }


def header_has_date_semantics(header_name: str) -> bool:
    return "date" in normalize_header(header_name)


def build_desktop_extrascreen_payloads(row: HandoverRow) -> list[tuple[str, dict[str, object]]]:
    payloads: list[tuple[str, dict[str, object]]] = []
    current_screen_label: str | None = None
    current_payload: dict[str, object] | None = None

    def flush_current() -> None:
        nonlocal current_screen_label, current_payload
        if current_screen_label and current_payload and len(current_payload) > 1:
            payloads.append((current_screen_label, current_payload))
        current_screen_label = None
        current_payload = None

    for header_name, raw_value in zip(row.headers, row.row_values):
        if not header_name:
            continue

        screen_match = re.fullmatch(r"DesktopExtraScreenID(\d+)", header_name.strip(), re.IGNORECASE)
        if screen_match:
            flush_current()
            screen_id = normalize_reference(raw_value)
            if screen_id:
                current_screen_label = f"DesktopExtraScreenID{screen_match.group(1)}"
                current_payload = {"docscreenid": screen_id}
            continue

        if current_payload is None:
            continue

        field_match = re.match(r"Desktop Extra Field\s+(\d+)\b", header_name.strip(), re.IGNORECASE)
        if not field_match:
            continue

        field_number = int(field_match.group(1))
        field_name = f"field{field_number}"
        if raw_value in (None, ""):
            continue

        if header_has_date_semantics(header_name):
            encoded = encode_legalsuite_date(raw_value)
            if encoded is None:
                continue
            current_payload[field_name] = encoded
        else:
            normalized = normalize_cell_value(raw_value)
            if normalized in (None, ""):
                continue
            current_payload[field_name] = normalized

    flush_current()
    return payloads


def extract_created_recordid(created_response: dict | str) -> str:
    if not isinstance(created_response, dict):
        raise ValueError(f"Unexpected create response: {created_response}")
    data = created_response.get("data")
    if not isinstance(data, list) or not data:
        raise ValueError(f"Create response has no data row: {created_response}")
    recordid = data[0].get("recordid")
    if not recordid:
        raise ValueError(f"Create response has no recordid: {created_response}")
    return str(recordid)


def extract_matter_create_recordid(created_response: dict | str) -> tuple[str, bool]:
    try:
        return extract_created_recordid(created_response), False
    except ValueError:
        pass

    if not isinstance(created_response, dict):
        raise ValueError(f"Unexpected create response: {created_response}")

    errors = str(created_response.get("errors") or "")
    if "addMatterParties()" not in errors or "insert into [MatParty]" not in errors:
        raise ValueError(f"Create response has no data row: {created_response}")

    match = re.search(r"values\s*\(\s*1\s*,\s*(\d+)\s*,", errors, re.IGNORECASE)
    if not match:
        raise ValueError(f"Could not recover matter recordid from create response: {created_response}")

    return match.group(1), True


def extract_fetched_row(fetched_response: dict | str) -> dict:
    if not isinstance(fetched_response, dict):
        raise ValueError(f"Unexpected fetch response: {fetched_response}")
    data = fetched_response.get("data")
    if not isinstance(data, list) or not data:
        raise ValueError(f"Fetch response has no data row: {fetched_response}")
    return data[0]


def fetch_matter_row(client: LegalSuiteLookupClient, recordid: int | str) -> dict:
    fetched = client.get_matter_by_recordid(recordid)
    return extract_fetched_row(fetched)


def normalize_compare_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, (int, float, Decimal)):
        try:
            return format(Decimal(str(value)).normalize(), "f").rstrip("0").rstrip(".") or "0"
        except (InvalidOperation, ValueError):
            return str(value).strip()

    text = str(value).strip()
    if not text:
        return ""

    numeric_text = text.replace(",", "")
    if re.fullmatch(r"-?\d+(?:\.\d+)?", numeric_text):
        try:
            return format(Decimal(numeric_text).normalize(), "f").rstrip("0").rstrip(".") or "0"
        except InvalidOperation:
            pass

    return text


MATTER_VERIFY_IGNORE_FIELDS = {
    "formattedupdatedbydate",
    "formattedupdatedbytime",
    "updatedbydate",
    "updatedbytime",
    "updatedbydatetime",
    "loggedinemployeeid",
}


def find_changed_fields(before: dict, after: dict, field_names: list[str]) -> list[tuple[str, object, object]]:
    changes: list[tuple[str, object, object]] = []
    for field_name in field_names:
        before_value = before.get(field_name)
        after_value = after.get(field_name)
        if normalize_compare_value(before_value) != normalize_compare_value(after_value):
            changes.append((field_name, before_value, after_value))
    return changes


def compare_extrascreen_payload_to_row(payload: dict[str, object], fetched_row: dict) -> list[tuple[str, object, object]]:
    field_names = sorted(key for key in payload if key.startswith("field"))
    mismatches: list[tuple[str, object, object]] = []
    for field_name in field_names:
        sent_value = payload.get(field_name)
        fetched_value = fetched_row.get(field_name)
        if normalize_compare_value(sent_value) != normalize_compare_value(fetched_value):
            mismatches.append((field_name, sent_value, fetched_value))
    return mismatches


def describe_extrascreen_field_values(row: dict, field_names: list[str]) -> str:
    parts: list[str] = []
    for field_name in field_names:
        value = row.get(field_name)
        if normalize_compare_value(value):
            parts.append(f"{field_name}={value}")
    return ", ".join(parts) if parts else "<all blank>"


def compare_matter_payload_to_row(
    payload: dict[str, object],
    fetched_row: dict,
    field_names: list[str] | None = None,
    ignore_fields: set[str] | None = None,
) -> list[tuple[str, object, object]]:
    ignored = set(MATTER_VERIFY_IGNORE_FIELDS)
    if ignore_fields:
        ignored.update(ignore_fields)

    names = field_names or sorted(payload)
    mismatches: list[tuple[str, object, object]] = []
    for field_name in names:
        if field_name in ignored:
            continue
        sent_value = payload.get(field_name)
        fetched_value = fetched_row.get(field_name)
        if normalize_compare_value(sent_value) != normalize_compare_value(fetched_value):
            mismatches.append((field_name, sent_value, fetched_value))
    return mismatches


def print_matter_verification(
    label: str,
    payload: dict[str, object],
    fetched_row: dict,
    field_names: list[str] | None = None,
    ignore_fields: set[str] | None = None,
) -> None:
    mismatches = compare_matter_payload_to_row(payload, fetched_row, field_names=field_names, ignore_fields=ignore_fields)
    names = [name for name in (field_names or sorted(payload)) if name not in MATTER_VERIFY_IGNORE_FIELDS]
    if mismatches:
        mismatch_names = ", ".join(field_name for field_name, _, _ in mismatches)
        print(f"  {label} verification mismatches: {mismatch_names}")
        for field_name, sent_value, fetched_value in mismatches:
            print(f"    {field_name}: sent={sent_value!r} fetched={fetched_value!r}")
    elif names:
        print(f"  {label} verified: {', '.join(names)}")
    else:
        print(f"  {label} had no fields to verify.")


def find_existing_matter_for_row(client: LegalSuiteLookupClient, row: HandoverRow, file_ref: str) -> dict | None:
    fileref_matches = client.get_matters_by_fileref(file_ref)
    if fileref_matches:
        return fileref_matches[0]

    if row.reference:
        reference_matches = client.get_matters_by_clientid_and_reference(row.client_id, row.reference)
        if reference_matches:
            return reference_matches[0]

    return None


def update_handover_row_desktop_extrascreens(
    client: LegalSuiteLookupClient,
    row: HandoverRow,
    matter_recordid: int | str,
    dry_run: bool,
) -> None:
    payloads = build_desktop_extrascreen_payloads(row)
    if not payloads:
        print("  No desktop extrascreen data to update.")
        return

    for screen_label, payload in payloads:
        docscreenid = payload["docscreenid"]
        field_names = sorted(key for key in payload if key.startswith("field"))
        print(f"  Processing {screen_label}: docscreenid={docscreenid} | fields={len(field_names)}")
        if dry_run:
            print(f"  Dry-run: would update {screen_label}.")
            continue

        update_payload = {"matterid": matter_recordid, **payload}
        client.update_matter_extrascreen(update_payload)
        print(f"  Updated {screen_label}")
        fetched_rows = client.get_matter_extrascreen(matter_recordid, docscreenid)
        if not fetched_rows:
            print(f"  No data returned for {screen_label} after update.")
            continue

        fetched_row = fetched_rows[0]
        mismatches = compare_extrascreen_payload_to_row(payload, fetched_row)
        if mismatches:
            mismatch_names = ", ".join(field_name for field_name, _, _ in mismatches)
            print(f"  {screen_label} mismatched fields after update: {mismatch_names}")
            for field_name, sent_value, fetched_value in mismatches:
                print(f"    {field_name}: sent={sent_value!r} fetched={fetched_value!r}")
            print(
                "  Returned extrascreen values: "
                f"{describe_extrascreen_field_values(fetched_row, sorted(key for key in payload if key.startswith('field')))}"
            )
        else:
            if field_names:
                print(f"  {screen_label} fields verified after update: {', '.join(field_names)}")
                print(f"  Returned extrascreen values: {describe_extrascreen_field_values(fetched_row, field_names)}")
            else:
                print(f"  {screen_label} had no field values to verify.")


def create_and_update_handover_matters(
    rows: list[HandoverRow],
    next_refs_by_code: dict[str, str],
    client: LegalSuiteLookupClient,
    date_ctx: DateContext,
    logged_in_employee_id: str,
    create_matters: bool,
    create_limit: int | None,
) -> list[HandoverCreatedMatter]:
    next_numbers: dict[str, tuple[int, int]] = {
        code: next_ref_sequence_start(next_ref)
        for code, next_ref in next_refs_by_code.items()
    }

    processed_count = 0
    created_matters: list[HandoverCreatedMatter] = []
    print("\nMatter create/update:")
    for row in rows:
        if create_limit is not None and processed_count >= create_limit:
            print(f"Create limit reached: {create_limit}")
            break
        processed_count += 1

        next_number, width = next_numbers[row.client_code]
        file_ref = f"{row.client_code}/{next_number:0{width}d}"
        next_numbers[row.client_code] = (next_number + 1, width)
        payload = build_matter_create_payload(row, file_ref, date_ctx, logged_in_employee_id)
        party_payload = build_party_create_payload(row, date_ctx, logged_in_employee_id)

        print(
            f"- {row.client_code} row {row.row_number} | reference {row.reference or ''} | "
            f"new fileref {file_ref}"
        )

        print("  Checking for existing matter...")
        existing_matter = find_existing_matter_for_row(client, row, file_ref)
        if existing_matter:
            existing_recordid = str(existing_matter.get("recordid"))
            print(
                "  Skipped: matter already exists "
                f"recordid {existing_recordid} "
                f"fileref {existing_matter.get('fileref')} "
                f"theirref {existing_matter.get('theirref')}"
            )
            update_handover_row_desktop_extrascreens(client, row, existing_recordid, dry_run=not create_matters)
            continue

        if not create_matters:
            print("  Dry-run: would create matter.")
            print("  Dry-run: would update matter description.")
            print("  Dry-run: would create or reuse party.")
            print("  Dry-run: would create MatParty link if missing.")
            update_handover_row_desktop_extrascreens(client, row, "<created-matter-recordid>", dry_run=True)
            continue

        print("  Creating matter...")
        created = client.create_matter(payload)
        recordid, recovered_from_error = extract_matter_create_recordid(created)
        if recovered_from_error:
            print(f"  Matter created; recovered recordid from create error: {recordid}")
        else:
            print(f"  Created matter recordid: {recordid}")
        created_matters.append(
            HandoverCreatedMatter(
                file_ref=file_ref,
                their_reference=str(payload.get("theirref") or row.reference or ""),
                description=str(payload.get("description") or ""),
            )
        )
        matter_after_create = fetch_matter_row(client, recordid)
        print(f"  Matter claimamount after create: {matter_after_create.get('claimamount')}")

        print("  Updating matter description...")
        update_payload = build_matter_description_update_payload(payload, logged_in_employee_id)
        client.update_matter(recordid, update_payload)
        print("  Updated matter description")
        matter_after_update = fetch_matter_row(client, recordid)
        print_matter_verification(
            "Matter description update",
            update_payload,
            matter_after_update,
            field_names=sorted(update_payload),
        )
        print(f"  Matter claimamount after description update: {matter_after_update.get('claimamount')}")
        tracked_fields = sorted(
            set(payload) | {"claimamount", "debtorsbalance", "debtorsopeningbalance", "interestonamount", "debtorscapitalbalance"}
        )
        changed_after_update = find_changed_fields(matter_after_create, matter_after_update, tracked_fields)
        if changed_after_update:
            changed_names = ", ".join(field_name for field_name, _, _ in changed_after_update)
            print(f"  Fields changed after description update: {changed_names}")
        else:
            print("  No tracked matter fields changed after description update.")

        print("  Checking for existing party...")
        identity_number = digits_only(get_row_value(row, "ID Number"))
        existing_party = None
        if identity_number:
            matches = client.get_party_by_identitynumber(identity_number)
            if matches:
                existing_party = matches[0]

        if existing_party:
            partyid = str(existing_party["recordid"])
            print(f"  Reusing existing partyid: {partyid}")
        else:
            print("  Creating party...")
            created_party = client.create_party(party_payload)
            partyid = extract_created_recordid(created_party)
            print(f"  Created partyid: {partyid}")

        print("  Checking MatParty link...")
        existing_matparty = client.get_matparty_by_matter_and_party(recordid, partyid)
        if existing_matparty:
            print(f"  MatParty link already exists: recordid {existing_matparty[0].get('recordid')}")
        else:
            print("  Creating MatParty link...")
            matparty_payload = build_matparty_create_payload(recordid, partyid)
            created_matparty = client.create_matparty(matparty_payload)
            matparty_recordid = extract_created_recordid(created_matparty)
            print(f"  Linked matparty recordid: {matparty_recordid}")

        matter_after_matparty = fetch_matter_row(client, recordid)
        print(f"  Matter claimamount after MatParty step: {matter_after_matparty.get('claimamount')}")
        changed_after_matparty = find_changed_fields(matter_after_update, matter_after_matparty, tracked_fields)
        if changed_after_matparty:
            changed_names = ", ".join(field_name for field_name, _, _ in changed_after_matparty)
            print(f"  Fields changed after MatParty step: {changed_names}")
        else:
            print("  No tracked matter fields changed after MatParty step.")

        update_handover_row_desktop_extrascreens(client, row, recordid, dry_run=False)

    return created_matters


def process_handover_files(
    paths: list[str],
    api_base: str,
    api_key: str,
    date_ctx: DateContext,
    create_matters: bool,
    create_dry_run: bool,
    create_limit: int | None,
    logged_in_employee_id: str,
) -> list[HandoverCreatedMatter]:
    code_counts: dict[str, int] = {}
    unknown_codes: list[str] = []
    code_references: dict[str, list[str]] = {}
    next_refs_by_code: dict[str, str] = {}

    for path in paths:
        file_counts, file_unknown_codes, file_references = read_client_codes_from_file(path)
        total_rows = sum(file_counts.values())
        print(f"Processed {path}: {total_rows} row(s), {len(file_counts)} unique client code(s)")
        for code, count in file_counts.items():
            code_counts[code] = code_counts.get(code, 0) + count
            refs = code_references.setdefault(code, [])
            for reference in file_references.get(code, []):
                if reference not in refs:
                    refs.append(reference)
        for code in file_unknown_codes:
            if code not in unknown_codes:
                unknown_codes.append(code)

    if not code_counts:
        print("No Client Code values found in the downloaded handover files.")
        return []

    if unknown_codes:
        print("Unknown client codes in Excel:", ", ".join(sorted(unknown_codes)))

    client = LegalSuiteLookupClient(api_base, api_key)
    print("\nLegalSuite latest matter lookup:")
    for client_code in sorted(code_counts):
        client_id = CLIENT_CODE_MAP.get(client_code)
        if not client_id:
            print(f"- {client_code}: no client-id mapping found")
            continue

        matters = client.get_matters_by_clientid_and_prefix(client_id, client_code)
        latest_ref, next_ref = find_latest_fileref(matters, client_code)
        next_refs_by_code[client_code] = next_ref
        if latest_ref:
            print(
                f"- {client_code} -> clientid {client_id} | rows {code_counts[client_code]} | "
                f"latest {latest_ref} | next {next_ref}"
            )
        else:
            print(
                f"- {client_code} -> clientid {client_id} | rows {code_counts[client_code]} | "
                f"latest none | next {next_ref}"
            )
        references = code_references.get(client_code, [])
        if references:
            print(f"  Reference values: {', '.join(references)}")

    if create_matters or create_dry_run:
        rows, row_unknown_codes = read_handover_rows(paths)
        for code in row_unknown_codes:
            if code not in unknown_codes:
                unknown_codes.append(code)
        if unknown_codes:
            print("Rows skipped for unknown client codes:", ", ".join(sorted(unknown_codes)))
        return create_and_update_handover_matters(
            rows=rows,
            next_refs_by_code=next_refs_by_code,
            client=client,
            date_ctx=date_ctx,
            logged_in_employee_id=logged_in_employee_id,
            create_matters=create_matters,
            create_limit=create_limit,
        )

    return []


def build_handover_report_preview_entries(
    rows: list[HandoverRow],
    create_limit: int | None,
) -> list[HandoverCreatedMatter]:
    preview_entries: list[HandoverCreatedMatter] = []
    for row in rows:
        if create_limit is not None and len(preview_entries) >= create_limit:
            break
        preview_entries.append(
            HandoverCreatedMatter(
                file_ref=f"{row.client_code}/TEST-{row.row_number:04d}",
                their_reference=str(row.reference or ""),
                description=build_description(row),
            )
        )
    return preview_entries


class Cleaner:
    def __init__(self, handover_prefixes: list[str]) -> None:
        self._rules = {"accountnumber": self._digits_only}
        self._handover_prefixes = handover_prefixes

    def clean_downloads(
        self,
        download_dir: str,
        cleaned_dir: str,
        report_lines: list[str],
        source_paths: list[str] | None = None,
    ) -> None:
        if load_workbook is None or Workbook is None:
            report_lines.append("Cleaning skipped: openpyxl is not installed.")
            print("Cleaning skipped: openpyxl is not installed.", file=sys.stderr)
            return

        cleaned_files = 0
        converted_files = 0
        skipped_files = 0
        failed_files = 0
        total_cells = 0
        matched_columns = 0
        blanked_columns = 0
        headers_removed = 0
        copied_cells = 0
        cleaned_dir_abs = os.path.abspath(cleaned_dir)

        for source_path in self._iter_source_paths(download_dir, cleaned_dir_abs, source_paths):
            name = os.path.basename(source_path)
            lowered = name.lower()
            if not (lowered.endswith(".xlsx") or lowered.endswith(".csv")):
                continue
            if name.startswith("~$"):
                skipped_files += 1
                continue

            rel_path = os.path.relpath(source_path, download_dir)
            if lowered.endswith(".csv"):
                rel_path = os.path.splitext(rel_path)[0] + ".xlsx"
            destination_path = os.path.join(cleaned_dir, rel_path)

            blank_headers = set()
            drop_header = True
            if fnmatch.fnmatch(name, CLAIM_AMOUNT_PATTERN):
                blank_headers.add("matter")
                drop_header = False

            rel_norm = rel_path.replace(os.sep, "/")
            copy_reference = any(rel_norm.startswith(prefix) for prefix in self._handover_prefixes)
            if copy_reference:
                drop_header = False

            try:
                if lowered.endswith(".csv"):
                    raw_copy_path = os.path.splitext(source_path)[0] + ".xlsx"
                    cleaned, matched, blanked, removed, copied = self._convert_csv_to_excel(
                        source_path,
                        destination_path,
                        blank_headers,
                        copy_reference,
                        drop_header,
                        raw_copy_path,
                    )
                    converted_files += 1
                    action = "Converted"
                    report_lines.append(f"Saved CSV Excel copy: {raw_copy_path}")
                else:
                    cleaned, matched, blanked, removed, copied = self._clean_excel_file(
                        source_path, destination_path, blank_headers, copy_reference, drop_header
                    )
                    cleaned_files += 1
                    action = "Cleaned"

                total_cells += cleaned
                matched_columns += matched
                blanked_columns += blanked
                headers_removed += removed
                copied_cells += copied
                report_lines.append(
                    f"{action}: {source_path} -> {destination_path} (cells updated: {cleaned})"
                )
            except Exception as exc:
                failed_files += 1
                report_lines.append(f"Clean failed: {source_path} ({exc})")
                print(f"Clean failed: {source_path}: {exc}", file=sys.stderr)

        report_lines.append(
            "Clean summary: cleaned_files={cleaned}, converted_files={converted}, "
            "skipped_files={skipped}, failed_files={failed}, "
            "columns_matched={columns}, blanked_columns={blanked}, "
            "headers_removed={headers}, copied_cells={copied}, cells_updated={cells}".format(
                cleaned=cleaned_files,
                converted=converted_files,
                skipped=skipped_files,
                failed=failed_files,
                columns=matched_columns,
                blanked=blanked_columns,
                headers=headers_removed,
                copied=copied_cells,
                cells=total_cells,
            )
        )

    @staticmethod
    def _iter_source_paths(
        download_dir: str,
        cleaned_dir_abs: str,
        source_paths: list[str] | None,
    ) -> list[str]:
        if source_paths is None:
            paths: list[str] = []
            for root, _, files in os.walk(download_dir):
                root_abs = os.path.abspath(root)
                if root_abs.startswith(cleaned_dir_abs):
                    continue
                for name in files:
                    paths.append(os.path.join(root, name))
            return paths

        unique_paths: list[str] = []
        seen: set[str] = set()
        for path in source_paths:
            path_abs = os.path.abspath(path)
            if path_abs.startswith(cleaned_dir_abs):
                continue
            if not os.path.exists(path):
                continue
            if path_abs in seen:
                continue
            seen.add(path_abs)
            unique_paths.append(path)
        return unique_paths

    def _clean_excel_file(
        self,
        source_path: str,
        destination_path: str,
        blank_headers: set[str],
        copy_reference: bool,
        drop_header: bool,
    ) -> tuple[int, int, int, int, int]:
        workbook = load_workbook(source_path)
        cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells = self._apply_cleaning(
            workbook, blank_headers, drop_header, copy_reference
        )
        os.makedirs(os.path.dirname(destination_path) or ".", exist_ok=True)
        workbook.save(destination_path)
        return cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells

    def _convert_csv_to_excel(
        self,
        source_path: str,
        destination_path: str,
        blank_headers: set[str],
        copy_reference: bool,
        drop_header: bool,
        raw_copy_path: str | None = None,
    ) -> tuple[int, int, int, int, int]:
        workbook = Workbook()
        worksheet = workbook.active
        with open(source_path, newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle)
            for row in reader:
                worksheet.append(row)
        if raw_copy_path:
            os.makedirs(os.path.dirname(raw_copy_path) or ".", exist_ok=True)
            workbook.save(raw_copy_path)
        cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells = self._apply_cleaning(
            workbook, blank_headers, drop_header, copy_reference
        )
        os.makedirs(os.path.dirname(destination_path) or ".", exist_ok=True)
        workbook.save(destination_path)
        return cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells

    def _apply_cleaning(
        self,
        workbook,
        blank_headers: set[str],
        drop_header: bool,
        copy_reference: bool,
    ) -> tuple[int, int, int, int, int]:
        cleaned_cells = 0
        matched_columns = 0
        blanked_columns = 0
        headers_removed = 0
        copied_cells = 0

        for worksheet in workbook.worksheets:
            header_row = list(worksheet.iter_rows(min_row=1, max_row=1))
            if not header_row:
                continue
            header_cells = header_row[0]
            column_rules: dict[int, Callable[[object], object]] = {}
            blank_columns: set[int] = set()
            account_col = None
            reference_col = None

            for idx, cell in enumerate(header_cells, start=1):
                key = self._normalize_header(cell.value)
                if key in self._rules:
                    column_rules[idx] = self._rules[key]
                if key in blank_headers:
                    blank_columns.add(idx)
                if key == "accountnumber":
                    account_col = idx
                if key == "reference":
                    reference_col = idx

            matched_columns += len(column_rules)
            blanked_columns += len(blank_columns)

            if column_rules or blank_columns or (copy_reference and account_col and reference_col):
                for row in worksheet.iter_rows(min_row=2):
                    if copy_reference and account_col and reference_col:
                        account_cell = row[account_col - 1]
                        reference_cell = row[reference_col - 1]
                        if account_cell.value != reference_cell.value:
                            account_cell.value = reference_cell.value
                            copied_cells += 1
                    for col_idx, rule in column_rules.items():
                        cell = row[col_idx - 1]
                        new_value = rule(cell.value)
                        if new_value != cell.value:
                            cell.value = new_value
                            cleaned_cells += 1
                    for col_idx in blank_columns:
                        cell = row[col_idx - 1]
                        if cell.value not in (None, ""):
                            cell.value = None
                            cleaned_cells += 1

            if drop_header and worksheet.max_row >= 1:
                worksheet.delete_rows(1)
                headers_removed += 1

        return cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells

    @staticmethod
    def _normalize_header(header: object) -> str:
        if header is None:
            return ""
        return "".join(ch for ch in str(header).strip().lower() if ch.isalnum())

    @staticmethod
    def _digits_only(value: object) -> object:
        if value is None:
            return None
        text = str(value)
        return "".join(ch for ch in text if ch.isdigit())


class App:
    def __init__(self, args: argparse.Namespace) -> None:
        self._args = args
        self._date_ctx = self._resolve_date(args.date, args.days_ago)
        self._targets = self._build_targets(self._date_ctx)
        self._cleaner = Cleaner(HANDOVER_PREFIXES)
        self._verification_recorder = VerificationWorkbookRecorder(
            self._args.verification_dir,
            [self._args.cleaned_dir, self._args.download_dir],
        )

    def run(self) -> int:
        report_lines = [
            f"Report date: {self._date_ctx.date_str}",
            f"Month folder: {self._date_ctx.month_year}",
        ]
        return_code = 0
        try:
            download_counts, source_paths = self._download_files(report_lines)
            if download_counts is None:
                return_code = 1
                return return_code

            if not self._args.skip_clean:
                print("Cleaning downloaded files...")
                self._cleaner.clean_downloads(
                    self._args.download_dir,
                    self._args.cleaned_dir,
                    report_lines,
                    source_paths=source_paths,
                )

            if not self._args.skip_handover:
                print("Processing handover files...")
                self._process_handover(report_lines)

            if self._args.update_extrascreen:
                print("Updating matter extra screens...")
                self._update_matter_extrascreens(report_lines)

            if self._args.update_claim_amount:
                print("Updating claim amounts...")
                self._update_claim_amounts(report_lines)

            if self._args.archive_closed:
                print("Reading closed files and calling LegalSuite...")
                self._archive_closed_matters(report_lines)

            if self._args.reopen_matters:
                print("Reading reopen files and calling LegalSuite...")
                self._reopen_matters(report_lines)

            report_lines.append(
                "Summary: downloaded={downloaded}, reused={reused}, missing_dirs={missing_dirs}, "
                "missing_files={missing_files}, failed_downloads={failed}".format(
                    downloaded=download_counts["downloaded"],
                    reused=download_counts["reused"],
                    missing_dirs=download_counts["missing_dirs"],
                    missing_files=download_counts["missing_files"],
                    failed=download_counts["failed_downloads"],
                )
            )
            return return_code
        finally:
            self._finalize_verification_workbooks(report_lines)
            self._write_report(report_lines)

    def _finalize_verification_workbooks(self, report_lines: list[str]) -> None:
        try:
            verification_paths = self._verification_recorder.finalize()
        except Exception as exc:
            report_lines.append(f"Verification workbook save failed: {exc}")
            print(f"Verification workbook save failed: {exc}", file=sys.stderr)
            return

        if not verification_paths:
            return

        report_lines.append(
            "Verification workbook summary: created={count}, directory={directory}".format(
                count=len(verification_paths),
                directory=os.path.abspath(self._args.verification_dir),
            )
        )
        print(
            "Verification workbook summary: created={count}, directory={directory}".format(
                count=len(verification_paths),
                directory=os.path.abspath(self._args.verification_dir),
            )
        )

    def _download_files(self, report_lines: list[str]) -> tuple[dict[str, int] | None, list[str]]:
        counts = {
            "downloaded": 0,
            "reused": 0,
            "missing_dirs": 0,
            "missing_files": 0,
            "failed_downloads": 0,
        }
        source_paths: list[str] = []

        if self._args.clean_only:
            report_lines.append("Download skipped: clean-only mode.")
            source_paths = self._resolve_existing_files(report_lines, counts)
            return counts, source_paths

        print("Connecting to FTP...")
        client = FTPClient(FTP_HOST, FTP_USER, FTP_PASS, self._args.timeout)
        try:
            client.connect()
        except ftplib.all_errors as exc:
            error_line = f"FTP connection failed: {exc}"
            print(error_line, file=sys.stderr)
            report_lines.append(error_line)
            return None, []

        try:
            print("Downloading files...")
            for remote_dir, filename in self._targets:
                resolved_dir = None
                resolved_name = None
                reason = "missing_dir"
                for candidate_dir in self._target_dirs(remote_dir):
                    resolved_name, reason = client.resolve_remote_file(candidate_dir, filename)
                    if resolved_name:
                        resolved_dir = candidate_dir
                        break

                if not resolved_name or not resolved_dir:
                    if reason == "missing_dir":
                        line = f"Missing directory: {remote_dir}"
                        counts["missing_dirs"] += 1
                    else:
                        line = f"Missing file: {remote_dir}/{filename}"
                        counts["missing_files"] += 1
                    print(line)
                    report_lines.append(line)
                    continue

                remote_path = f"{resolved_dir}/{resolved_name}"
                local_path = self._ensure_local_path(self._args.download_dir, resolved_dir, resolved_name)
                try:
                    client.download_file(remote_path, local_path)
                    line = f"Downloaded: {remote_path} -> {local_path}"
                    print(line)
                    report_lines.append(line)
                    counts["downloaded"] += 1
                    source_paths.append(local_path)
                except ftplib.all_errors as exc:
                    line = f"Download failed for {remote_path}: {exc}"
                    print(line, file=sys.stderr)
                    report_lines.append(line)
                    counts["failed_downloads"] += 1
        finally:
            client.close()

        return counts, source_paths

    def _resolve_existing_files(
        self,
        report_lines: list[str],
        counts: dict[str, int],
    ) -> list[str]:
        source_paths: list[str] = []
        for remote_dir, filename in self._targets:
            local_path = self._resolve_existing_local_file(remote_dir, filename)
            if local_path:
                line = f"Using existing file: {local_path}"
                print(line)
                report_lines.append(line)
                counts["reused"] += 1
                source_paths.append(local_path)
                continue

            line = f"Existing file not found: {remote_dir}/{filename}"
            print(line)
            report_lines.append(line)
            counts["missing_files"] += 1
        return source_paths

    def _resolve_existing_local_file(self, remote_dir: str, filename: str) -> str | None:
        for candidate_dir in self._target_dirs(remote_dir):
            local_dir = os.path.join(self._args.download_dir, *candidate_dir.split("/"))
            if not os.path.isdir(local_dir):
                continue

            try:
                names = sorted(
                    name for name in os.listdir(local_dir) if os.path.isfile(os.path.join(local_dir, name))
                )
            except OSError:
                continue

            if any(ch in filename for ch in ["*", "?", "["]):
                matches = [name for name in names if fnmatch.fnmatch(name, filename)]
                if matches:
                    return os.path.join(local_dir, matches[-1])
                continue

            if filename in names:
                return os.path.join(local_dir, filename)
        return None

    @staticmethod
    def _target_dirs(remote_dir: str) -> list[str]:
        dirs = [remote_dir]
        dirs.extend(REMOTE_DIR_ALIASES.get(remote_dir, []))
        seen: set[str] = set()
        unique_dirs: list[str] = []
        for item in dirs:
            if item in seen:
                continue
            seen.add(item)
            unique_dirs.append(item)
        return unique_dirs

    def _write_report(self, report_lines: list[str]) -> None:
        log_path = self._args.log_file
        if not log_path:
            log_path = os.path.join(self._args.download_dir, f"report_{self._date_ctx.date_str}.txt")
        os.makedirs(os.path.dirname(log_path) or ".", exist_ok=True)
        try:
            print("Writing report log...")
            with open(log_path, "w", encoding="utf-8") as handle:
                handle.write("\n".join(report_lines) + "\n")
        except OSError as exc:
            print(f"Failed to write log file {log_path}: {exc}", file=sys.stderr)

    def _resolve_handover_files(self, base_dir: str, label: str) -> list[str]:
        resolved_files: list[str] = []
        targets = [
            (
                "Debt Review handover",
                [
                    "SBSA/Debt Review/Debt_Review_Handover_APT_LWS",
                    "SBSA/Debt Review/Debt_Review_ Handover_APT_LWS",
                ],
                f"Standard_Bank_Panel_L_Handover_{self._date_ctx.date_str}_DR.xlsx",
            ),
            (
                "Panel L handover",
                [f"SBSA/Panel L/Handover_APT_LSW/{self._date_ctx.month_year}"],
                f"*_{self._date_ctx.date_str}.xlsx",
            ),
        ]

        for target_label, remote_dirs, filename_pattern in targets:
            found_path = None
            for remote_dir in remote_dirs:
                local_dir = os.path.join(base_dir, *remote_dir.split("/"))
                if not os.path.isdir(local_dir):
                    continue

                try:
                    names = sorted(
                        name for name in os.listdir(local_dir) if os.path.isfile(os.path.join(local_dir, name))
                    )
                except OSError:
                    continue

                if any(ch in filename_pattern for ch in ["*", "?", "["]):
                    matches = [name for name in names if fnmatch.fnmatch(name, filename_pattern)]
                    if matches:
                        found_path = os.path.join(local_dir, matches[-1])
                        break
                elif filename_pattern in names:
                    found_path = os.path.join(local_dir, filename_pattern)
                    break

            if found_path:
                resolved_files.append(found_path)
                print(f"Using {label} {target_label}: {found_path}")
            else:
                print(f"{label.capitalize()} {target_label} file not found for {self._date_ctx.date_str} in {base_dir}")

        return resolved_files

    def _process_handover(self, report_lines: list[str]) -> None:
        if self._args.skip_clean:
            working_files = self._resolve_handover_files(self._args.download_dir, "downloaded")
        else:
            working_files = self._resolve_handover_files(self._args.cleaned_dir, "cleaned")
            if not working_files:
                print("No cleaned handover files were available; checking downloaded files.")
                working_files = self._resolve_handover_files(self._args.download_dir, "downloaded")

        if not working_files:
            report_lines.append(f"Handover processing skipped: no handover files found for {self._date_ctx.date_str}.")
            print("No handover files were available.")
            return

        if self._args.handover_email_test:
            print("Handover email test mode: generating report from handover rows only.")
            preview_rows, row_unknown_codes = read_handover_rows(working_files)
            if row_unknown_codes:
                report_lines.append(
                    "Handover email test skipped unknown client codes: {codes}".format(
                        codes=", ".join(sorted(row_unknown_codes))
                    )
                )
            preview_entries = build_handover_report_preview_entries(
                preview_rows,
                self._args.handover_create_limit,
            )
            if not preview_entries:
                report_lines.append("Handover email test skipped: no handover rows available for report.")
                print("No handover rows were available for the email test report.")
                return
            report_path = self._write_handover_report(preview_entries)
            report_lines.append(f"Handover test report generated: {report_path}")
            print(f"Handover test report generated: {report_path}")
            self._send_handover_report_email(report_path, preview_entries, report_lines)
            return

        api_key = self._args.api_key or os.getenv("LEGALSUITE_API_KEY") or LEGALSUITE_API_KEY
        if not api_key:
            report_lines.append("Handover processing skipped: missing API key.")
            print("Handover processing skipped: missing API key.", file=sys.stderr)
            return

        created_matters = process_handover_files(
            paths=working_files,
            api_base=self._args.api_base,
            api_key=api_key,
            date_ctx=self._date_ctx,
            create_matters=not self._args.handover_dry_run,
            create_dry_run=self._args.handover_dry_run,
            create_limit=self._args.handover_create_limit,
            logged_in_employee_id=self._args.handover_logged_in_employee_id,
        )
        report_lines.append(
            "Handover processing completed for {count} file(s).".format(count=len(working_files))
        )
        if self._args.handover_dry_run:
            report_lines.append("Handover email skipped: dry-run mode.")
            return
        if not created_matters:
            report_lines.append("Handover email skipped: no new matters were created.")
            print("No new handover matters were created; report email skipped.")
            return

        report_path = self._write_handover_report(created_matters)
        report_lines.append(f"Handover report generated: {report_path}")
        print(f"Handover report generated: {report_path}")
        self._send_handover_report_email(report_path, created_matters, report_lines)

    def _write_handover_report(self, created_matters: list[HandoverCreatedMatter]) -> str:
        if Workbook is None:
            raise RuntimeError("openpyxl is not installed")

        report_dir = os.path.join(self._args.download_dir, "handover_reports")
        os.makedirs(report_dir, exist_ok=True)
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = os.path.join(report_dir, f"handover_created_matters_{self._date_ctx.date_str}_{timestamp}.xlsx")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Handover Report"
        worksheet.append(
            [
                "Matter File Reference",
                "Their Reference",
                "Matter Description",
            ]
        )
        for item in created_matters:
            worksheet.append([item.file_ref, item.their_reference, item.description])
        workbook.save(report_path)
        workbook.close()
        return report_path

    def _send_handover_report_email(
        self,
        report_path: str,
        created_matters: list[HandoverCreatedMatter],
        report_lines: list[str],
    ) -> None:
        smtp_host = os.getenv("MAIL_HOST", os.getenv("SMTP_HOST", "")).strip()
        smtp_port = int(os.getenv("MAIL_PORT", os.getenv("SMTP_PORT", "587")).strip() or "587")
        smtp_user = os.getenv("MAIL_USERNAME", os.getenv("SMTP_USER", "")).strip()
        smtp_pass = os.getenv("MAIL_PASSWORD", os.getenv("SMTP_PASS", "")).strip()
        smtp_from = os.getenv("MAIL_FROM_ADDRESS", os.getenv("SMTP_FROM", smtp_user)).strip()
        encryption = os.getenv("MAIL_ENCRYPTION", "").strip().lower()
        smtp_use_tls = encryption not in {"", "null", "none", "false", "0", "no"} if "MAIL_ENCRYPTION" in os.environ else (
            os.getenv("SMTP_USE_TLS", "true").strip().lower() not in {"0", "false", "no"}
        )

        missing = [name for name, value in (("SMTP_HOST", smtp_host), ("SMTP_FROM", smtp_from)) if not value]
        if missing:
            message = f"Handover email skipped: missing SMTP settings: {', '.join(missing)}"
            report_lines.append(message)
            print(message)
            return

        to_recipients = HANDOVER_REPORT_TEST_TO if self._args.handover_email_test else HANDOVER_REPORT_TO
        cc_recipients = HANDOVER_REPORT_TEST_CC if self._args.handover_email_test else HANDOVER_REPORT_CC
        subject_prefix = "[TEST] " if self._args.handover_email_test else ""

        message = EmailMessage()
        message["Subject"] = f"{subject_prefix}Handover Matter Creation Report - {self._date_ctx.date_str}"
        message["From"] = smtp_from
        message["To"] = ", ".join(to_recipients)
        if cc_recipients:
            message["Cc"] = ", ".join(cc_recipients)
        message.set_content(
            "Good Day All,\n\n"
            "I trust you are well.\n\n"
            "Please find attached the references for the new matters that have been imported over into legal suite.\n\n"
            "Kind Regards,\n"
        )

        with open(report_path, "rb") as handle:
            message.add_attachment(
                handle.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=os.path.basename(report_path),
            )

        all_recipients = to_recipients + cc_recipients
        try:
            with smtplib.SMTP(smtp_host, smtp_port, timeout=60) as server:
                if smtp_use_tls:
                    server.starttls()
                if smtp_user:
                    server.login(smtp_user, smtp_pass)
                server.send_message(message, from_addr=smtp_from, to_addrs=all_recipients)
        except Exception as exc:
            report_lines.append(f"Handover email failed: {exc}")
            print(f"Handover email failed: {exc}", file=sys.stderr)
            return

        report_lines.append(
            "Handover email sent: to={to_count}, cc={cc_count}, test_mode={test_mode}".format(
                to_count=len(to_recipients),
                cc_count=len(cc_recipients),
                test_mode=self._args.handover_email_test,
            )
        )
        print(
            "Handover email sent: to={to_count}, cc={cc_count}, test_mode={test_mode}".format(
                to_count=len(to_recipients),
                cc_count=len(cc_recipients),
                test_mode=self._args.handover_email_test,
            )
        )

    def _update_matter_extrascreens(self, report_lines: list[str]) -> None:
        if load_workbook is None:
            report_lines.append("Extrascreen update skipped: openpyxl is not installed.")
            print("Extrascreen update skipped: openpyxl is not installed.", file=sys.stderr)
            return

        api_key = self._args.api_key or os.getenv("LEGALSUITE_API_KEY") or LEGALSUITE_API_KEY
        if not api_key:
            report_lines.append("Extrascreen update skipped: missing API key.")
            print("Extrascreen update skipped: missing API key.", file=sys.stderr)
            return

        client = LegalSuiteClient(self._args.api_base, api_key)
        feedback_files = self._resolve_feedback_files()
        ptp_files = self._resolve_ptp_files()
        poc_summons_files = self._resolve_poc_summons_files()
        if self._args.extrascreen_only:
            allowed = {self._args.extrascreen_only}
            feedback_files = self._filter_extrascreen_files(feedback_files, allowed)
            ptp_files = self._filter_extrascreen_files(ptp_files, allowed)
            poc_summons_files = self._filter_extrascreen_files(poc_summons_files, allowed)
        if not feedback_files and not ptp_files and not poc_summons_files:
            report_lines.append("Extrascreen update skipped: no matching files found.")
            print("Extrascreen update skipped: no matching files found.")
            return

        processed = 0
        updated = 0
        failed = 0

        for cleaned_path, mapping in feedback_files + ptp_files + poc_summons_files:
            if self._args.extrascreen_verbose:
                label = self._extrascreen_mapping_label(mapping)
                print(f"Extrascreen file ({label}): {cleaned_path}")
            download_path = self._download_path_for_cleaned(cleaned_path)
            verification_source_path = download_path if os.path.exists(download_path) else cleaned_path
            header_row = self._read_header_row(download_path)
            if not header_row:
                report_lines.append(f"Header not found for extrascreen update: {cleaned_path}")
                continue

            col_map, file_ref_idx, screen_id_idx = self._build_extrascreen_column_map(header_row, mapping)
            if not file_ref_idx or not screen_id_idx:
                report_lines.append(f"Missing File Reference or Desktop Extra ScreenID: {cleaned_path}")
                continue

            workbook = load_workbook(cleaned_path, read_only=False, data_only=True)
            try:
                for worksheet in workbook.worksheets:
                    max_col = worksheet.max_column or max(col_map.values(), default=0)
                    header_row_values = tuple(header_row)
                    data_start_row = self._worksheet_data_start_row(worksheet, file_ref_idx)
                    verification_row_offset = 1 if verification_source_path != cleaned_path and data_start_row == 1 else 0
                    processing_rows = self._iter_processing_rows(
                        worksheet,
                        max_col=max_col,
                        data_start_row=data_start_row,
                        header_row=header_row,
                    )
                    if self._args.extrascreen_verbose and self._find_header_index(header_row, {"ptpcapturedate"}):
                        print(
                            "  Sorting worksheet rows by PTPCaptureDate ascending before processing: "
                            f"{worksheet.title}"
                        )
                    for row_number, row in processing_rows:
                        verification_row_number = row_number + verification_row_offset
                        file_ref = self._cell_text(row, file_ref_idx)
                        docscreenid = self._cell_text(row, screen_id_idx)
                        if not file_ref or not docscreenid:
                            continue
                        field_payload = self._build_extrascreen_payload(row, col_map, mapping)
                        if not field_payload:
                            continue
                        try:
                            matter = client.get_matter_by_fileref(file_ref)
                            recordid = matter.get("recordid")
                            if not recordid:
                                report_lines.append(f"Extrascreen update skipped (missing recordid): {file_ref}")
                                self._record_verification_result(
                                    verification_source_path,
                                    worksheet.title,
                                    verification_row_number,
                                    "Skipped",
                                    "Missing recordid on fetched matter.",
                                    matter,
                                )
                                continue
                            payload = {
                                "matterid": recordid,
                                "docscreenid": docscreenid,
                                **field_payload,
                            }
                            processed += 1
                            if self._args.extrascreen_dry_run:
                                print(f"Extrascreen dry-run: {file_ref} -> docscreenid={docscreenid}")
                                if self._args.extrascreen_verbose:
                                    print(json.dumps(payload, indent=2, default=str))
                                continue
                            if self._args.extrascreen_verbose:
                                print(f"Extrascreen update payload for {file_ref}:")
                                print(json.dumps(payload, indent=2, default=str))
                            result = client.update_matter_extrascreen(payload)
                            if self._args.extrascreen_verbose:
                                print(f"Extrascreen update response for {file_ref}:")
                                print(json.dumps(result, indent=2, default=str))
                            print(f"Extrascreen updated: {file_ref} -> docscreenid={docscreenid}")
                            fetched_rows = client.get_matter_extrascreen(recordid, docscreenid)
                            verification_status = "Verified"
                            verification_notes = ""
                            verification_response: object = fetched_rows
                            verification_values: dict[str, object] | None = None
                            if not fetched_rows:
                                verification_status = "Missing GET data"
                                verification_notes = "No extrascreen data returned after update."
                                print(
                                    f"  Extrascreen verification failed: no data returned for {file_ref} -> "
                                    f"docscreenid={docscreenid}"
                                )
                            else:
                                fetched_row = fetched_rows[0]
                                verification_response = fetched_row
                                verification_values = self._build_extrascreen_verification_values(
                                    header_row_values,
                                    col_map,
                                    mapping,
                                    fetched_row,
                                )
                                mismatches = compare_extrascreen_payload_to_row(field_payload, fetched_row)
                                if mismatches:
                                    mismatch_names = ", ".join(field_name for field_name, _, _ in mismatches)
                                    verification_status = "Mismatch"
                                    verification_notes = f"Mismatched fields: {mismatch_names}"
                                    print(
                                        f"  Extrascreen verification mismatches for {file_ref} -> "
                                        f"docscreenid={docscreenid}: {mismatch_names}"
                                    )
                                    for field_name, sent_value, fetched_value in mismatches:
                                        print(f"    {field_name}: sent={sent_value!r} fetched={fetched_value!r}")
                                else:
                                    verified_fields = sorted(key for key in field_payload if key.startswith("field"))
                                    print(
                                        f"  Extrascreen verified for {file_ref} -> docscreenid={docscreenid}: "
                                        f"{', '.join(verified_fields)}"
                                    )
                            self._record_verification_result(
                                verification_source_path,
                                worksheet.title,
                                verification_row_number,
                                verification_status,
                                verification_notes,
                                verification_response,
                                verification_values,
                            )
                            updated += 1
                        except Exception as exc:
                            failed += 1
                            report_lines.append(f"Extrascreen update failed for {file_ref}: {exc}")
                            print(f"Extrascreen update failed for {file_ref}: {exc}", file=sys.stderr)
                            self._record_verification_result(
                                verification_source_path,
                                worksheet.title,
                                verification_row_number,
                                "Failed",
                                str(exc),
                                None,
                            )
            finally:
                workbook.close()

        report_lines.append(
            "Extrascreen summary: processed={processed}, updated={updated}, failed={failed}".format(
                processed=processed,
                updated=updated,
                failed=failed,
            )
        )

    def _resolve_feedback_files(self) -> list[tuple[str, list[tuple[str, str, bool]]]]:
        files: list[tuple[str, list[tuple[str, str, bool]]]] = []
        debt_review = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Debt Review",
            "Debt_Review_Feedback_APT_LSW",
            f"Standard_Bank_Panel_L_Update_{self._date_ctx.date_str}_DR.xlsx",
        )
        if os.path.exists(debt_review):
            files.append((debt_review, FEEDBACK_FIELD_MAP))

        panel_l_pattern = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Panel L",
            "Feedback_APT_LSW",
            self._date_ctx.month_year,
            f"*_{self._date_ctx.date_str}.xlsx",
        )
        for path in sorted(glob.glob(panel_l_pattern)):
            files.append((path, FEEDBACK_FIELD_MAP))
        return files

    def _resolve_ptp_files(self) -> list[tuple[str, list[tuple[str, str, bool]]]]:
        files: list[tuple[str, list[tuple[str, str, bool]]]] = []
        debt_review = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Debt Review",
            "Debt_Review_PTP_APT_LSW",
            f"Standard_Bank_Panel_L_PTP_{self._date_ctx.date_str}_DR.xlsx",
        )
        if os.path.exists(debt_review):
            files.append((debt_review, PTP_FIELD_MAP))

        panel_l_pattern = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Panel L",
            "PTP_APT_LSW",
            self._date_ctx.month_year,
            f"*_{self._date_ctx.date_str}.xlsx",
        )
        for path in sorted(glob.glob(panel_l_pattern)):
            files.append((path, PTP_FIELD_MAP))
        return files

    def _resolve_poc_summons_files(self) -> list[tuple[str, list[tuple[str, str, bool]]]]:
        files: list[tuple[str, list[tuple[str, str, bool]]]] = []
        poc_summons_path = os.path.join(
            self._args.cleaned_dir,
            "SBSA POC AND SUMMONS",
            f"{self._date_ctx.day}{self._date_ctx.month}{self._date_ctx.year}.xlsx",
        )
        if os.path.exists(poc_summons_path):
            files.append((poc_summons_path, POC_SUMMONS_FIELD_MAP))
        return files

    def _filter_extrascreen_files(
        self,
        files: list[tuple[str, list[tuple[str, str, bool]]]],
        allowed: set[str],
    ) -> list[tuple[str, list[tuple[str, str, bool]]]]:
        filtered: list[tuple[str, list[tuple[str, str, bool]]]] = []
        for path, mapping in files:
            if self._extrascreen_mapping_label(mapping) in allowed:
                filtered.append((path, mapping))
        return filtered

    @staticmethod
    def _extrascreen_mapping_label(mapping: list[tuple[str, str, bool]]) -> str:
        if mapping is FEEDBACK_FIELD_MAP:
            return "feedback"
        if mapping is PTP_FIELD_MAP:
            return "ptp"
        if mapping is POC_SUMMONS_FIELD_MAP:
            return "poc-summons"
        return "extrascreen"

    def _build_extrascreen_column_map(
        self,
        header_row: list[object],
        mapping: list[tuple[str, str, bool]],
    ) -> tuple[dict[str, int], int | None, int | None]:
        col_map: dict[str, int] = {}
        file_ref_idx = None
        screen_id_idx = None
        for idx, value in enumerate(header_row, start=1):
            key = self._normalize_header(value)
            if key in {"fileref", "filereference"}:
                file_ref_idx = idx
            if key == "desktopextrascreenid":
                screen_id_idx = idx
            for header_key, field_name, _ in mapping:
                if key == header_key:
                    col_map[field_name] = idx
        return col_map, file_ref_idx, screen_id_idx

    def _build_extrascreen_payload(
        self,
        row: tuple[object, ...],
        col_map: dict[str, int],
        mapping: list[tuple[str, str, bool]],
    ) -> dict:
        payload: dict[str, object] = {}
        for header_key, field_name, is_date in mapping:
            col_idx = col_map.get(field_name)
            if not col_idx or col_idx > len(row):
                continue
            raw_value = row[col_idx - 1]
            if raw_value in (None, ""):
                continue
            if is_date:
                encoded = self._encode_legalsuite_date(raw_value)
                if encoded is None:
                    continue
                payload[field_name] = encoded
            else:
                payload[field_name] = self._normalize_cell_value(raw_value)
        return payload

    @staticmethod
    def _normalize_cell_value(value: object) -> object:
        if isinstance(value, float) and value.is_integer():
            return int(value)
        if isinstance(value, (dt.datetime, dt.date)):
            return value.strftime("%Y-%m-%d")
        return value

    def _record_verification_result(
        self,
        source_path: str,
        worksheet_name: str,
        row_number: int,
        status: str,
        notes: str,
        get_response: object | None,
        verified_values: dict[str, object] | None = None,
    ) -> None:
        self._verification_recorder.record_row(
            source_path=source_path,
            worksheet_name=worksheet_name,
            row_number=row_number,
            status=status,
            notes=notes,
            get_response=get_response,
            verified_values=verified_values,
        )

    @staticmethod
    def _joined_notes(*notes: str) -> str:
        return "; ".join(note for note in notes if note)

    @staticmethod
    def _build_extrascreen_verification_values(
        header_row: tuple[object, ...],
        col_map: dict[str, int],
        mapping: list[tuple[str, str, bool]],
        fetched_row: dict,
    ) -> dict[str, object]:
        verified_values: dict[str, object] = {}
        for _, field_name, _ in mapping:
            col_idx = col_map.get(field_name)
            if not col_idx or col_idx > len(header_row):
                continue
            header_value = header_row[col_idx - 1]
            header_text = str(header_value).strip() if header_value not in (None, "") else field_name
            verified_values[f"Verified {header_text}"] = fetched_row.get(field_name)
        return verified_values

    @staticmethod
    def _build_claim_verification_values(fetched_matter: dict) -> dict[str, object]:
        return {
            "Verified Claim Amount": fetched_matter.get("claimamount"),
        }

    @staticmethod
    def _build_archive_verification_values(fetched_matter: dict) -> dict[str, object]:
        return {
            "Verified Archive Flag": fetched_matter.get("archiveflag"),
            "Verified Archive Status": fetched_matter.get("archivestatus"),
            "Verified Archive Status Description": fetched_matter.get("archivestatusdescription"),
            "Verified Archive No": fetched_matter.get("archiveno"),
        }

    @staticmethod
    def _build_reopen_verification_values(fetched_matter: dict) -> dict[str, object]:
        return {
            "Verified Archive Flag": fetched_matter.get("archiveflag"),
            "Verified Archive Status": fetched_matter.get("archivestatus"),
            "Verified Archive Status Description": fetched_matter.get("archivestatusdescription"),
            "Verified Archive No": fetched_matter.get("archiveno"),
            "Verified Archive Date": fetched_matter.get("archivedate"),
        }

    def _worksheet_data_start_row(self, worksheet, file_ref_idx: int) -> int:
        first_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, max_col=max(file_ref_idx, 1), values_only=True),
            tuple(),
        )
        file_ref_value = self._cell_text(first_row, file_ref_idx)
        if file_ref_value and self._is_header_value(file_ref_value):
            return 2
        return 1

    def _iter_processing_rows(
        self,
        worksheet,
        max_col: int,
        data_start_row: int,
        header_row: list[object] | tuple[object, ...],
    ) -> list[tuple[int, tuple[object, ...]]]:
        rows = [
            (row_number, tuple(row))
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=data_start_row, max_col=max_col, values_only=True),
                start=data_start_row,
            )
        ]
        ptp_capture_idx = self._find_header_index(header_row, {"ptpcapturedate"})
        if not ptp_capture_idx:
            return rows

        return sorted(
            rows,
            key=lambda item: self._ptp_capture_sort_key(item[1], ptp_capture_idx, item[0]),
        )

    @staticmethod
    def _find_header_index(header_row: list[object] | tuple[object, ...], expected_keys: set[str]) -> int | None:
        for idx, value in enumerate(header_row, start=1):
            if App._normalize_header(value) in expected_keys:
                return idx
        return None

    def _ptp_capture_sort_key(
        self,
        row: tuple[object, ...],
        column_idx: int,
        row_number: int,
    ) -> tuple[int, dt.datetime, int]:
        value = self._cell_value(row, column_idx)
        parsed = self._parse_processing_date(value)
        if parsed is None:
            return (1, dt.datetime.max, row_number)
        return (0, parsed, row_number)

    @staticmethod
    def _parse_processing_date(value: object) -> dt.datetime | None:
        if value in (None, ""):
            return None
        if isinstance(value, dt.datetime):
            return value
        if isinstance(value, dt.date):
            return dt.datetime.combine(value, dt.time())
        if isinstance(value, (int, float)):
            try:
                return EXCEL_BASE + dt.timedelta(days=float(value))
            except (OverflowError, ValueError):
                return None

        text = str(value).strip()
        if not text:
            return None
        date_text = text.split()[0].split("T")[0].replace("/", "-")
        parts = date_text.split("-")
        if len(parts) == 3 and all(parts):
            if len(parts[0]) == 4:
                normalized = f"{parts[0]}-{parts[1]}-{parts[2]}"
            elif len(parts[2]) == 4:
                normalized = f"{parts[2]}-{parts[1]}-{parts[0]}"
            else:
                normalized = date_text
        else:
            normalized = date_text
        try:
            return dt.datetime.strptime(normalized, "%Y-%m-%d")
        except ValueError:
            return None

    @staticmethod
    def _encode_legalsuite_date(value: object) -> int | None:
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            return int(value) + LEGALSUITE_OFFSET
        if isinstance(value, dt.datetime):
            dt_value = value
        elif isinstance(value, dt.date):
            dt_value = dt.datetime.combine(value, dt.time())
        else:
            text = str(value).strip()
            if not text:
                return None
            date_text = text.split()[0].split("T")[0].replace("/", "-")
            parts = date_text.split("-")
            if len(parts) == 3 and all(parts):
                if len(parts[0]) == 4:
                    date_text = f"{parts[0]}-{parts[1]}-{parts[2]}"
                elif len(parts[2]) == 4:
                    date_text = f"{parts[2]}-{parts[1]}-{parts[0]}"
            try:
                dt_value = dt.datetime.strptime(date_text, "%Y-%m-%d")
            except ValueError:
                return None
        excel_serial = (dt_value - EXCEL_BASE).days
        return excel_serial + LEGALSUITE_OFFSET

    @staticmethod
    def _cell_text(row: tuple[object, ...], col_idx: int) -> str | None:
        if col_idx > len(row):
            return None
        value = row[col_idx - 1]
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def _archive_closed_matters(self, report_lines: list[str]) -> None:
        if load_workbook is None:
            report_lines.append("Archive skipped: openpyxl is not installed.")
            print("Archive skipped: openpyxl is not installed.", file=sys.stderr)
            return

        api_key = self._args.api_key or os.getenv("LEGALSUITE_API_KEY") or LEGALSUITE_API_KEY
        if not api_key:
            report_lines.append("Archive skipped: missing API key (use --api-key or LEGALSUITE_API_KEY).")
            print("Archive skipped: missing API key.", file=sys.stderr)
            return

        client = LegalSuiteClient(self._args.api_base, api_key)
        logged_in_employee_id = self._args.logged_in_employee_id or LEGALSUITE_EMPLOYEE_ID
        archive_status = self._args.archive_status or LEGALSUITE_ARCHIVE_STATUS
        closed_dir_panel = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Panel L",
            "Closed_APT_LSW",
            self._date_ctx.month_year,
        )
        closed_dir_debt = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Debt Review",
            "Debt_Review_Close_APT_LSW",
        )
        panel_pattern = os.path.join(closed_dir_panel, f"*_{self._date_ctx.date_str}.xlsx")
        debt_pattern = os.path.join(
            closed_dir_debt,
            f"Standard_Bank_Panel_L_Close_{self._date_ctx.date_str}_DR.xlsx",
        )
        cleaned_files = sorted(glob.glob(panel_pattern)) + sorted(glob.glob(debt_pattern))
        if not cleaned_files:
            report_lines.append(
                f"No cleaned closed files found: {panel_pattern} or {debt_pattern}"
            )
            print(f"No cleaned closed files found: {panel_pattern} or {debt_pattern}")
            return

        file_ref_occurrences: dict[str, list[tuple[str, str, int]]] = {}
        missing_header_files = 0
        for cleaned_path in cleaned_files:
            download_path = self._download_path_for_cleaned(cleaned_path)
            verification_source_path = download_path if os.path.exists(download_path) else cleaned_path
            col_info = self._find_fileref_column(download_path, cleaned_path)
            if not col_info:
                report_lines.append(f"FileRef column not found: {cleaned_path}")
                missing_header_files += 1
                continue
            col_idx, col_name = col_info
            refs = self._collect_file_ref_occurrences(verification_source_path, col_idx)
            if not refs:
                report_lines.append(f"No file refs found using {col_name}: {cleaned_path}")
                continue
            for file_ref, locations in refs.items():
                file_ref_occurrences.setdefault(file_ref, []).extend(
                    (verification_source_path, worksheet_name, row_number)
                    for worksheet_name, row_number in locations
                )

        file_refs = set(file_ref_occurrences)
        if not file_refs:
            report_lines.append("No file references found to archive.")
            print("No file references found to archive.")
            return
        print(f"Found {len(file_refs)} file references to process.")

        archived = 0
        failed = 0
        for file_ref in sorted(file_refs):
            try:
                matter = client.get_matter_by_fileref(file_ref)
                if self._args.archive_dry_run:
                    archived += 1
                    print(f"Dry-run: {file_ref}")
                    print(matter)
                    continue
                archive_no = matter.get("archiveno") or matter.get("archive_no") or matter.get("archivenumber")
                payload = client.build_archive_payload(
                    matter=matter,
                    logged_in_employee_id=logged_in_employee_id,
                    archive_no=archive_no,
                    archive_status=archive_status,
                )
                if self._args.archive_verbose:
                    print(f"Archive request for {file_ref}:")
                    print(json.dumps(payload, indent=2, default=str))
                print(f"Archiving matter: {file_ref}")
                result = client.update_matter(payload)
                if self._args.archive_verbose:
                    print(f"Archive response for {file_ref}:")
                    print(json.dumps(result, indent=2, default=str))
                verification_payload = payload
                verification_fields = [
                    "archiveflag",
                    "archivestatus",
                    "archivestatusdescription",
                    "archiveno",
                    "actual",
                    "reserved",
                    "invested",
                    "transfer",
                    "batchednormal",
                ]
                verification_status = "Verified"
                verification_notes = ""
                if is_archive_rejected_error(result):
                    print(f"Archive rejected for {file_ref}; setting Pending Deletion...")
                    pending_payload = client.build_pending_deletion_payload(
                        matter=matter,
                        logged_in_employee_id=logged_in_employee_id,
                        archive_no=archive_no,
                    )
                    if self._args.archive_verbose:
                        print(f"Pending Deletion request for {file_ref}:")
                        print(json.dumps(pending_payload, indent=2, default=str))
                    pending_result = client.update_matter(pending_payload)
                    if self._args.archive_verbose:
                        print(f"Pending Deletion response for {file_ref}:")
                        print(json.dumps(pending_result, indent=2, default=str))
                    archived_matter = client.get_matter_by_fileref(file_ref)
                    print_matter_verification(
                        "Pending Deletion fallback",
                        pending_payload,
                        archived_matter,
                        field_names=["archiveflag", "archivestatus", "archivestatusdescription", "archiveno"],
                    )
                    verification_payload = pending_payload
                    verification_fields = ["archiveflag", "archivestatus", "archivestatusdescription", "archiveno"]
                    verification_status = "Verified (fallback)"
                    verification_notes = "Archive rejected; Pending Deletion fallback used."
                else:
                    archived_matter = client.get_matter_by_fileref(file_ref)
                    print_matter_verification(
                        "Archive update",
                        payload,
                        archived_matter,
                        field_names=[
                            "archiveflag",
                            "archivestatus",
                            "archivestatusdescription",
                            "archiveno",
                            "actual",
                            "reserved",
                            "invested",
                            "transfer",
                            "batchednormal",
                        ],
                    )
                fetched_archive_flag = archived_matter.get("archiveflag")
                fetched_archive_status = archived_matter.get("archivestatus")
                fetched_archive_status_desc = archived_matter.get("archivestatusdescription")

                if normalize_compare_value(fetched_archive_status_desc).lower() == "live":
                    print(f"Archive did not stick for {file_ref}; setting Pending Deletion...")
                    pending_payload = client.build_pending_deletion_payload(
                        matter=archived_matter,
                        logged_in_employee_id=logged_in_employee_id,
                        archive_no=archive_no,
                    )
                    if self._args.archive_verbose:
                        print(f"Pending Deletion request for {file_ref}:")
                        print(json.dumps(pending_payload, indent=2, default=str))
                    pending_result = client.update_matter(pending_payload)
                    if self._args.archive_verbose:
                        print(f"Pending Deletion response for {file_ref}:")
                        print(json.dumps(pending_result, indent=2, default=str))
                    archived_matter = client.get_matter_by_fileref(file_ref)
                    print_matter_verification(
                        "Pending Deletion fallback",
                        pending_payload,
                        archived_matter,
                        field_names=["archiveflag", "archivestatus", "archivestatusdescription", "archiveno"],
                    )
                    verification_payload = pending_payload
                    verification_fields = ["archiveflag", "archivestatus", "archivestatusdescription", "archiveno"]
                    verification_status = "Verified (fallback)"
                    verification_notes = self._joined_notes(
                        verification_notes,
                        "Archive remained Live; Pending Deletion fallback used.",
                    )
                    fetched_archive_flag = archived_matter.get("archiveflag")
                    fetched_archive_status = archived_matter.get("archivestatus")
                    fetched_archive_status_desc = archived_matter.get("archivestatusdescription")

                final_mismatches = compare_matter_payload_to_row(
                    verification_payload,
                    archived_matter,
                    field_names=verification_fields,
                )
                if final_mismatches:
                    mismatch_names = ", ".join(field_name for field_name, _, _ in final_mismatches)
                    verification_status = "Mismatch"
                    verification_notes = self._joined_notes(
                        verification_notes,
                        f"Mismatched fields: {mismatch_names}",
                    )

                for cleaned_path, worksheet_name, row_number in file_ref_occurrences.get(file_ref, []):
                    self._record_verification_result(
                        cleaned_path,
                        worksheet_name,
                        row_number,
                        verification_status,
                        verification_notes,
                        archived_matter,
                        self._build_archive_verification_values(archived_matter),
                    )

                archived += 1
                report_lines.append(
                    "Archived matter: {file_ref} | archiveflag={archive_flag} | "
                    "archivestatus={archive_status} | archivestatusdescription={archive_status_desc}".format(
                        file_ref=file_ref,
                        archive_flag=fetched_archive_flag,
                        archive_status=fetched_archive_status,
                        archive_status_desc=fetched_archive_status_desc,
                    )
                )
                print(
                    "Archived matter: {file_ref} | archiveflag={archive_flag} | "
                    "archivestatus={archive_status} | archivestatusdescription={archive_status_desc}".format(
                        file_ref=file_ref,
                        archive_flag=fetched_archive_flag,
                        archive_status=fetched_archive_status,
                        archive_status_desc=fetched_archive_status_desc,
                    )
                )
            except Exception as exc:
                failed += 1
                report_lines.append(f"Archive failed for {file_ref}: {exc}")
                print(f"Archive failed for {file_ref}: {exc}", file=sys.stderr)
                for cleaned_path, worksheet_name, row_number in file_ref_occurrences.get(file_ref, []):
                    self._record_verification_result(
                        cleaned_path,
                        worksheet_name,
                        row_number,
                        "Failed",
                        str(exc),
                        None,
                    )

        report_lines.append(
            "Archive summary: archived={archived}, failed={failed}, files_missing_header={missing}".format(
                archived=archived,
                failed=failed,
                missing=missing_header_files,
            )
        )

    def _reopen_matters(self, report_lines: list[str]) -> None:
        if load_workbook is None:
            report_lines.append("Reopen skipped: openpyxl is not installed.")
            print("Reopen skipped: openpyxl is not installed.", file=sys.stderr)
            return

        api_key = self._args.api_key or os.getenv("LEGALSUITE_API_KEY") or LEGALSUITE_API_KEY
        if not api_key:
            report_lines.append("Reopen skipped: missing API key (use --api-key or LEGALSUITE_API_KEY).")
            print("Reopen skipped: missing API key.", file=sys.stderr)
            return

        client = LegalSuiteClient(self._args.api_base, api_key)
        logged_in_employee_id = self._args.logged_in_employee_id or LEGALSUITE_EMPLOYEE_ID
        panel_pattern = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Panel L",
            "Reopen_APT_LSW",
            self._date_ctx.month_year,
            f"*_{self._date_ctx.date_str}.xlsx",
        )
        debt_pattern = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Debt Review",
            "Debt_Review_Reopen_APT_LSW",
            f"Standard_Bank_Panel_L_Reopen_{self._date_ctx.date_str}_DR.xlsx",
        )
        cleaned_files = sorted(glob.glob(panel_pattern)) + sorted(glob.glob(debt_pattern))
        if not cleaned_files:
            report_lines.append(
                f"No cleaned reopen files found: {panel_pattern} or {debt_pattern}"
            )
            print(f"No cleaned reopen files found: {panel_pattern} or {debt_pattern}")
            return

        file_ref_occurrences: dict[str, list[tuple[str, str, int]]] = {}
        missing_header_files = 0
        for cleaned_path in cleaned_files:
            download_path = self._download_path_for_cleaned(cleaned_path)
            verification_source_path = download_path if os.path.exists(download_path) else cleaned_path
            col_info = self._find_fileref_column(download_path, cleaned_path)
            if not col_info:
                report_lines.append(f"FileRef column not found: {cleaned_path}")
                missing_header_files += 1
                continue
            col_idx, col_name = col_info
            refs = self._collect_file_ref_occurrences(verification_source_path, col_idx)
            if not refs:
                report_lines.append(f"No file refs found using {col_name}: {cleaned_path}")
                continue
            for file_ref, locations in refs.items():
                file_ref_occurrences.setdefault(file_ref, []).extend(
                    (verification_source_path, worksheet_name, row_number)
                    for worksheet_name, row_number in locations
                )

        file_refs = set(file_ref_occurrences)
        if not file_refs:
            report_lines.append("No file references found to reopen.")
            print("No file references found to reopen.")
            return
        print(f"Found {len(file_refs)} file references to process for reopen.")

        reopened = 0
        failed = 0
        for file_ref in sorted(file_refs):
            try:
                matter = client.get_matter_by_fileref(file_ref)
                if self._args.reopen_dry_run:
                    reopened += 1
                    print(f"Reopen dry-run: {file_ref}")
                    print(matter)
                    continue

                payload = client.build_reopen_payload(
                    matter=matter,
                    logged_in_employee_id=logged_in_employee_id,
                )
                if self._args.reopen_verbose:
                    print(f"Reopen request for {file_ref}:")
                    print(json.dumps(payload, indent=2, default=str))
                print(f"Reopening matter: {file_ref}")
                result = client.update_matter(payload)
                if self._args.reopen_verbose:
                    print(f"Reopen response for {file_ref}:")
                    print(json.dumps(result, indent=2, default=str))

                reopened_matter = client.get_matter_by_fileref(file_ref)
                print_matter_verification(
                    "Reopen update",
                    payload,
                    reopened_matter,
                    field_names=["archiveflag", "archivestatus", "archivestatusdescription", "archiveno"],
                )
                verification_status = "Verified"
                verification_notes = ""
                final_mismatches = compare_matter_payload_to_row(
                    payload,
                    reopened_matter,
                    field_names=["archiveflag", "archivestatus", "archivestatusdescription", "archiveno"],
                )
                if final_mismatches:
                    mismatch_names = ", ".join(field_name for field_name, _, _ in final_mismatches)
                    verification_status = "Mismatch"
                    verification_notes = f"Mismatched fields: {mismatch_names}"

                fetched_archive_flag = reopened_matter.get("archiveflag")
                fetched_archive_status = reopened_matter.get("archivestatus")
                fetched_archive_status_desc = reopened_matter.get("archivestatusdescription")

                for cleaned_path, worksheet_name, row_number in file_ref_occurrences.get(file_ref, []):
                    self._record_verification_result(
                        cleaned_path,
                        worksheet_name,
                        row_number,
                        verification_status,
                        verification_notes,
                        reopened_matter,
                        self._build_reopen_verification_values(reopened_matter),
                    )

                reopened += 1
                report_lines.append(
                    "Reopened matter: {file_ref} | archiveflag={archive_flag} | "
                    "archivestatus={archive_status} | archivestatusdescription={archive_status_desc}".format(
                        file_ref=file_ref,
                        archive_flag=fetched_archive_flag,
                        archive_status=fetched_archive_status,
                        archive_status_desc=fetched_archive_status_desc,
                    )
                )
                print(
                    "Reopened matter: {file_ref} | archiveflag={archive_flag} | "
                    "archivestatus={archive_status} | archivestatusdescription={archive_status_desc}".format(
                        file_ref=file_ref,
                        archive_flag=fetched_archive_flag,
                        archive_status=fetched_archive_status,
                        archive_status_desc=fetched_archive_status_desc,
                    )
                )
            except Exception as exc:
                failed += 1
                report_lines.append(f"Reopen failed for {file_ref}: {exc}")
                print(f"Reopen failed for {file_ref}: {exc}", file=sys.stderr)
                for cleaned_path, worksheet_name, row_number in file_ref_occurrences.get(file_ref, []):
                    self._record_verification_result(
                        cleaned_path,
                        worksheet_name,
                        row_number,
                        "Failed",
                        str(exc),
                        None,
                    )

        report_lines.append(
            "Reopen summary: reopened={reopened}, failed={failed}, files_missing_header={missing}".format(
                reopened=reopened,
                failed=failed,
                missing=missing_header_files,
            )
        )

    def _update_claim_amounts(self, report_lines: list[str]) -> None:
        if load_workbook is None:
            report_lines.append("Claim amount update skipped: openpyxl is not installed.")
            print("Claim amount update skipped: openpyxl is not installed.", file=sys.stderr)
            return

        api_key = self._args.api_key or os.getenv("LEGALSUITE_API_KEY") or LEGALSUITE_API_KEY
        if not api_key:
            report_lines.append("Claim amount update skipped: missing API key.")
            print("Claim amount update skipped: missing API key.", file=sys.stderr)
            return

        claim_files = self._resolve_claim_amount_files()
        if not claim_files:
            report_lines.append("Claim amount update skipped: no matching files found.")
            print("Claim amount update skipped: no matching files found.")
            return

        client = LegalSuiteClient(self._args.api_base, api_key)
        logged_in_employee_id = self._args.logged_in_employee_id or LEGALSUITE_EMPLOYEE_ID
        processed = 0
        updated = 0
        failed = 0
        missing_headers = 0

        for claim_path in claim_files:
            header_row = self._read_header_row(claim_path)
            if not header_row:
                report_lines.append(f"Claim amount header not found: {claim_path}")
                missing_headers += 1
                continue
            file_ref_idx, claim_amount_idx = self._find_claim_amount_columns(header_row)
            if not file_ref_idx or not claim_amount_idx:
                report_lines.append(f"Claim amount columns missing: {claim_path}")
                missing_headers += 1
                continue

            workbook = load_workbook(claim_path, read_only=False, data_only=True)
            try:
                for worksheet in workbook.worksheets:
                    max_col = worksheet.max_column or max(file_ref_idx, claim_amount_idx)
                    for row_number, row in enumerate(
                        worksheet.iter_rows(min_row=2, max_col=max_col, values_only=True),
                        start=2,
                    ):
                        file_ref = self._cell_text(row, file_ref_idx)
                        claim_amount = self._normalize_claim_amount(self._cell_value(row, claim_amount_idx))
                        if not file_ref or claim_amount is None:
                            continue
                        try:
                            matter = client.get_matter_by_fileref(file_ref)
                            payload = client.build_claim_amount_payload(
                                matter=matter,
                                logged_in_employee_id=logged_in_employee_id,
                                claim_amount=claim_amount,
                            )
                            processed += 1
                            if self._args.claim_amount_dry_run:
                                print(f"Claim amount dry-run: {file_ref} -> {claim_amount}")
                                if self._args.claim_amount_verbose:
                                    print(json.dumps(payload, indent=2, default=str))
                                continue
                            if self._args.claim_amount_verbose:
                                print(f"Claim amount update payload for {file_ref}:")
                                print(json.dumps(payload, indent=2, default=str))
                            result = client.update_matter(payload)
                            if self._args.claim_amount_verbose:
                                print(f"Claim amount update response for {file_ref}:")
                                print(json.dumps(result, indent=2, default=str))
                            print(f"Claim amount updated: {file_ref} -> {claim_amount}")
                            fetched_matter = client.get_matter_by_fileref(file_ref)
                            verification_status = "Verified"
                            verification_notes = ""
                            claim_mismatches = compare_matter_payload_to_row(
                                payload,
                                fetched_matter,
                                field_names=["claimamount"],
                            )
                            if claim_mismatches:
                                print("  Claim amount update verification mismatches: claimamount")
                                for field_name, sent_value, fetched_value in claim_mismatches:
                                    print(f"    {field_name}: sent={sent_value!r} fetched={fetched_value!r}")
                                update_error_text = extract_update_error_text(result).strip()
                                verification_status = "Mismatch"
                                verification_notes = self._joined_notes(
                                    "Mismatched fields: claimamount",
                                    update_error_text if update_error_text not in {"", "{}", "[]"} else "",
                                )
                                if update_error_text and update_error_text not in {"{}", "[]"}:
                                    print(f"  Claim amount update response message: {update_error_text}")
                                if is_old_code_unique_error(result):
                                    print(
                                        f"  Retrying claim amount update for {file_ref} using File Ref-only payload..."
                                    )
                                    fallback_payload = client.build_claim_amount_fileref_only_payload(
                                        matter=matter,
                                        file_ref=file_ref,
                                        logged_in_employee_id=logged_in_employee_id,
                                        claim_amount=claim_amount,
                                    )
                                    if self._args.claim_amount_verbose:
                                        print(f"Claim amount fallback payload for {file_ref}:")
                                        print(json.dumps(fallback_payload, indent=2, default=str))
                                    fallback_result = client.update_matter(fallback_payload)
                                    if self._args.claim_amount_verbose:
                                        print(f"Claim amount fallback response for {file_ref}:")
                                        print(json.dumps(fallback_result, indent=2, default=str))
                                    fetched_matter = client.get_matter_by_fileref(file_ref)
                                    fallback_mismatches = compare_matter_payload_to_row(
                                        fallback_payload,
                                        fetched_matter,
                                        field_names=["claimamount"],
                                    )
                                    if fallback_mismatches:
                                        print("  Claim amount fallback verification mismatches: claimamount")
                                        for field_name, sent_value, fetched_value in fallback_mismatches:
                                            print(f"    {field_name}: sent={sent_value!r} fetched={fetched_value!r}")
                                        fallback_error_text = extract_update_error_text(fallback_result).strip()
                                        verification_status = "Mismatch"
                                        verification_notes = self._joined_notes(
                                            "Fallback mismatch on claimamount",
                                            fallback_error_text if fallback_error_text not in {"", "{}", "[]"} else "",
                                        )
                                        if fallback_error_text and fallback_error_text not in {"{}", "[]"}:
                                            print(f"  Claim amount fallback response message: {fallback_error_text}")
                                    else:
                                        verification_status = "Verified (fallback)"
                                        verification_notes = "Fallback File Ref-only payload verified."
                                        print("  Claim amount fallback verified: claimamount")
                            else:
                                print("  Claim amount update verified: claimamount")

                            self._record_verification_result(
                                claim_path,
                                worksheet.title,
                                row_number,
                                verification_status,
                                verification_notes,
                                fetched_matter,
                                self._build_claim_verification_values(fetched_matter),
                            )
                            updated += 1
                        except Exception as exc:
                            failed += 1
                            report_lines.append(f"Claim amount update failed for {file_ref}: {exc}")
                            print(f"Claim amount update failed for {file_ref}: {exc}", file=sys.stderr)
                            self._record_verification_result(
                                claim_path,
                                worksheet.title,
                                row_number,
                                "Failed",
                                str(exc),
                                None,
                            )
            finally:
                workbook.close()

        report_lines.append(
            "Claim amount summary: processed={processed}, updated={updated}, "
            "failed={failed}, files_missing_header={missing}".format(
                processed=processed,
                updated=updated,
                failed=failed,
                missing=missing_headers,
            )
        )

    def _download_path_for_cleaned(self, cleaned_path: str) -> str:
        rel_path = os.path.relpath(cleaned_path, self._args.cleaned_dir)
        return os.path.join(self._args.download_dir, rel_path)

    def _collect_file_ref_occurrences(
        self,
        cleaned_path: str,
        col_idx: int,
    ) -> dict[str, list[tuple[str, int]]]:
        occurrences: dict[str, list[tuple[str, int]]] = {}
        workbook = load_workbook(cleaned_path, read_only=False, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                max_col = worksheet.max_column or col_idx
                for row_number, row in enumerate(
                    worksheet.iter_rows(min_row=1, max_col=max_col, values_only=True),
                    start=1,
                ):
                    if not row or len(row) < col_idx:
                        continue
                    value = row[col_idx - 1]
                    if value is None:
                        continue
                    text = str(value).strip()
                    if not text or self._is_header_value(text):
                        continue
                    occurrences.setdefault(text, []).append((worksheet.title, row_number))
        finally:
            workbook.close()
        return occurrences

    def _find_fileref_column(self, download_path: str, cleaned_path: str) -> tuple[int, str] | None:
        header_row = self._read_header_row(download_path) or self._read_header_row(cleaned_path)
        if not header_row:
            return None
        for idx, value in enumerate(header_row, start=1):
            key = self._normalize_header(value)
            if key in {"fileref", "filereference", "matterref", "matterreference"}:
                return idx, "fileref"
        return None

    def _read_header_row(self, path: str) -> list[object] | None:
        if not os.path.exists(path):
            return None
        workbook = load_workbook(path, read_only=False, data_only=True)
        if not workbook.worksheets:
            return None
        worksheet = workbook.worksheets[0]
        max_col = worksheet.max_column or 1
        for row in worksheet.iter_rows(min_row=1, max_row=1, max_col=max_col, values_only=True):
            return list(row)
        return None

    def _resolve_claim_amount_files(self) -> list[str]:
        pattern = os.path.join(
            self._args.cleaned_dir,
            "Standard Bank_ClaimsAmount",
            f"Standard Bank Legal Claim Amount_Panel_L{self._date_ctx.year}_{self._date_ctx.month}_{self._date_ctx.day}_*.xlsx",
        )
        return sorted(glob.glob(pattern))

    def _find_claim_amount_columns(self, header_row: list[object]) -> tuple[int | None, int | None]:
        file_ref_idx = None
        claim_amount_idx = None
        for idx, value in enumerate(header_row, start=1):
            key = self._normalize_header(value)
            if key in {"fileref", "filereference"}:
                file_ref_idx = idx
            if key in {"claimamount", "claimamounts", "claimamt"}:
                claim_amount_idx = idx
        return file_ref_idx, claim_amount_idx

    @staticmethod
    def _cell_value(row: tuple[object, ...], col_idx: int) -> object | None:
        if col_idx > len(row):
            return None
        return row[col_idx - 1]

    @staticmethod
    def _normalize_claim_amount(value: object) -> object | None:
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            return value
        text = str(value).strip()
        if not text:
            return None
        cleaned = re.sub(r"[^0-9.\-]", "", text)
        if not cleaned or cleaned in {"-", ".", "-."}:
            return None
        try:
            if "." in cleaned:
                return float(cleaned)
            return int(cleaned)
        except ValueError:
            return None

    @staticmethod
    def _normalize_header(header: object) -> str:
        if header is None:
            return ""
        return "".join(ch for ch in str(header).strip().lower() if ch.isalnum())

    @staticmethod
    def _is_header_value(value: str) -> bool:
        return App._normalize_header(value) in {"fileref", "filereference", "matterref", "matterreference"}

    @staticmethod
    def _resolve_date(date_arg: str | None, days_ago: int) -> DateContext:
        if date_arg:
            try:
                date_val = dt.datetime.strptime(date_arg, "%Y%m%d")
            except ValueError as exc:
                raise ValueError("Date must be in YYYYMMDD format.") from exc
        else:
            if days_ago < 0:
                raise ValueError("days_ago must be 0 or greater.")
            date_val = dt.datetime.now() - dt.timedelta(days=days_ago)
        return DateContext(
            date_str=date_val.strftime("%Y%m%d"),
            month_year=date_val.strftime("%b %Y"),
            year=date_val.strftime("%Y"),
            month=date_val.strftime("%m"),
            day=date_val.strftime("%d"),
        )

    @staticmethod
    def _build_targets(date_ctx: DateContext) -> list[tuple[str, str]]:
        items: list[tuple[str, str]] = []
        for dir_tmpl, file_tmpl in TARGETS:
            remote_dir = dir_tmpl.format(
                date=date_ctx.date_str,
                month_year=date_ctx.month_year,
                year=date_ctx.year,
                month=date_ctx.month,
                day=date_ctx.day,
            )
            filename = file_tmpl.format(
                date=date_ctx.date_str,
                month_year=date_ctx.month_year,
                year=date_ctx.year,
                month=date_ctx.month,
                day=date_ctx.day,
            )
            items.append((remote_dir, filename))
        return items

    @staticmethod
    def _ensure_local_path(base_dir: str, remote_dir: str, filename: str) -> str:
        local_dir = os.path.join(base_dir, *remote_dir.split("/"))
        os.makedirs(local_dir, exist_ok=True)
        return os.path.join(local_dir, filename)

    @staticmethod
    def _normalize_cli_args(argv: list[str]) -> list[str]:
        normalized: list[str] = []
        for arg in argv:
            match = re.fullmatch(r"--days-(\d+)", arg)
            if match:
                normalized.extend(["--days-ago", match.group(1)])
                continue
            normalized.append(arg)
        return normalized

    @staticmethod
    def parse_args() -> argparse.Namespace:
        # Example: python3 ftp_download_today.py --date 20260402 --clean-only --log-file logs/report.txt
        parser = argparse.ArgumentParser(
            description="Download LegalSuite FTP files for a specific date (default: today)."
        )
        parser.add_argument(
            "--date",
            help="Date in YYYYMMDD format; defaults to today.",
        )
        parser.add_argument(
            "--days-ago",
            type=int,
            default=0,
            help="Download files from N days ago (default: 0). Also accepts shorthand like --days-15. Ignored if --date is provided.",
        )
        parser.add_argument(
            "--download-dir",
            default="downloads",
            help="Local base directory for downloads (default: downloads).",
        )
        parser.add_argument(
            "--timeout",
            type=int,
            default=30,
            help="FTP connection timeout in seconds (default: 30).",
        )
        parser.add_argument(
            "--log-file",
            help="Write a report log to this path (default: <download-dir>/report_YYYYMMDD.txt).",
        )
        parser.add_argument(
            "--cleaned-dir",
            default="cleaned",
            help="Local base directory for cleaned files (default: cleaned).",
        )
        parser.add_argument(
            "--verification-dir",
            default="verification",
            help="Local base directory for verification workbooks (default: verification).",
        )
        parser.add_argument(
            "--skip-clean",
            action="store_true",
            help="Skip cleaning downloaded files.",
        )
        parser.add_argument(
            "--clean-only",
            action="store_true",
            help="Reuse existing selected-date downloads; do not connect to FTP.",
        )
        parser.add_argument(
            "--skip-handover",
            action="store_true",
            help="Skip handover matter/party/MatParty processing.",
        )
        parser.add_argument(
            "--handover-dry-run",
            action="store_true",
            help="Preview handover matter processing without creating or updating records.",
        )
        parser.add_argument(
            "--handover-create-limit",
            type=int,
            help="Limit how many handover rows are processed.",
        )
        parser.add_argument(
            "--handover-logged-in-employee-id",
            default=str(CREATE_DEFAULTS["loggedinemployeeid"]),
            help="LegalSuite logged-in employee ID for handover-created matters and parties (default: 174).",
        )
        parser.add_argument(
            "--handover-email-test",
            action="store_true",
            help="Send the handover report to the test recipients instead of the production recipients.",
        )
        parser.add_argument(
            "--archive-closed",
            action="store_true",
            help="Update closed matters in LegalSuite using cleaned closed files.",
        )
        parser.add_argument(
            "--archive-dry-run",
            action="store_true",
            help="Only fetch matters for closed files; do not update.",
        )
        parser.add_argument(
            "--archive-verbose",
            action="store_true",
            help="Print LegalSuite update payload and response to the console.",
        )
        parser.add_argument(
            "--reopen-matters",
            action="store_true",
            help="Reopen matters in LegalSuite using cleaned reopen files.",
        )
        parser.add_argument(
            "--reopen-dry-run",
            action="store_true",
            help="Only fetch matters for reopen files; do not update.",
        )
        parser.add_argument(
            "--reopen-verbose",
            action="store_true",
            help="Print LegalSuite reopen payload and response to the console.",
        )
        parser.add_argument(
            "--update-extrascreen",
            action="store_true",
            help="Update Matter ExtraScreen data using feedback/PTP files.",
        )
        parser.add_argument(
            "--extrascreen-only",
            choices=["feedback", "ptp", "poc-summons"],
            help="Limit extrascreen updates to one file type.",
        )
        parser.add_argument(
            "--extrascreen-dry-run",
            action="store_true",
            help="Only fetch matters for extrascreen updates; do not update.",
        )
        parser.add_argument(
            "--extrascreen-verbose",
            action="store_true",
            help="Print extrascreen update payload and response to the console.",
        )
        parser.add_argument(
            "--update-claim-amount",
            action="store_true",
            help="Update LegalSuite claim amount using the claims file.",
        )
        parser.add_argument(
            "--claim-amount-dry-run",
            action="store_true",
            help="Only fetch matters for claim updates; do not update.",
        )
        parser.add_argument(
            "--claim-amount-verbose",
            action="store_true",
            help="Print claim amount update payload and response to the console.",
        )
        parser.add_argument(
            "--api-base",
            default=LEGALSUITE_API_BASE,
            help="LegalSuite API base URL (default: https://api.legalsuite.net).",
        )
        parser.add_argument(
            "--api-key",
            default=os.getenv("LEGALSUITE_API_KEY") or LEGALSUITE_API_KEY,
            help="LegalSuite API key (or set LEGALSUITE_API_KEY).",
        )
        parser.add_argument(
            "--logged-in-employee-id",
            default=LEGALSUITE_EMPLOYEE_ID,
            help="LegalSuite logged-in employee ID for updates (default: 1).",
        )
        parser.add_argument(
            "--archive-status",
            default=LEGALSUITE_ARCHIVE_STATUS,
            help="Archive status ID to send with updates (default: 2).",
        )
        return parser.parse_args(App._normalize_cli_args(sys.argv[1:]))


def main() -> int:
    try:
        args = App.parse_args()
        app = App(args)
        return app.run()
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 2
    except requests.RequestException as exc:
        print(f"LegalSuite API error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
