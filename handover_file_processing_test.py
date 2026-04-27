#!/usr/bin/env python3
import argparse
import datetime as dt
import fnmatch
import os
import re
import shutil
import sys
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass

import requests

from ftp_download_today import (
    EXCEL_BASE,
    FTP_HOST,
    FTP_PASS,
    FTP_USER,
    FTPClient,
    LEGALSUITE_API_BASE,
    LEGALSUITE_API_KEY,
    LEGALSUITE_OFFSET,
)


CLIENT_CODE_MAP = {
    "STA387": "150307",
    "DR387": "334695",
    "STD9": "155128",
    "DRR9": "334565",
    "STA482": "209250",
    "DR482": "334568",
    "STA822": "283850",
    "DR822": "334567",
    "STA614": "267742",
    "DR614": "334569",
}

CREATE_FIELD_COLUMN_MAP = {
    "theirref": "Reference",
    "claimamount": "Claim Amount",
    "interestrate": "Interest Rate",
    "employerid": "EmployerID",
    "tracingagentid": "TracingAgentID",
    "sheriffareaid": "SheriffAreaID",
    "sheriffid": "SheriffID",
    "branchid": "BranchID",
    "employeeid": "EmployeeID",
    "stagegroupid": "StageGroupID",
    "mattertypeid": "MatterTypeID",
    "debtorfeesheetid": "DebtorFeeSheetID",
    "clientfeesheetid": "ClientFeeSheetID",
    "debtorcollcommoption": "DebtorCollCommOption",
    "debtorcollcommpercent": "DebtorCollCommPercent",
    "collcommoption": "CollCommOption",
    "clientcollcommpercent": "ClientCollCommPercent",
    "costcentreid": "CostCentreID",
    "defendantemail": "DefendantEmail",
    "magcourtdistrict": "MagCourtDistrict",
    "magcourtheldat": "MagCourtHeldAt",
    "extrascreenid": "ExtraScreenID",
    "induplumamount": "In Duplum Amount",
    "maximuminterestamount": "Maximum Interest Amount",
    "alternateref": "Alternate Reference",
}

CREATE_DEFAULTS = {
    "mattertypeid": 4,
    "clientfeesheetid": 1,
    "docgenid": 5,
    "costcentreid": 193,
    "todogroupid": 1,
    "stagegroupid": 59,
    "debtorfeesheetid": 45,
    "businessbankid": 1103,
    "trustbankid": 1198,
    "employeeid": 174,
    "loggedinemployeeid": 174,
    "archivestatusdescription": "LIVE",
    "archivestatus": 2,
    "partyname": "STANDARD BANK",
    "employeename": "Angie Reddy",
    "mattypedescription": "Collections",
    "docgendescription": "Magistrate Court",
    "branchdescription": "STRAUSS DALY UMHLANGA",
    "docgencode": "MAG",
    "docgentype": "LIT",
    "costcentredescription": "A Reddy",
    "planofactiondescription": "Collections",
    "stagegroupdescription": "Collections APT - E4",
    "clientfeesheetdescription": "Conveyancing",
    "debtorfeesheetdescription": "Mag Court Tariff",
    "backgroundcolour": "FF0000",
}

PARTY_DEFAULTS = {
    "party[partytypeid]": 1,
    "party[defaultlanguageid]": 1,
    "party[createdid]": 1,
    "party[parlang][languageid]": 1,
}

MATPARTY_DEFAULTS = {
    "roleid": 103,
    "sorter": 1,
    "languageid": 0,
    "reference": "",
}

XLSX_NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
}


@dataclass(frozen=True)
class DateContext:
    date_str: str
    month_year: str


@dataclass(frozen=True)
class DownloadTarget:
    label: str
    remote_dirs: tuple[str, ...]
    filename_pattern: str


@dataclass
class HandoverRow:
    source_path: str
    row_number: int
    headers: list[str]
    row_values: tuple[object, ...]
    values_by_header: dict[str, object]
    client_code: str
    client_id: str
    reference: str | None


class LegalSuiteLookupClient:
    def __init__(self, api_base: str, api_key: str) -> None:
        self._api_base = api_base.rstrip("/")
        self._api_key = api_key

    def get_matters_by_clientid_and_prefix(self, client_id: str, prefix: str) -> list[dict]:
        url = f"{self._api_base}/matter/get"
        data = [
            ("where[]", f"Matter.ClientID,=,{client_id}"),
            ("where[]", f"Matter.FileRef,like,{prefix}/%"),
        ]
        response = requests.post(url, headers=self._headers(), data=data, timeout=120)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def get_matters_by_fileref(self, file_ref: str) -> list[dict]:
        url = f"{self._api_base}/matter/get"
        data = {
            "where[]": f"Matter.FileRef,=,{file_ref}",
        }
        response = requests.post(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def get_matters_by_clientid_and_reference(self, client_id: str, reference: str) -> list[dict]:
        url = f"{self._api_base}/matter/get"
        data = [
            ("where[]", f"Matter.ClientID,=,{client_id}"),
            ("where[]", f"Matter.TheirRef,=,{reference}"),
        ]
        response = requests.post(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def create_matter(self, data: dict) -> dict | str:
        url = f"{self._api_base}/matter/store"
        payload = {key: str(value) for key, value in data.items()}
        response = requests.post(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def get_matter_by_recordid(self, recordid: int | str) -> dict | str:
        url = f"{self._api_base}/matter/get"
        data = {
            "where[]": f"Matter.RecordID,=,{recordid}",
        }
        response = requests.post(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def update_matter(self, recordid: int | str, updates: dict) -> dict | str:
        url = f"{self._api_base}/matter/update"
        payload = {
            "recordid": str(recordid),
        }
        for key, value in updates.items():
            payload[key] = str(value)
        response = requests.post(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def create_party(self, data: dict) -> dict | str:
        url = f"{self._api_base}/party/store"
        payload = {key: str(value) for key, value in data.items()}
        response = requests.post(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def get_party_by_identitynumber(self, identity_number: str) -> list[dict]:
        url = f"{self._api_base}/party/get"
        data = {
            "where[]": f"Party.IdentityNumber,=,{identity_number}",
        }
        response = requests.post(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def get_matparty_by_matter_and_party(self, matter_id: int | str, party_id: int | str) -> list[dict]:
        url = f"{self._api_base}/matparty/get"
        data = [
            ("where[]", f"MatParty.MatterID,=,{matter_id}"),
            ("where[]", f"MatParty.PartyID,=,{party_id}"),
        ]
        response = requests.post(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def create_matparty(self, data: dict) -> dict | str:
        url = f"{self._api_base}/matparty/store"
        payload = {key: str(value) for key, value in data.items()}
        response = requests.post(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def update_matter_extrascreen(self, data: dict) -> dict | str:
        url = f"{self._api_base}/matdocsc/update"
        payload = {key: str(value) for key, value in data.items()}
        response = requests.post(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            return response.json()
        except ValueError:
            return response.text

    def get_matter_extrascreen(self, matter_id: int | str, docscreenid: int | str) -> list[dict]:
        url = f"{self._api_base}/matdocsc/get"
        data = [
            ("where[]", f"MatDocSc.MatterID,=,{matter_id}"),
            ("where[]", f"MatDocSc.DocScreenID,=,{docscreenid}"),
        ]
        response = requests.post(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def _headers(self) -> dict[str, str]:
        return {
            "Authorization": f"Bearer {self._api_key}",
            "Content-Type": "application/x-www-form-urlencoded",
        }


def resolve_date(date_arg: str | None, days_ago: int) -> DateContext:
    if date_arg:
        try:
            date_value = dt.datetime.strptime(date_arg, "%Y%m%d")
        except ValueError as exc:
            raise ValueError("Date must be in YYYYMMDD format.") from exc
    else:
        if days_ago < 0:
            raise ValueError("days_ago must be 0 or greater.")
        date_value = dt.datetime.now() - dt.timedelta(days=days_ago)

    return DateContext(
        date_str=date_value.strftime("%Y%m%d"),
        month_year=date_value.strftime("%b %Y"),
    )


def build_targets(date_ctx: DateContext) -> list[DownloadTarget]:
    return [
        DownloadTarget(
            label="Debt Review handover",
            remote_dirs=(
                "SBSA/Debt Review/Debt_Review_Handover_APT_LWS",
                "SBSA/Debt Review/Debt_Review_ Handover_APT_LWS",
            ),
            filename_pattern=f"Standard_Bank_Panel_L_Handover_{date_ctx.date_str}_DR.xlsx",
        ),
        DownloadTarget(
            label="Panel L handover",
            remote_dirs=(f"SBSA/Panel L/Handover_APT_LSW/{date_ctx.month_year}",),
            filename_pattern=f"*_{date_ctx.date_str}.xlsx",
        ),
    ]


def ensure_local_path(base_dir: str, remote_dir: str, filename: str) -> str:
    local_dir = os.path.join(base_dir, *remote_dir.split("/"))
    os.makedirs(local_dir, exist_ok=True)
    return os.path.join(local_dir, filename)


def download_handover_files(
    date_ctx: DateContext,
    download_dir: str,
    timeout: int,
) -> list[str]:
    downloaded_files: list[str] = []
    targets = build_targets(date_ctx)
    client = FTPClient(FTP_HOST, FTP_USER, FTP_PASS, timeout)

    print("Connecting to FTP...")
    client.connect()
    try:
        for target in targets:
            downloaded = False
            last_reason = "missing_dir"
            for remote_dir in target.remote_dirs:
                resolved_name, reason = client.resolve_remote_file(remote_dir, target.filename_pattern)
                if resolved_name:
                    remote_path = f"{remote_dir}/{resolved_name}"
                    local_path = ensure_local_path(download_dir, remote_dir, resolved_name)
                    client.download_file(remote_path, local_path)
                    downloaded_files.append(local_path)
                    downloaded = True
                    print(f"Downloaded {target.label}: {remote_path} -> {local_path}")
                    break
                if reason:
                    last_reason = reason

            if not downloaded:
                if last_reason == "missing_file":
                    print(f"Missing {target.label} file for {date_ctx.date_str}")
                else:
                    tried = ", ".join(target.remote_dirs)
                    print(f"Missing {target.label} directory on FTP: {tried}")
    finally:
        client.close()

    return downloaded_files


def resolve_local_handover_files(date_ctx: DateContext, base_dir: str, label: str = "existing") -> list[str]:
    resolved_files: list[str] = []

    for target in build_targets(date_ctx):
        found_path = None
        for remote_dir in target.remote_dirs:
            local_dir = os.path.join(base_dir, *remote_dir.split("/"))
            if not os.path.isdir(local_dir):
                continue

            try:
                names = sorted(
                    name for name in os.listdir(local_dir) if os.path.isfile(os.path.join(local_dir, name))
                )
            except OSError:
                continue

            if any(ch in target.filename_pattern for ch in ["*", "?", "["]):
                matches = [name for name in names if fnmatch.fnmatch(name, target.filename_pattern)]
                if matches:
                    found_path = os.path.join(local_dir, matches[-1])
                    break
            elif target.filename_pattern in names:
                found_path = os.path.join(local_dir, target.filename_pattern)
                break

        if found_path:
            resolved_files.append(found_path)
            print(f"Using {label} {target.label}: {found_path}")
        else:
            print(f"{label.capitalize()} {target.label} file not found for {date_ctx.date_str} in {base_dir}")

    return resolved_files


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def cell_value(row: tuple[object, ...], col_idx: int) -> object | None:
    if col_idx >= len(row):
        return None
    return row[col_idx]


def normalize_reference(value: object) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("'"):
        text = text[1:].strip()
    return text or None


def normalize_cell_value(value: object) -> object | None:
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("'"):
            text = text[1:].strip()
        return text or None
    return value


def normalize_money(value: object) -> object | None:
    value = normalize_cell_value(value)
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if not cleaned or cleaned in {"-", ".", "-."}:
        return None
    try:
        return float(cleaned) if "." in cleaned else int(cleaned)
    except ValueError:
        return None


def digits_only(value: object) -> str | None:
    if value in (None, ""):
        return None
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    return digits or None


def encode_legalsuite_date(value: object) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return int(value) + LEGALSUITE_OFFSET
    if isinstance(value, dt.datetime):
        date_value = value
    elif isinstance(value, dt.date):
        date_value = dt.datetime.combine(value, dt.time())
    else:
        text = str(value).strip()
        if not text:
            return None
        if re.fullmatch(r"-?\d+(?:\.0+)?", text):
            return int(float(text)) + LEGALSUITE_OFFSET
        date_text = text.split()[0].split("T")[0].replace("/", "-")
        parts = date_text.split("-")
        if len(parts) == 3 and all(parts):
            if len(parts[0]) == 4:
                date_text = f"{parts[0]}-{parts[1]}-{parts[2]}"
            elif len(parts[2]) == 4:
                date_text = f"{parts[2]}-{parts[1]}-{parts[0]}"
        try:
            date_value = dt.datetime.strptime(date_text, "%Y-%m-%d")
        except ValueError:
            return None
    excel_serial = (date_value - EXCEL_BASE).days
    return excel_serial + LEGALSUITE_OFFSET


def encode_legalsuite_time(value: dt.datetime | dt.time | None = None) -> int:
    if value is None:
        value = dt.datetime.now()
    if isinstance(value, dt.datetime):
        time_value = value.time()
    else:
        time_value = value
    return int(time_value.strftime("%H%M%S"))


def find_column_index(header_row: tuple[object, ...], expected_name: str) -> int | None:
    expected_key = normalize_header(expected_name)
    for idx, value in enumerate(header_row):
        if normalize_header(value) == expected_key:
            return idx
    return None


def find_header_row(
    rows,
    expected_name: str,
    max_scan_rows: int = 10,
) -> tuple[tuple[object, ...], int, int]:
    scanned_rows: list[tuple[object, ...]] = []

    for row_number, row in enumerate(rows, start=1):
        row_tuple = tuple(row)
        scanned_rows.append(row_tuple)
        column_idx = find_column_index(row_tuple, expected_name)
        if column_idx is not None:
            return row_tuple, column_idx, row_number
        if row_number >= max_scan_rows:
            break

    sample_rows = [
        ", ".join(str(value or "") for value in row[:8])
        for row in scanned_rows[:3]
    ]
    sample_text = " | ".join(sample_rows) if sample_rows else "<no rows scanned>"
    raise ValueError(
        f"{expected_name} column not found in first {max_scan_rows} row(s). "
        f"Header sample: {sample_text}"
    )


def read_client_codes_from_file(path: str) -> tuple[dict[str, int], list[str], dict[str, list[str]]]:
    rows = iter_excel_rows(path)
    header_row, client_code_idx, header_row_number = find_header_row(rows, "Client Code")
    reference_idx = find_column_index(header_row, "Reference")
    if reference_idx is None:
        raise ValueError(f"Reference column not found in {path}")

    counts: dict[str, int] = {}
    unknown_codes: list[str] = []
    references_by_code: dict[str, list[str]] = {}
    for row in rows:
        raw_value = cell_value(row, client_code_idx)
        if raw_value in (None, ""):
            continue
        client_code = str(raw_value).strip().upper()
        if not client_code or normalize_header(client_code) == "clientcode":
            continue
        counts[client_code] = counts.get(client_code, 0) + 1
        reference = normalize_reference(cell_value(row, reference_idx))
        if reference:
            references = references_by_code.setdefault(client_code, [])
            if reference not in references:
                references.append(reference)
        if client_code not in CLIENT_CODE_MAP and client_code not in unknown_codes:
            unknown_codes.append(client_code)

    if header_row_number > 1:
        print(f"Found header row at line {header_row_number} in {path}")

    return counts, unknown_codes, references_by_code


def read_handover_rows_from_file(path: str) -> tuple[list[HandoverRow], list[str]]:
    rows = iter_excel_rows(path)
    header_row, client_code_idx, header_row_number = find_header_row(rows, "Client Code")
    reference_idx = find_column_index(header_row, "Reference")
    if reference_idx is None:
        raise ValueError(f"Reference column not found in {path}")

    headers = [str(value).strip() if value not in (None, "") else "" for value in header_row]
    items: list[HandoverRow] = []
    unknown_codes: list[str] = []

    for offset, row in enumerate(rows, start=1):
        row_number = header_row_number + offset
        raw_client_code = cell_value(row, client_code_idx)
        if raw_client_code in (None, ""):
            continue

        client_code = str(raw_client_code).strip().upper()
        if not client_code or normalize_header(client_code) == "clientcode":
            continue

        client_id = CLIENT_CODE_MAP.get(client_code)
        if not client_id:
            if client_code not in unknown_codes:
                unknown_codes.append(client_code)
            continue

        values_by_header: dict[str, object] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            values_by_header[header] = cell_value(row, idx)

        items.append(
            HandoverRow(
                source_path=path,
                row_number=row_number,
                headers=headers,
                row_values=tuple(row),
                values_by_header=values_by_header,
                client_code=client_code,
                client_id=client_id,
                reference=normalize_reference(cell_value(row, reference_idx)),
            )
        )

    return items, unknown_codes


def read_handover_rows(paths: list[str]) -> tuple[list[HandoverRow], list[str]]:
    all_rows: list[HandoverRow] = []
    unknown_codes: list[str] = []
    for path in paths:
        file_rows, file_unknown_codes = read_handover_rows_from_file(path)
        all_rows.extend(file_rows)
        for code in file_unknown_codes:
            if code not in unknown_codes:
                unknown_codes.append(code)
    return all_rows, unknown_codes


def clean_handover_files(paths: list[str], download_dir: str, cleaned_dir: str) -> list[str]:
    cleaned_paths: list[str] = []

    for source_path in paths:
        rel_path = os.path.relpath(source_path, download_dir)
        destination_path = os.path.join(cleaned_dir, rel_path)
        os.makedirs(os.path.dirname(destination_path), exist_ok=True)
        clean_handover_file(source_path, destination_path)
        cleaned_paths.append(destination_path)
        print(f"Cleaned handover file: {source_path} -> {destination_path}")

    return cleaned_paths


def clean_handover_file(source_path: str, destination_path: str) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        shutil.copy2(source_path, destination_path)
        print("Cleaning skipped: openpyxl is not installed; copied file without changes.", file=sys.stderr)
        return

    workbook = load_workbook(source_path)
    try:
        if not workbook.worksheets:
            workbook.save(destination_path)
            return

        for worksheet in workbook.worksheets:
            header_values: tuple[object, ...] | None = None
            header_row_number = None

            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row), values_only=True),
                start=1,
            ):
                row_values = tuple(row)
                if find_column_index(row_values, "Client Code") is not None:
                    header_values = row_values
                    header_row_number = row_number
                    break

            if not header_values or not header_row_number:
                continue

            reference_idx = find_column_index(header_values, "Reference")
            account_idx = find_column_index(header_values, "Account number")
            if reference_idx is None or account_idx is None:
                continue

            for row_number in range(header_row_number + 1, worksheet.max_row + 1):
                reference_cell = worksheet.cell(row=row_number, column=reference_idx + 1)
                account_cell = worksheet.cell(row=row_number, column=account_idx + 1)

                if account_cell.value != reference_cell.value:
                    account_cell.value = reference_cell.value

                cleaned_account = digits_only(account_cell.value)
                if cleaned_account != account_cell.value:
                    account_cell.value = cleaned_account

        workbook.save(destination_path)
    finally:
        workbook.close()


def iter_excel_rows(path: str):
    if path.lower().endswith(".xlsx"):
        try:
            yield from iter_xlsx_rows_stdlib(path)
            return
        except (zipfile.BadZipFile, KeyError, ET.ParseError, ValueError):
            pass

    try:
        from openpyxl import load_workbook
    except ImportError:
        yield from iter_xlsx_rows_stdlib(path)
        return

    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        if not workbook.worksheets:
            return
        worksheet = workbook.worksheets[0]
        for row in worksheet.iter_rows(values_only=True):
            yield tuple(row)
    finally:
        workbook.close()


def iter_xlsx_rows_stdlib(path: str):
    with zipfile.ZipFile(path) as workbook_zip:
        shared_strings = read_shared_strings(workbook_zip)
        sheet_path = first_sheet_path(workbook_zip)
        sheet_root = ET.fromstring(workbook_zip.read(sheet_path))
        for row in sheet_root.findall("a:sheetData/a:row", XLSX_NS):
            yield read_xlsx_row(row, shared_strings)


def read_shared_strings(workbook_zip: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in workbook_zip.namelist():
        return []

    root = ET.fromstring(workbook_zip.read("xl/sharedStrings.xml"))
    values: list[str] = []
    for item in root.findall("a:si", XLSX_NS):
        text = "".join(node.text or "" for node in item.findall(".//a:t", XLSX_NS))
        values.append(text)
    return values


def first_sheet_path(workbook_zip: zipfile.ZipFile) -> str:
    workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
    first_sheet = workbook_root.find("a:sheets/a:sheet", XLSX_NS)
    if first_sheet is None:
        raise ValueError("Workbook contains no sheets")

    rel_id = first_sheet.attrib.get(
        "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    )
    if not rel_id:
        raise ValueError("Workbook sheet relationship missing")

    rel_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
    target = None
    for rel in rel_root:
        if rel.attrib.get("Id") == rel_id:
            target = rel.attrib.get("Target")
            break
    if not target:
        raise ValueError("Workbook sheet target missing")
    if target.startswith("xl/"):
        return target
    return f"xl/{target.lstrip('/')}"


def read_xlsx_row(row_node: ET.Element, shared_strings: list[str]) -> tuple[object, ...]:
    values: list[object] = []
    last_col = 0
    for cell in row_node.findall("a:c", XLSX_NS):
        ref = cell.attrib.get("r", "A1")
        col_idx = column_index(ref)
        while last_col + 1 < col_idx:
            values.append(None)
            last_col += 1
        values.append(read_xlsx_cell(cell, shared_strings))
        last_col = col_idx
    return tuple(values)


def read_xlsx_cell(cell_node: ET.Element, shared_strings: list[str]) -> object | None:
    cell_type = cell_node.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell_node.findall(".//a:t", XLSX_NS))

    value_node = cell_node.find("a:v", XLSX_NS)
    if value_node is None:
        return None

    value = value_node.text
    if cell_type == "s" and value is not None:
        return shared_strings[int(value)]
    if cell_type == "b":
        return value == "1"
    return value


def column_index(cell_ref: str) -> int:
    letters = ""
    for char in cell_ref:
        if char.isalpha():
            letters += char
        else:
            break

    result = 0
    for char in letters.upper():
        result = (result * 26) + (ord(char) - 64)
    return result


def find_latest_fileref(matters: list[dict], prefix: str) -> tuple[str | None, str]:
    best_number = 0
    best_ref = None
    best_width = 4
    pattern = re.compile(rf"^{re.escape(prefix)}/(\d+)$", re.IGNORECASE)

    for matter in matters:
        file_ref = str(matter.get("fileref") or "").strip()
        match = pattern.match(file_ref)
        if not match:
            continue
        digits = match.group(1)
        number = int(digits)
        if number >= best_number:
            best_number = number
            best_ref = file_ref
            best_width = max(4, len(digits))

    next_ref = f"{prefix}/{best_number + 1:0{best_width}d}"
    return best_ref, next_ref


def next_ref_sequence_start(next_ref: str) -> tuple[int, int]:
    match = re.search(r"/(\d+)$", next_ref)
    if not match:
        return 1, 4
    digits = match.group(1)
    return int(digits), max(4, len(digits))


def get_row_value(row: HandoverRow, header_name: str) -> object | None:
    expected_key = normalize_header(header_name)
    for key, value in row.values_by_header.items():
        if normalize_header(key) == expected_key:
            return value
    return None


def build_debtor_name(row: HandoverRow) -> str | None:
    surname = normalize_cell_value(get_row_value(row, "Debtor Surname"))
    first_name = normalize_cell_value(get_row_value(row, "Debtor First Name"))
    name_parts = [str(value) for value in (surname, first_name) if value not in (None, "")]
    if name_parts:
        return " ".join(name_parts)
    return None


def build_description(row: HandoverRow) -> str:
    debtor_name = build_debtor_name(row)
    if debtor_name:
        return debtor_name
    matter_description = normalize_cell_value(get_row_value(row, "Matter Description"))
    if matter_description:
        return str(matter_description)
    if row.reference:
        return row.reference
    return "New Matter"


def build_party_prefix(row: HandoverRow) -> str:
    surname = normalize_cell_value(get_row_value(row, "Debtor Surname"))
    base_text = str(surname or build_description(row) or "PTY")
    letters = "".join(ch for ch in base_text.upper() if ch.isalpha())
    prefix = (letters[:3] or "PTY").ljust(3, "X")

    suffix_source = digits_only(row.reference) or digits_only(get_row_value(row, "ID Number")) or str(row.row_number)
    suffix = suffix_source[-3:].rjust(3, "0")
    return f"{prefix}{suffix}"


def build_matter_create_payload(
    row: HandoverRow,
    file_ref: str,
    date_ctx: DateContext,
    logged_in_employee_id: str,
) -> dict[str, object]:
    now = dt.datetime.now()
    payload: dict[str, object] = {}
    payload.update(CREATE_DEFAULTS)
    payload["clientid"] = row.client_id
    payload["fileref"] = file_ref
    payload["description"] = build_description(row)
    payload["dateinstructed"] = encode_legalsuite_date(dt.datetime.strptime(date_ctx.date_str, "%Y%m%d"))
    payload["updatedbydate"] = encode_legalsuite_date(now)
    payload["updatedbytime"] = encode_legalsuite_time(now)
    payload["partymatterprefix"] = row.client_code
    payload["internalcomment"] = (
        f"Imported from handover file on "
        f"{dt.datetime.strptime(date_ctx.date_str, '%Y%m%d').strftime('%d %B %Y')}"
    )
    payload["loggedinemployeeid"] = logged_in_employee_id

    for field_name, header_name in CREATE_FIELD_COLUMN_MAP.items():
        raw_value = get_row_value(row, header_name)
        if raw_value in (None, ""):
            continue
        value = normalize_reference(raw_value) if field_name == "theirref" else normalize_cell_value(raw_value)
        if value not in (None, ""):
            payload[field_name] = value

    if row.reference:
        payload["theirref"] = row.reference

    claim_amount = normalize_money(get_row_value(row, "Claim Amount"))
    if claim_amount is not None:
        payload["claimamount"] = claim_amount
        payload["debtorsbalance"] = claim_amount
        payload["debtorsopeningbalance"] = claim_amount
        payload["interestonamount"] = claim_amount
        payload["debtorscapitalbalance"] = claim_amount

    return {key: value for key, value in payload.items() if value not in (None, "")}


def build_matter_description_update_payload(
    create_payload: dict[str, object],
    logged_in_employee_id: str,
) -> dict[str, object]:
    payload = {
        "clientid": create_payload.get("clientid"),
        "archivestatusdescription": "Live",
        "loggedinemployeeid": logged_in_employee_id,
        "mattertypeid": create_payload.get("mattertypeid", CREATE_DEFAULTS["mattertypeid"]),
        "clientfeesheetid": create_payload.get("clientfeesheetid", CREATE_DEFAULTS["clientfeesheetid"]),
        "docgenid": create_payload.get("docgenid", CREATE_DEFAULTS["docgenid"]),
        "archiveflag": 0,
        "archivestatus": 0,
        "archiveno": 0,
        "description": create_payload.get("description"),
        "claimamount": create_payload.get("claimamount"),
        "debtorsbalance": create_payload.get("debtorsbalance"),
        "debtorsopeningbalance": create_payload.get("debtorsopeningbalance"),
        "interestonamount": create_payload.get("interestonamount"),
        "debtorscapitalbalance": create_payload.get("debtorscapitalbalance"),
    }
    return {key: value for key, value in payload.items() if value not in (None, "")}


def add_payload_value(payload: dict[str, object], field_name: str, value: object) -> None:
    normalized = normalize_cell_value(value)
    if normalized not in (None, ""):
        payload[field_name] = normalized


def build_party_create_payload(
    row: HandoverRow,
    date_ctx: DateContext,
    logged_in_employee_id: str,
) -> dict[str, object]:
    now = dt.datetime.now()
    name = build_debtor_name(row) or build_description(row)
    imported_date = dt.datetime.strptime(date_ctx.date_str, "%Y%m%d")
    notes = f"Imported on {imported_date.strftime('%d %B %Y')}"

    payload: dict[str, object] = {}
    payload.update(PARTY_DEFAULTS)
    payload["party[name]"] = name
    payload["party[matterprefix]"] = build_party_prefix(row)
    payload["party[updatedbydate]"] = encode_legalsuite_date(now)
    payload["party[updatedbytime]"] = encode_legalsuite_time(now)
    payload["party[createdid]"] = logged_in_employee_id
    payload["party[notes]"] = notes

    identity_number = digits_only(get_row_value(row, "ID Number"))
    if identity_number:
        payload["party[identitynumber]"] = identity_number

    add_payload_value(payload, "party[parlang][salutation]", get_row_value(row, "Debtor Title"))

    add_payload_value(payload, "party[parlang][physicalline1]", get_row_value(row, "Physical Address Line 1"))
    add_payload_value(payload, "party[parlang][physicalline2]", get_row_value(row, "Physical Address Line 2"))
    add_payload_value(payload, "party[parlang][physicalline3]", get_row_value(row, "Physical Address Line 3"))
    add_payload_value(payload, "party[parlang][physicalcode]", get_row_value(row, "Physical Postal Code"))
    add_payload_value(payload, "party[parlang][postalline1]", get_row_value(row, "Postal Address Line 1"))
    add_payload_value(payload, "party[parlang][postalline2]", get_row_value(row, "Postal Address Line 2"))
    add_payload_value(payload, "party[parlang][postalline3]", get_row_value(row, "Postal Address Line 3"))
    add_payload_value(payload, "party[parlang][postalcode]", get_row_value(row, "Postal Code"))
    add_payload_value(payload, "party[parlang][homenumber]", get_row_value(row, "Telephone (Home)"))
    add_payload_value(payload, "party[parlang][worknumber]", get_row_value(row, "Telephone (Work)"))
    add_payload_value(payload, "party[parlang][cellnumber]", get_row_value(row, "Cell Phone"))
    add_payload_value(payload, "party[parlang][emailaddress]", get_row_value(row, "DefendantEmail"))

    return {key: value for key, value in payload.items() if value not in (None, "")}


def build_matparty_create_payload(matterid: int | str, partyid: int | str) -> dict[str, object]:
    return {
        "matterid": matterid,
        "partyid": partyid,
        **MATPARTY_DEFAULTS,
    }


def header_has_date_semantics(header_name: str) -> bool:
    return "date" in normalize_header(header_name)


def build_desktop_extrascreen_payloads(row: HandoverRow) -> list[tuple[str, dict[str, object]]]:
    payloads: list[tuple[str, dict[str, object]]] = []
    current_screen_label: str | None = None
    current_payload: dict[str, object] | None = None

    def flush_current() -> None:
        nonlocal current_screen_label, current_payload
        if current_screen_label and current_payload and len(current_payload) > 1:
            payloads.append((current_screen_label, current_payload))
        current_screen_label = None
        current_payload = None

    for header_name, raw_value in zip(row.headers, row.row_values):
        if not header_name:
            continue

        screen_match = re.fullmatch(r"DesktopExtraScreenID(\d+)", header_name.strip(), re.IGNORECASE)
        if screen_match:
            flush_current()
            screen_id = normalize_reference(raw_value)
            if screen_id:
                current_screen_label = f"DesktopExtraScreenID{screen_match.group(1)}"
                current_payload = {"docscreenid": screen_id}
            continue

        if current_payload is None:
            continue

        field_match = re.match(r"Desktop Extra Field\s+(\d+)\b", header_name.strip(), re.IGNORECASE)
        if not field_match:
            continue

        field_number = int(field_match.group(1))
        field_name = f"field{field_number}"
        if raw_value in (None, ""):
            continue

        if header_has_date_semantics(header_name):
            encoded = encode_legalsuite_date(raw_value)
            if encoded is None:
                continue
            current_payload[field_name] = encoded
        else:
            normalized = normalize_cell_value(raw_value)
            if normalized in (None, ""):
                continue
            current_payload[field_name] = normalized

    flush_current()
    return payloads


def extract_created_recordid(created_response: dict | str) -> str:
    if not isinstance(created_response, dict):
        raise ValueError(f"Unexpected create response: {created_response}")
    data = created_response.get("data")
    if not isinstance(data, list) or not data:
        raise ValueError(f"Create response has no data row: {created_response}")
    recordid = data[0].get("recordid")
    if not recordid:
        raise ValueError(f"Create response has no recordid: {created_response}")
    return str(recordid)


def extract_fetched_row(fetched_response: dict | str) -> dict:
    if not isinstance(fetched_response, dict):
        raise ValueError(f"Unexpected fetch response: {fetched_response}")
    data = fetched_response.get("data")
    if not isinstance(data, list) or not data:
        raise ValueError(f"Fetch response has no data row: {fetched_response}")
    return data[0]


def fetch_matter_row(
    client: LegalSuiteLookupClient,
    recordid: int | str,
) -> dict:
    fetched = client.get_matter_by_recordid(recordid)
    return extract_fetched_row(fetched)


def fetch_matter_claimamount(
    client: LegalSuiteLookupClient,
    recordid: int | str,
) -> object | None:
    matter_row = fetch_matter_row(client, recordid)
    return matter_row.get("claimamount")


def normalize_compare_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def find_changed_fields(
    before: dict,
    after: dict,
    field_names: list[str],
) -> list[tuple[str, object, object]]:
    changes: list[tuple[str, object, object]] = []
    for field_name in field_names:
        before_value = before.get(field_name)
        after_value = after.get(field_name)
        if normalize_compare_value(before_value) != normalize_compare_value(after_value):
            changes.append((field_name, before_value, after_value))
    return changes


def compare_extrascreen_payload_to_row(payload: dict[str, object], fetched_row: dict) -> list[tuple[str, object, object]]:
    field_names = sorted(key for key in payload if key.startswith("field"))
    mismatches: list[tuple[str, object, object]] = []
    for field_name in field_names:
        sent_value = payload.get(field_name)
        fetched_value = fetched_row.get(field_name)
        if normalize_compare_value(sent_value) != normalize_compare_value(fetched_value):
            mismatches.append((field_name, sent_value, fetched_value))
    return mismatches


def describe_extrascreen_field_values(row: dict, field_names: list[str]) -> str:
    parts: list[str] = []
    for field_name in field_names:
        value = row.get(field_name)
        if normalize_compare_value(value):
            parts.append(f"{field_name}={value}")
    return ", ".join(parts) if parts else "<all blank>"


def compare_fields(sent_data: dict, returned_data: dict) -> dict[str, dict[str, object]]:
    missing_or_changed: dict[str, dict[str, object]] = {}

    for key, sent_value in sent_data.items():
        returned_value = returned_data.get(key)
        if str(returned_value) != str(sent_value):
            missing_or_changed[key] = {
                "sent": sent_value,
                "returned": returned_value,
            }

    return missing_or_changed


def find_existing_matter_for_row(
    client: LegalSuiteLookupClient,
    row: HandoverRow,
    file_ref: str,
) -> dict | None:
    fileref_matches = client.get_matters_by_fileref(file_ref)
    if fileref_matches:
        return fileref_matches[0]

    if row.reference:
        reference_matches = client.get_matters_by_clientid_and_reference(row.client_id, row.reference)
        if reference_matches:
            return reference_matches[0]

    return None


def update_handover_row_desktop_extrascreens(
    client: LegalSuiteLookupClient,
    row: HandoverRow,
    matter_recordid: int | str,
    dry_run: bool,
) -> None:
    payloads = build_desktop_extrascreen_payloads(row)
    if not payloads:
        print("  No desktop extrascreen data to update.")
        return

    for screen_label, payload in payloads:
        docscreenid = payload["docscreenid"]
        field_names = sorted(key for key in payload if key.startswith("field"))
        print(
            f"  Processing {screen_label}: docscreenid={docscreenid} | "
            f"fields={len(field_names)}"
        )
        if dry_run:
            print(f"  Dry-run: would update {screen_label}.")
            continue

        update_payload = {"matterid": matter_recordid, **payload}
        client.update_matter_extrascreen(update_payload)
        print(f"  Updated {screen_label}")
        fetched_rows = client.get_matter_extrascreen(matter_recordid, docscreenid)
        if not fetched_rows:
            print(f"  No data returned for {screen_label} after update.")
            continue

        fetched_row = fetched_rows[0]
        mismatches = compare_extrascreen_payload_to_row(payload, fetched_row)
        if mismatches:
            field_names = ", ".join(field_name for field_name, _, _ in mismatches)
            print(f"  {screen_label} mismatched fields after update: {field_names}")
            for field_name, sent_value, fetched_value in mismatches:
                print(f"    {field_name}: sent={sent_value!r} fetched={fetched_value!r}")
            print(
                "  Returned extrascreen values: "
                f"{describe_extrascreen_field_values(fetched_row, sorted(key for key in payload if key.startswith('field')))}"
            )
        else:
            updated_fields = sorted(key for key in payload if key.startswith("field"))
            if updated_fields:
                print(f"  {screen_label} fields verified after update: {', '.join(updated_fields)}")
                print(
                    "  Returned extrascreen values: "
                    f"{describe_extrascreen_field_values(fetched_row, updated_fields)}"
                )
            else:
                print(f"  {screen_label} had no field values to verify.")


def create_and_update_handover_matters(
    rows: list[HandoverRow],
    next_refs_by_code: dict[str, str],
    client: LegalSuiteLookupClient,
    date_ctx: DateContext,
    logged_in_employee_id: str,
    create_matters: bool,
    create_limit: int | None,
) -> None:
    next_numbers: dict[str, tuple[int, int]] = {
        code: next_ref_sequence_start(next_ref)
        for code, next_ref in next_refs_by_code.items()
    }

    processed_count = 0
    print("\nMatter create/update:")
    for row in rows:
        if create_limit is not None and processed_count >= create_limit:
            print(f"Create limit reached: {create_limit}")
            break
        processed_count += 1

        next_number, width = next_numbers[row.client_code]
        file_ref = f"{row.client_code}/{next_number:0{width}d}"
        next_numbers[row.client_code] = (next_number + 1, width)
        payload = build_matter_create_payload(row, file_ref, date_ctx, logged_in_employee_id)
        party_payload = build_party_create_payload(row, date_ctx, logged_in_employee_id)

        print(
            f"- {row.client_code} row {row.row_number} | reference {row.reference or ''} | "
            f"new fileref {file_ref}"
        )

        print("  Checking for existing matter...")
        existing_matter = find_existing_matter_for_row(client, row, file_ref)
        if existing_matter:
            existing_recordid = str(existing_matter.get("recordid"))
            print(
                "  Skipped: matter already exists "
                f"recordid {existing_recordid} "
                f"fileref {existing_matter.get('fileref')} "
                f"theirref {existing_matter.get('theirref')}"
            )
            update_handover_row_desktop_extrascreens(client, row, existing_recordid, dry_run=not create_matters)
            continue

        if not create_matters:
            print("  Dry-run: would create matter.")
            print("  Dry-run: would update matter description.")
            print("  Dry-run: would create or reuse party.")
            print("  Dry-run: would create MatParty link if missing.")
            update_handover_row_desktop_extrascreens(client, row, "<created-matter-recordid>", dry_run=True)
            continue

        print("  Creating matter...")
        created = client.create_matter(payload)
        recordid = extract_created_recordid(created)
        print(f"  Created matter recordid: {recordid}")
        matter_after_create = fetch_matter_row(client, recordid)
        claimamount_after_create = matter_after_create.get("claimamount")
        print(f"  Matter claimamount after create: {claimamount_after_create}")

        print("  Updating matter description...")
        update_payload = build_matter_description_update_payload(payload, logged_in_employee_id)
        client.update_matter(recordid, update_payload)
        print("  Updated matter description")
        matter_after_update = fetch_matter_row(client, recordid)
        claimamount_after_update = matter_after_update.get("claimamount")
        print(f"  Matter claimamount after description update: {claimamount_after_update}")
        tracked_fields = sorted(set(payload) | {"claimamount", "debtorsbalance", "debtorsopeningbalance", "interestonamount", "debtorscapitalbalance"})
        changed_after_update = find_changed_fields(matter_after_create, matter_after_update, tracked_fields)
        if changed_after_update:
            changed_names = ", ".join(field_name for field_name, _, _ in changed_after_update)
            print(f"  Fields changed after description update: {changed_names}")
        else:
            print("  No tracked matter fields changed after description update.")

        create_parties = True
        if not create_parties:
            print("  Party and MatParty creation skipped for preview.")
            continue

        print("  Checking for existing party...")
        identity_number = digits_only(get_row_value(row, "ID Number"))
        existing_party = None
        if identity_number:
            matches = client.get_party_by_identitynumber(identity_number)
            if matches:
                existing_party = matches[0]

        if existing_party:
            partyid = str(existing_party["recordid"])
            print(f"  Reusing existing partyid: {partyid}")
        else:
            print("  Creating party...")
            created_party = client.create_party(party_payload)
            partyid = extract_created_recordid(created_party)
            print(f"  Created partyid: {partyid}")

        print("  Checking MatParty link...")
        existing_matparty = client.get_matparty_by_matter_and_party(recordid, partyid)
        if existing_matparty:
            print(
                "  MatParty link already exists: "
                f"recordid {existing_matparty[0].get('recordid')}"
            )
        else:
            print("  Creating MatParty link...")
            matparty_payload = build_matparty_create_payload(recordid, partyid)
            created_matparty = client.create_matparty(matparty_payload)
            matparty_recordid = extract_created_recordid(created_matparty)
            print(f"  Linked matparty recordid: {matparty_recordid}")

        matter_after_matparty = fetch_matter_row(client, recordid)
        claimamount_after_matparty = matter_after_matparty.get("claimamount")
        print(f"  Matter claimamount after MatParty step: {claimamount_after_matparty}")
        changed_after_matparty = find_changed_fields(matter_after_update, matter_after_matparty, tracked_fields)
        if changed_after_matparty:
            changed_names = ", ".join(field_name for field_name, _, _ in changed_after_matparty)
            print(f"  Fields changed after MatParty step: {changed_names}")
        else:
            print("  No tracked matter fields changed after MatParty step.")

        update_handover_row_desktop_extrascreens(client, row, recordid, dry_run=False)


def process_handover_files(
    paths: list[str],
    api_base: str,
    api_key: str,
    date_ctx: DateContext,
    create_matters: bool,
    create_dry_run: bool,
    create_limit: int | None,
    logged_in_employee_id: str,
) -> int:
    code_counts: dict[str, int] = {}
    unknown_codes: list[str] = []
    code_references: dict[str, list[str]] = {}
    next_refs_by_code: dict[str, str] = {}

    for path in paths:
        file_counts, file_unknown_codes, file_references = read_client_codes_from_file(path)
        total_rows = sum(file_counts.values())
        print(f"Processed {path}: {total_rows} row(s), {len(file_counts)} unique client code(s)")
        for code, count in file_counts.items():
            code_counts[code] = code_counts.get(code, 0) + count
            refs = code_references.setdefault(code, [])
            for reference in file_references.get(code, []):
                if reference not in refs:
                    refs.append(reference)
        for code in file_unknown_codes:
            if code not in unknown_codes:
                unknown_codes.append(code)

    if not code_counts:
        print("No Client Code values found in the downloaded handover files.")
        return 0

    if unknown_codes:
        print("Unknown client codes in Excel:", ", ".join(sorted(unknown_codes)))

    client = LegalSuiteLookupClient(api_base, api_key)
    print("\nLegalSuite latest matter lookup:")
    for client_code in sorted(code_counts):
        client_id = CLIENT_CODE_MAP.get(client_code)
        if not client_id:
            print(f"- {client_code}: no client-id mapping found")
            continue

        matters = client.get_matters_by_clientid_and_prefix(client_id, client_code)
        latest_ref, next_ref = find_latest_fileref(matters, client_code)
        next_refs_by_code[client_code] = next_ref
        if latest_ref:
            print(
                f"- {client_code} -> clientid {client_id} | rows {code_counts[client_code]} | "
                f"latest {latest_ref} | next {next_ref}"
            )
        else:
            print(
                f"- {client_code} -> clientid {client_id} | rows {code_counts[client_code]} | "
                f"latest none | next {next_ref}"
            )
        references = code_references.get(client_code, [])
        if references:
            print(f"  Reference values: {', '.join(references)}")

    if create_matters or create_dry_run:
        rows, row_unknown_codes = read_handover_rows(paths)
        for code in row_unknown_codes:
            if code not in unknown_codes:
                unknown_codes.append(code)
        if unknown_codes:
            print("Rows skipped for unknown client codes:", ", ".join(sorted(unknown_codes)))
        create_and_update_handover_matters(
            rows=rows,
            next_refs_by_code=next_refs_by_code,
            client=client,
            date_ctx=date_ctx,
            logged_in_employee_id=logged_in_employee_id,
            create_matters=create_matters,
            create_limit=create_limit,
        )

    return 0


def normalize_cli_args(argv: list[str]) -> list[str]:
    normalized: list[str] = []
    for arg in argv:
        match = re.fullmatch(r"--days-(\d+)", arg)
        if match:
            normalized.extend(["--days-ago", match.group(1)])
            continue
        normalized.append(arg)
    return normalized


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    if argv is None:
        argv = sys.argv[1:]

    parser = argparse.ArgumentParser(
        description="Download handover files and display the next LegalSuite file ref per Client Code."
    )
    parser.add_argument("--date", help="Date in YYYYMMDD format.")
    parser.add_argument(
        "--days-ago",
        type=int,
        default=0,
        help="Download files from N days ago (default: 0). Also accepts shorthand like --days-15. Ignored if --date is provided.",
    )
    parser.add_argument(
        "--download-dir",
        default="downloads",
        help="Local base directory for downloads (default: downloads).",
    )
    parser.add_argument(
        "--cleaned-dir",
        default="cleaned_handover",
        help="Local base directory for cleaned handover files (default: cleaned_handover).",
    )
    parser.add_argument(
        "--clean-only",
        action="store_true",
        help="Skip FTP and use already-downloaded handover files for the selected date.",
    )
    parser.add_argument(
        "--skip-clean",
        action="store_true",
        help="Use the selected handover files directly without creating cleaned working copies.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="FTP connection timeout in seconds (default: 30).",
    )
    parser.add_argument(
        "--api-base",
        default=LEGALSUITE_API_BASE,
        help="LegalSuite API base URL.",
    )
    parser.add_argument(
        "--api-key",
        default=os.getenv("LEGALSUITE_API_KEY") or LEGALSUITE_API_KEY,
        help="LegalSuite API key.",
    )
    parser.add_argument(
        "--logged-in-employee-id",
        default=str(CREATE_DEFAULTS["loggedinemployeeid"]),
        help="LegalSuite logged-in employee ID for created/updated matters (default: 1).",
    )
    parser.add_argument(
        "--create-dry-run",
        action="store_true",
        help="Build and print matter/store payloads without creating matters.",
    )
    parser.add_argument(
        "--create-matters",
        action="store_true",
        help="Create matters in LegalSuite, fetch them by recordid, then update fields that did not save.",
    )
    parser.add_argument(
        "--create-limit",
        type=int,
        help="Limit how many handover rows are created or previewed.",
    )
    return parser.parse_args(normalize_cli_args(argv))


def main() -> int:
    try:
        args = parse_args()
        date_ctx = resolve_date(args.date, args.days_ago)
        print(f"Target date: {date_ctx.date_str}")
        print(f"Panel month folder: {date_ctx.month_year}")
        if args.clean_only:
            working_files = resolve_local_handover_files(date_ctx, args.cleaned_dir, label="cleaned")
            if not working_files:
                print("No cleaned handover files were available; checking downloaded files.")
                source_files = resolve_local_handover_files(date_ctx, args.download_dir)
                if not source_files:
                    print("No handover files were available.")
                    return 1
                if args.skip_clean:
                    working_files = source_files
                else:
                    working_files = clean_handover_files(source_files, args.download_dir, args.cleaned_dir)
        else:
            source_files = download_handover_files(date_ctx, args.download_dir, args.timeout)
            if not source_files:
                print("No handover files were available.")
                return 1
            if args.skip_clean:
                working_files = source_files
            else:
                working_files = clean_handover_files(source_files, args.download_dir, args.cleaned_dir)

        return process_handover_files(
            paths=working_files,
            api_base=args.api_base,
            api_key=args.api_key,
            date_ctx=date_ctx,
            create_matters=args.create_matters,
            create_dry_run=args.create_dry_run,
            create_limit=args.create_limit,
            logged_in_employee_id=args.logged_in_employee_id,
        )
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 2
    except requests.RequestException as exc:
        print(f"LegalSuite API error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
