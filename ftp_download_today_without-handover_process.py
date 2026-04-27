#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import glob
import json
import fnmatch
import ftplib
import os
import re
import requests
import sys
import warnings
from dataclasses import dataclass
from typing import Callable

from env_config import load_env_file

load_env_file()

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - handled at runtime
    load_workbook = None
    Workbook = None
else:
    warnings.filterwarnings(
        "ignore",
        message="Workbook contains no default style, apply openpyxl's default",
        module="openpyxl.styles.stylesheet",
    )

FTP_HOST = os.getenv("FTP_HOST", "")
FTP_USER = os.getenv("FTP_USER", "")
FTP_PASS = os.getenv("FTP_PASS", "")

LEGALSUITE_API_BASE = "https://api.legalsuite.net"
LEGALSUITE_API_KEY = os.getenv("LEGALSUITE_API_KEY", "")
LEGALSUITE_EMPLOYEE_ID = "1"
LEGALSUITE_ARCHIVE_STATUS = "2"
LEGALSUITE_OFFSET = 36161
EXCEL_BASE = dt.datetime(1899, 12, 30)

TARGETS = [
    ("SBSA/Debt Review/Debt_Review_Close_APT_LSW", "Standard_Bank_Panel_L_Close_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_PTP_APT_LSW", "Standard_Bank_Panel_L_PTP_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_Feedback_APT_LSW", "Standard_Bank_Panel_L_Update_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_Reopen_APT_LSW", "Standard_Bank_Panel_L_Reopen_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_Handover_APT_LWS", "Standard_Bank_Panel_L_Handover_{date}_DR.xlsx"),
    ("SBSA/Panel L/PTP_APT_LSW/{month_year}", "*_{date}.xlsx"),
    ("SBSA/Panel L/Feedback_APT_LSW/{month_year}", "*_{date}.xlsx"),
    ("SBSA/Panel L/Handover_APT_LSW/{month_year}", "*_{date}.xlsx"),
    ("SBSA/Panel L/Closed_APT_LSW/{month_year}", "*_{date}.xlsx"),
    ("SBSA/Panel L/Reopen_APT_LSW/{month_year}", "*_{date}.xlsx"),
    ("Standard Bank_ClaimsAmount", "Standard Bank Legal Claim Amount_Panel_L{year}_{month}_{day}_*.xlsx"),
    ("SBSA POC AND SUMMONS", "{day}{month}{year}.csv"),
]

REMOTE_DIR_ALIASES = {
    "SBSA/Debt Review/Debt_Review_Handover_APT_LWS": [
        "SBSA/Debt Review/Debt_Review_ Handover_APT_LWS",
    ],
}

CLAIM_AMOUNT_PATTERN = "Standard Bank Legal Claim Amount_Panel_L*.xlsx"
HANDOVER_PREFIXES = [
    "SBSA/Debt Review/Debt_Review_Handover_APT_LWS/",
    "SBSA/Debt Review/Debt_Review_ Handover_APT_LWS/",
    "SBSA/Panel L/Handover_APT_LSW/",
]

FEEDBACK_FIELD_MAP = [
    ("accountnumber", "field1", False),
    ("ptpcapturedate", "field2", True),
    ("ptpduedate", "field3", True),
    ("ptpamount", "field4", False),
    ("lastpaymentdate", "field5", True),
    ("lastpaymentamount", "field6", False),
    ("lastquickcomment", "field7", False),
    ("lastquickcommentdate", "field8", True),
    ("lastmemo", "field9", False),
    ("lastmemodate", "field10", True),
    ("accountcloseddate", "field11", True),
    ("reasonforclosure", "field12", False),
    ("branchid", "field13", False),
]

PTP_FIELD_MAP = [
    ("accountnumber", "field1", False),
    ("ptpcapturedate", "field2", True),
    ("ptpduedate", "field3", True),
    ("ptpamount", "field4", False),
    ("lastpaymentdate", "field5", True),
    ("branchid", "field6", False),
    ("lastquickcomment", "field7", False),
    ("ptpamount2", "field8", False),
    ("ptpduedate2", "field9", True),
    ("ptpamount3", "field10", False),
    ("ptpduedate3", "field11", True),
    ("ptpamount4", "field12", False),
    ("ptpduedate4", "field13", True),
    ("ptpamount5", "field14", False),
    ("ptpduedate5", "field15", True),
    ("ptpamount6", "field16", False),
    ("ptpduedate6", "field17", True),
    ("ptpamount7", "field18", False),
    ("ptpduedate7", "field19", True),
    ("ptpamount8", "field20", False),
    ("ptpduedate8", "field21", True),
    ("ptpamount9", "field22", False),
    ("ptpduedate9", "field23", True),
    ("ptpamount10", "field24", False),
    ("ptpduedate10", "field25", True),
    ("ptpamount11", "field26", False),
    ("ptpduedate11", "field27", True),
    ("ptpamount12", "field28", False),
    ("ptpduedate12", "field29", True),
]

POC_SUMMONS_FIELD_MAP = [
    ("noofcallattempts", "field2", False),
    ("noofdispatchedsmss", "field3", False),
    ("noofdispatchedemails", "field4", False),
    ("noofbrokenptps", "field5", False),
]

@dataclass(frozen=True)
class DateContext:
    date_str: str
    month_year: str
    year: str
    month: str
    day: str


class FTPClient:
    def __init__(self, host: str, user: str, password: str, timeout: int) -> None:
        self._host = host
        self._user = user
        self._password = password
        self._timeout = timeout
        self._ftp: ftplib.FTP | None = None

    def connect(self) -> None:
        self._ftp = ftplib.FTP(self._host, timeout=self._timeout)
        self._ftp.login(self._user, self._password)
        self._ftp.set_pasv(True)

    def close(self) -> None:
        if not self._ftp:
            return
        try:
            self._ftp.quit()
        except ftplib.all_errors:
            self._ftp.close()

    def list_dir(self, remote_dir: str) -> list[str] | None:
        ftp = self._require()
        try:
            return ftp.nlst(remote_dir)
        except ftplib.error_perm:
            pass
        current_dir = None
        try:
            current_dir = ftp.pwd()
            ftp.cwd(remote_dir)
            listing = ftp.nlst()
            return [f"{remote_dir}/{name}" for name in listing]
        except ftplib.error_perm:
            return None
        finally:
            if current_dir:
                try:
                    ftp.cwd(current_dir)
                except ftplib.all_errors:
                    pass

    def mdtm_timestamp(self, remote_path: str) -> dt.datetime | None:
        ftp = self._require()
        try:
            response = ftp.sendcmd(f"MDTM {remote_path}")
        except ftplib.all_errors:
            return None
        parts = response.split()
        if len(parts) != 2 or parts[0] != "213":
            return None
        try:
            return dt.datetime.strptime(parts[1], "%Y%m%d%H%M%S")
        except ValueError:
            return None

    def select_newest_by_mdtm(self, remote_dir: str, names: list[str]) -> str:
        candidates: list[tuple[dt.datetime, str]] = []
        for name in names:
            ts = self.mdtm_timestamp(f"{remote_dir}/{name}")
            if ts:
                candidates.append((ts, name))
        if candidates:
            candidates.sort()
            return candidates[-1][1]
        names.sort()
        return names[-1]

    def resolve_remote_file(self, remote_dir: str, filename_tmpl: str) -> tuple[str | None, str | None]:
        listing = self.list_dir(remote_dir)
        if listing is None:
            return None, "missing_dir"
        names = [os.path.basename(item) for item in listing]
        if any(ch in filename_tmpl for ch in ["*", "?", "["]):
            matches = [name for name in names if fnmatch.fnmatch(name, filename_tmpl)]
            if not matches:
                return None, "missing_file"
            return self.select_newest_by_mdtm(remote_dir, matches), None
        if filename_tmpl in names:
            return filename_tmpl, None
        ftp = self._require()
        try:
            ftp.size(f"{remote_dir}/{filename_tmpl}")
            return filename_tmpl, None
        except ftplib.all_errors:
            return None, "missing_file"

    def download_file(self, remote_path: str, local_path: str) -> None:
        ftp = self._require()
        with open(local_path, "wb") as handle:
            ftp.retrbinary(f"RETR {remote_path}", handle.write)

    def _require(self) -> ftplib.FTP:
        if not self._ftp:
            raise RuntimeError("FTP client not connected")
        return self._ftp


class LegalSuiteClient:
    def __init__(self, api_base: str, api_key: str) -> None:
        self._api_base = api_base.rstrip("/")
        self._api_key = api_key

    def get_matter_by_fileref(self, file_ref: str) -> dict:
        url = f"{self._api_base}/matter/get"
        data = {
            "where[]": f"Matter.FileRef,=,{file_ref}",
        }
        resp = requests.post(url, headers=self._headers(), data=data, timeout=60)
        resp.raise_for_status()
        payload = resp.json()
        items = payload.get("data", [])
        if not items:
            raise ValueError(f"No matter found for FileRef: {file_ref}")
        return items[0]

    def update_matter(self, payload: dict) -> dict:
        url = f"{self._api_base}/matter/update"
        resp = requests.post(url, headers=self._headers(), data=payload, timeout=60)
        resp.raise_for_status()
        try:
            return resp.json()
        except ValueError:
            return {"raw_response": resp.text}

    def update_matter_extrascreen(self, payload: dict) -> dict:
        url = f"{self._api_base}/matdocsc/update"
        resp = requests.post(url, headers=self._headers(), data=payload, timeout=60)
        resp.raise_for_status()
        try:
            return resp.json()
        except ValueError:
            return {"raw_response": resp.text}

    @staticmethod
    def build_archive_payload(
        matter: dict,
        logged_in_employee_id: str,
        archive_no: str | None = None,
        archive_status: str | None = None,
    ) -> dict:
        now = dt.datetime.now()
        payload = {}
        for key, value in matter.items():
            if isinstance(value, (dict, list, tuple, set)):
                continue
            payload[key] = value
        payload.update(
            {
                "loggedinemployeeid": str(logged_in_employee_id),
                "archiveflag": "1",
                "archivestatusdescription": "Archived",
                "formattedupdatedbydate": now.strftime("%d %b %Y"),
                "formattedupdatedbytime": now.strftime("%H:%M:%S"),
            }
        )
        if archive_status is not None:
            payload["archivestatus"] = str(archive_status)
        if archive_no is not None:
            payload["archiveno"] = str(archive_no)
        return {k: v for k, v in payload.items() if v not in ("", None)}

    @staticmethod
    def build_claim_amount_payload(
        matter: dict,
        logged_in_employee_id: str,
        claim_amount: object,
    ) -> dict:
        payload = {}
        for key, value in matter.items():
            if isinstance(value, (dict, list, tuple, set)):
                continue
            payload[key] = value
        payload["loggedinemployeeid"] = str(logged_in_employee_id)
        payload["claimamount"] = claim_amount
        return {k: v for k, v in payload.items() if v not in ("", None)}

    def _headers(self) -> dict:
        return {
            "Authorization": f"Bearer {self._api_key}",
            "Content-Type": "application/x-www-form-urlencoded",
        }


class Cleaner:
    def __init__(self, handover_prefixes: list[str]) -> None:
        self._rules = {"accountnumber": self._digits_only}
        self._handover_prefixes = handover_prefixes

    def clean_downloads(
        self,
        download_dir: str,
        cleaned_dir: str,
        report_lines: list[str],
        source_paths: list[str] | None = None,
    ) -> None:
        if load_workbook is None or Workbook is None:
            report_lines.append("Cleaning skipped: openpyxl is not installed.")
            print("Cleaning skipped: openpyxl is not installed.", file=sys.stderr)
            return

        cleaned_files = 0
        converted_files = 0
        skipped_files = 0
        failed_files = 0
        total_cells = 0
        matched_columns = 0
        blanked_columns = 0
        headers_removed = 0
        copied_cells = 0
        cleaned_dir_abs = os.path.abspath(cleaned_dir)

        for source_path in self._iter_source_paths(download_dir, cleaned_dir_abs, source_paths):
            name = os.path.basename(source_path)
            lowered = name.lower()
            if not (lowered.endswith(".xlsx") or lowered.endswith(".csv")):
                continue
            if name.startswith("~$"):
                skipped_files += 1
                continue

            rel_path = os.path.relpath(source_path, download_dir)
            if lowered.endswith(".csv"):
                rel_path = os.path.splitext(rel_path)[0] + ".xlsx"
            destination_path = os.path.join(cleaned_dir, rel_path)

            blank_headers = set()
            drop_header = True
            if fnmatch.fnmatch(name, CLAIM_AMOUNT_PATTERN):
                blank_headers.add("matter")
                drop_header = False

            rel_norm = rel_path.replace(os.sep, "/")
            copy_reference = any(rel_norm.startswith(prefix) for prefix in self._handover_prefixes)

            try:
                if lowered.endswith(".csv"):
                    raw_copy_path = os.path.splitext(source_path)[0] + ".xlsx"
                    cleaned, matched, blanked, removed, copied = self._convert_csv_to_excel(
                        source_path,
                        destination_path,
                        blank_headers,
                        copy_reference,
                        drop_header,
                        raw_copy_path,
                    )
                    converted_files += 1
                    action = "Converted"
                    report_lines.append(f"Saved CSV Excel copy: {raw_copy_path}")
                else:
                    cleaned, matched, blanked, removed, copied = self._clean_excel_file(
                        source_path, destination_path, blank_headers, copy_reference, drop_header
                    )
                    cleaned_files += 1
                    action = "Cleaned"

                total_cells += cleaned
                matched_columns += matched
                blanked_columns += blanked
                headers_removed += removed
                copied_cells += copied
                report_lines.append(
                    f"{action}: {source_path} -> {destination_path} (cells updated: {cleaned})"
                )
            except Exception as exc:
                failed_files += 1
                report_lines.append(f"Clean failed: {source_path} ({exc})")
                print(f"Clean failed: {source_path}: {exc}", file=sys.stderr)

        report_lines.append(
            "Clean summary: cleaned_files={cleaned}, converted_files={converted}, "
            "skipped_files={skipped}, failed_files={failed}, "
            "columns_matched={columns}, blanked_columns={blanked}, "
            "headers_removed={headers}, copied_cells={copied}, cells_updated={cells}".format(
                cleaned=cleaned_files,
                converted=converted_files,
                skipped=skipped_files,
                failed=failed_files,
                columns=matched_columns,
                blanked=blanked_columns,
                headers=headers_removed,
                copied=copied_cells,
                cells=total_cells,
            )
        )

    @staticmethod
    def _iter_source_paths(
        download_dir: str,
        cleaned_dir_abs: str,
        source_paths: list[str] | None,
    ) -> list[str]:
        if source_paths is None:
            paths: list[str] = []
            for root, _, files in os.walk(download_dir):
                root_abs = os.path.abspath(root)
                if root_abs.startswith(cleaned_dir_abs):
                    continue
                for name in files:
                    paths.append(os.path.join(root, name))
            return paths

        unique_paths: list[str] = []
        seen: set[str] = set()
        for path in source_paths:
            path_abs = os.path.abspath(path)
            if path_abs.startswith(cleaned_dir_abs):
                continue
            if not os.path.exists(path):
                continue
            if path_abs in seen:
                continue
            seen.add(path_abs)
            unique_paths.append(path)
        return unique_paths

    def _clean_excel_file(
        self,
        source_path: str,
        destination_path: str,
        blank_headers: set[str],
        copy_reference: bool,
        drop_header: bool,
    ) -> tuple[int, int, int, int, int]:
        workbook = load_workbook(source_path)
        cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells = self._apply_cleaning(
            workbook, blank_headers, drop_header, copy_reference
        )
        os.makedirs(os.path.dirname(destination_path) or ".", exist_ok=True)
        workbook.save(destination_path)
        return cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells

    def _convert_csv_to_excel(
        self,
        source_path: str,
        destination_path: str,
        blank_headers: set[str],
        copy_reference: bool,
        drop_header: bool,
        raw_copy_path: str | None = None,
    ) -> tuple[int, int, int, int, int]:
        workbook = Workbook()
        worksheet = workbook.active
        with open(source_path, newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle)
            for row in reader:
                worksheet.append(row)
        if raw_copy_path:
            os.makedirs(os.path.dirname(raw_copy_path) or ".", exist_ok=True)
            workbook.save(raw_copy_path)
        cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells = self._apply_cleaning(
            workbook, blank_headers, drop_header, copy_reference
        )
        os.makedirs(os.path.dirname(destination_path) or ".", exist_ok=True)
        workbook.save(destination_path)
        return cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells

    def _apply_cleaning(
        self,
        workbook,
        blank_headers: set[str],
        drop_header: bool,
        copy_reference: bool,
    ) -> tuple[int, int, int, int, int]:
        cleaned_cells = 0
        matched_columns = 0
        blanked_columns = 0
        headers_removed = 0
        copied_cells = 0

        for worksheet in workbook.worksheets:
            header_row = list(worksheet.iter_rows(min_row=1, max_row=1))
            if not header_row:
                continue
            header_cells = header_row[0]
            column_rules: dict[int, Callable[[object], object]] = {}
            blank_columns: set[int] = set()
            account_col = None
            reference_col = None

            for idx, cell in enumerate(header_cells, start=1):
                key = self._normalize_header(cell.value)
                if key in self._rules:
                    column_rules[idx] = self._rules[key]
                if key in blank_headers:
                    blank_columns.add(idx)
                if key == "accountnumber":
                    account_col = idx
                if key == "reference":
                    reference_col = idx

            matched_columns += len(column_rules)
            blanked_columns += len(blank_columns)

            if column_rules or blank_columns or (copy_reference and account_col and reference_col):
                for row in worksheet.iter_rows(min_row=2):
                    if copy_reference and account_col and reference_col:
                        account_cell = row[account_col - 1]
                        reference_cell = row[reference_col - 1]
                        if account_cell.value != reference_cell.value:
                            account_cell.value = reference_cell.value
                            copied_cells += 1
                    for col_idx, rule in column_rules.items():
                        cell = row[col_idx - 1]
                        new_value = rule(cell.value)
                        if new_value != cell.value:
                            cell.value = new_value
                            cleaned_cells += 1
                    for col_idx in blank_columns:
                        cell = row[col_idx - 1]
                        if cell.value not in (None, ""):
                            cell.value = None
                            cleaned_cells += 1

            if drop_header and worksheet.max_row >= 1:
                worksheet.delete_rows(1)
                headers_removed += 1

        return cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells

    @staticmethod
    def _normalize_header(header: object) -> str:
        if header is None:
            return ""
        return "".join(ch for ch in str(header).strip().lower() if ch.isalnum())

    @staticmethod
    def _digits_only(value: object) -> object:
        if value is None:
            return None
        text = str(value)
        return "".join(ch for ch in text if ch.isdigit())


class App:
    def __init__(self, args: argparse.Namespace) -> None:
        self._args = args
        self._date_ctx = self._resolve_date(args.date, args.days_ago)
        self._targets = self._build_targets(self._date_ctx)
        self._cleaner = Cleaner(HANDOVER_PREFIXES)

    def run(self) -> int:
        report_lines = [
            f"Report date: {self._date_ctx.date_str}",
            f"Month folder: {self._date_ctx.month_year}",
        ]

        download_counts, source_paths = self._download_files(report_lines)
        if download_counts is None:
            self._write_report(report_lines)
            return 1

        if not self._args.skip_clean:
            print("Cleaning downloaded files...")
            self._cleaner.clean_downloads(
                self._args.download_dir,
                self._args.cleaned_dir,
                report_lines,
                source_paths=source_paths,
            )

        if self._args.update_extrascreen:
            print("Updating matter extra screens...")
            self._update_matter_extrascreens(report_lines)

        if self._args.update_claim_amount:
            print("Updating claim amounts...")
            self._update_claim_amounts(report_lines)

        if self._args.archive_closed:
            print("Reading closed files and calling LegalSuite...")
            self._archive_closed_matters(report_lines)

        report_lines.append(
            "Summary: downloaded={downloaded}, reused={reused}, missing_dirs={missing_dirs}, "
            "missing_files={missing_files}, failed_downloads={failed}".format(
                downloaded=download_counts["downloaded"],
                reused=download_counts["reused"],
                missing_dirs=download_counts["missing_dirs"],
                missing_files=download_counts["missing_files"],
                failed=download_counts["failed_downloads"],
            )
        )

        self._write_report(report_lines)
        return 0

    def _download_files(self, report_lines: list[str]) -> tuple[dict[str, int] | None, list[str]]:
        counts = {
            "downloaded": 0,
            "reused": 0,
            "missing_dirs": 0,
            "missing_files": 0,
            "failed_downloads": 0,
        }
        source_paths: list[str] = []

        if self._args.clean_only:
            report_lines.append("Download skipped: clean-only mode.")
            source_paths = self._resolve_existing_files(report_lines, counts)
            return counts, source_paths

        print("Connecting to FTP...")
        client = FTPClient(FTP_HOST, FTP_USER, FTP_PASS, self._args.timeout)
        try:
            client.connect()
        except ftplib.all_errors as exc:
            error_line = f"FTP connection failed: {exc}"
            print(error_line, file=sys.stderr)
            report_lines.append(error_line)
            return None, []

        try:
            print("Downloading files...")
            for remote_dir, filename in self._targets:
                resolved_dir = None
                resolved_name = None
                reason = "missing_dir"
                for candidate_dir in self._target_dirs(remote_dir):
                    resolved_name, reason = client.resolve_remote_file(candidate_dir, filename)
                    if resolved_name:
                        resolved_dir = candidate_dir
                        break

                if not resolved_name or not resolved_dir:
                    if reason == "missing_dir":
                        line = f"Missing directory: {remote_dir}"
                        counts["missing_dirs"] += 1
                    else:
                        line = f"Missing file: {remote_dir}/{filename}"
                        counts["missing_files"] += 1
                    print(line)
                    report_lines.append(line)
                    continue

                remote_path = f"{resolved_dir}/{resolved_name}"
                local_path = self._ensure_local_path(self._args.download_dir, resolved_dir, resolved_name)
                try:
                    client.download_file(remote_path, local_path)
                    line = f"Downloaded: {remote_path} -> {local_path}"
                    print(line)
                    report_lines.append(line)
                    counts["downloaded"] += 1
                    source_paths.append(local_path)
                except ftplib.all_errors as exc:
                    line = f"Download failed for {remote_path}: {exc}"
                    print(line, file=sys.stderr)
                    report_lines.append(line)
                    counts["failed_downloads"] += 1
        finally:
            client.close()

        return counts, source_paths

    def _resolve_existing_files(
        self,
        report_lines: list[str],
        counts: dict[str, int],
    ) -> list[str]:
        source_paths: list[str] = []
        for remote_dir, filename in self._targets:
            local_path = self._resolve_existing_local_file(remote_dir, filename)
            if local_path:
                line = f"Using existing file: {local_path}"
                print(line)
                report_lines.append(line)
                counts["reused"] += 1
                source_paths.append(local_path)
                continue

            line = f"Existing file not found: {remote_dir}/{filename}"
            print(line)
            report_lines.append(line)
            counts["missing_files"] += 1
        return source_paths

    def _resolve_existing_local_file(self, remote_dir: str, filename: str) -> str | None:
        for candidate_dir in self._target_dirs(remote_dir):
            local_dir = os.path.join(self._args.download_dir, *candidate_dir.split("/"))
            if not os.path.isdir(local_dir):
                continue

            try:
                names = sorted(
                    name for name in os.listdir(local_dir) if os.path.isfile(os.path.join(local_dir, name))
                )
            except OSError:
                continue

            if any(ch in filename for ch in ["*", "?", "["]):
                matches = [name for name in names if fnmatch.fnmatch(name, filename)]
                if matches:
                    return os.path.join(local_dir, matches[-1])
                continue

            if filename in names:
                return os.path.join(local_dir, filename)
        return None

    @staticmethod
    def _target_dirs(remote_dir: str) -> list[str]:
        dirs = [remote_dir]
        dirs.extend(REMOTE_DIR_ALIASES.get(remote_dir, []))
        seen: set[str] = set()
        unique_dirs: list[str] = []
        for item in dirs:
            if item in seen:
                continue
            seen.add(item)
            unique_dirs.append(item)
        return unique_dirs

    def _write_report(self, report_lines: list[str]) -> None:
        log_path = self._args.log_file
        if not log_path:
            log_path = os.path.join(self._args.download_dir, f"report_{self._date_ctx.date_str}.txt")
        os.makedirs(os.path.dirname(log_path) or ".", exist_ok=True)
        try:
            print("Writing report log...")
            with open(log_path, "w", encoding="utf-8") as handle:
                handle.write("\n".join(report_lines) + "\n")
        except OSError as exc:
            print(f"Failed to write log file {log_path}: {exc}", file=sys.stderr)

    def _update_matter_extrascreens(self, report_lines: list[str]) -> None:
        if load_workbook is None:
            report_lines.append("Extrascreen update skipped: openpyxl is not installed.")
            print("Extrascreen update skipped: openpyxl is not installed.", file=sys.stderr)
            return

        api_key = self._args.api_key or os.getenv("LEGALSUITE_API_KEY") or LEGALSUITE_API_KEY
        if not api_key:
            report_lines.append("Extrascreen update skipped: missing API key.")
            print("Extrascreen update skipped: missing API key.", file=sys.stderr)
            return

        client = LegalSuiteClient(self._args.api_base, api_key)
        feedback_files = self._resolve_feedback_files()
        ptp_files = self._resolve_ptp_files()
        poc_summons_files = self._resolve_poc_summons_files()
        if self._args.extrascreen_only:
            allowed = {self._args.extrascreen_only}
            feedback_files = self._filter_extrascreen_files(feedback_files, allowed)
            ptp_files = self._filter_extrascreen_files(ptp_files, allowed)
            poc_summons_files = self._filter_extrascreen_files(poc_summons_files, allowed)
        if not feedback_files and not ptp_files and not poc_summons_files:
            report_lines.append("Extrascreen update skipped: no matching files found.")
            print("Extrascreen update skipped: no matching files found.")
            return

        processed = 0
        updated = 0
        failed = 0

        for cleaned_path, mapping in feedback_files + ptp_files + poc_summons_files:
            if self._args.extrascreen_verbose:
                label = self._extrascreen_mapping_label(mapping)
                print(f"Extrascreen file ({label}): {cleaned_path}")
            download_path = self._download_path_for_cleaned(cleaned_path)
            header_row = self._read_header_row(download_path)
            if not header_row:
                report_lines.append(f"Header not found for extrascreen update: {cleaned_path}")
                continue

            col_map, file_ref_idx, screen_id_idx = self._build_extrascreen_column_map(header_row, mapping)
            if not file_ref_idx or not screen_id_idx:
                report_lines.append(f"Missing File Reference or Desktop Extra ScreenID: {cleaned_path}")
                continue

            workbook = load_workbook(cleaned_path, read_only=False, data_only=True)
            for worksheet in workbook.worksheets:
                max_col = worksheet.max_column or max(col_map.values(), default=0)
                for row in worksheet.iter_rows(min_row=1, max_col=max_col, values_only=True):
                    file_ref = self._cell_text(row, file_ref_idx)
                    docscreenid = self._cell_text(row, screen_id_idx)
                    if not file_ref or not docscreenid:
                        continue
                    payload = self._build_extrascreen_payload(row, col_map, mapping)
                    if not payload:
                        continue
                    try:
                        matter = client.get_matter_by_fileref(file_ref)
                        recordid = matter.get("recordid")
                        if not recordid:
                            report_lines.append(f"Extrascreen update skipped (missing recordid): {file_ref}")
                            continue
                        payload["matterid"] = recordid
                        payload["docscreenid"] = docscreenid
                        processed += 1
                        if self._args.extrascreen_dry_run:
                            print(f"Extrascreen dry-run: {file_ref} -> docscreenid={docscreenid}")
                            if self._args.extrascreen_verbose:
                                print(json.dumps(payload, indent=2, default=str))
                            continue
                        if self._args.extrascreen_verbose:
                            print(f"Extrascreen update payload for {file_ref}:")
                            print(json.dumps(payload, indent=2, default=str))
                        result = client.update_matter_extrascreen(payload)
                        if self._args.extrascreen_verbose:
                            print(f"Extrascreen update response for {file_ref}:")
                            print(json.dumps(result, indent=2, default=str))
                        print(f"Extrascreen updated: {file_ref} -> docscreenid={docscreenid}")
                        updated += 1
                    except Exception as exc:
                        failed += 1
                        report_lines.append(f"Extrascreen update failed for {file_ref}: {exc}")
                        print(f"Extrascreen update failed for {file_ref}: {exc}", file=sys.stderr)

        report_lines.append(
            "Extrascreen summary: processed={processed}, updated={updated}, failed={failed}".format(
                processed=processed,
                updated=updated,
                failed=failed,
            )
        )

    def _resolve_feedback_files(self) -> list[tuple[str, list[tuple[str, str, bool]]]]:
        files: list[tuple[str, list[tuple[str, str, bool]]]] = []
        debt_review = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Debt Review",
            "Debt_Review_Feedback_APT_LSW",
            f"Standard_Bank_Panel_L_Update_{self._date_ctx.date_str}_DR.xlsx",
        )
        if os.path.exists(debt_review):
            files.append((debt_review, FEEDBACK_FIELD_MAP))

        panel_l_pattern = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Panel L",
            "Feedback_APT_LSW",
            self._date_ctx.month_year,
            f"*_{self._date_ctx.date_str}.xlsx",
        )
        for path in sorted(glob.glob(panel_l_pattern)):
            files.append((path, FEEDBACK_FIELD_MAP))
        return files

    def _resolve_ptp_files(self) -> list[tuple[str, list[tuple[str, str, bool]]]]:
        files: list[tuple[str, list[tuple[str, str, bool]]]] = []
        debt_review = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Debt Review",
            "Debt_Review_PTP_APT_LSW",
            f"Standard_Bank_Panel_L_PTP_{self._date_ctx.date_str}_DR.xlsx",
        )
        if os.path.exists(debt_review):
            files.append((debt_review, PTP_FIELD_MAP))

        panel_l_pattern = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Panel L",
            "PTP_APT_LSW",
            self._date_ctx.month_year,
            f"*_{self._date_ctx.date_str}.xlsx",
        )
        for path in sorted(glob.glob(panel_l_pattern)):
            files.append((path, PTP_FIELD_MAP))
        return files

    def _resolve_poc_summons_files(self) -> list[tuple[str, list[tuple[str, str, bool]]]]:
        files: list[tuple[str, list[tuple[str, str, bool]]]] = []
        poc_summons_path = os.path.join(
            self._args.cleaned_dir,
            "SBSA POC AND SUMMONS",
            f"{self._date_ctx.day}{self._date_ctx.month}{self._date_ctx.year}.xlsx",
        )
        if os.path.exists(poc_summons_path):
            files.append((poc_summons_path, POC_SUMMONS_FIELD_MAP))
        return files

    def _filter_extrascreen_files(
        self,
        files: list[tuple[str, list[tuple[str, str, bool]]]],
        allowed: set[str],
    ) -> list[tuple[str, list[tuple[str, str, bool]]]]:
        filtered: list[tuple[str, list[tuple[str, str, bool]]]] = []
        for path, mapping in files:
            if self._extrascreen_mapping_label(mapping) in allowed:
                filtered.append((path, mapping))
        return filtered

    @staticmethod
    def _extrascreen_mapping_label(mapping: list[tuple[str, str, bool]]) -> str:
        if mapping is FEEDBACK_FIELD_MAP:
            return "feedback"
        if mapping is PTP_FIELD_MAP:
            return "ptp"
        if mapping is POC_SUMMONS_FIELD_MAP:
            return "poc-summons"
        return "extrascreen"

    def _build_extrascreen_column_map(
        self,
        header_row: list[object],
        mapping: list[tuple[str, str, bool]],
    ) -> tuple[dict[str, int], int | None, int | None]:
        col_map: dict[str, int] = {}
        file_ref_idx = None
        screen_id_idx = None
        for idx, value in enumerate(header_row, start=1):
            key = self._normalize_header(value)
            if key in {"fileref", "filereference"}:
                file_ref_idx = idx
            if key == "desktopextrascreenid":
                screen_id_idx = idx
            for header_key, field_name, _ in mapping:
                if key == header_key:
                    col_map[field_name] = idx
        return col_map, file_ref_idx, screen_id_idx

    def _build_extrascreen_payload(
        self,
        row: tuple[object, ...],
        col_map: dict[str, int],
        mapping: list[tuple[str, str, bool]],
    ) -> dict:
        payload: dict[str, object] = {}
        for header_key, field_name, is_date in mapping:
            col_idx = col_map.get(field_name)
            if not col_idx or col_idx > len(row):
                continue
            raw_value = row[col_idx - 1]
            if raw_value in (None, ""):
                continue
            if is_date:
                encoded = self._encode_legalsuite_date(raw_value)
                if encoded is None:
                    continue
                payload[field_name] = encoded
            else:
                payload[field_name] = self._normalize_cell_value(raw_value)
        return payload

    @staticmethod
    def _normalize_cell_value(value: object) -> object:
        if isinstance(value, float) and value.is_integer():
            return int(value)
        if isinstance(value, (dt.datetime, dt.date)):
            return value.strftime("%Y-%m-%d")
        return value

    @staticmethod
    def _encode_legalsuite_date(value: object) -> int | None:
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            return int(value) + LEGALSUITE_OFFSET
        if isinstance(value, dt.datetime):
            dt_value = value
        elif isinstance(value, dt.date):
            dt_value = dt.datetime.combine(value, dt.time())
        else:
            text = str(value).strip()
            if not text:
                return None
            date_text = text.split()[0].split("T")[0].replace("/", "-")
            parts = date_text.split("-")
            if len(parts) == 3 and all(parts):
                if len(parts[0]) == 4:
                    date_text = f"{parts[0]}-{parts[1]}-{parts[2]}"
                elif len(parts[2]) == 4:
                    date_text = f"{parts[2]}-{parts[1]}-{parts[0]}"
            try:
                dt_value = dt.datetime.strptime(date_text, "%Y-%m-%d")
            except ValueError:
                return None
        excel_serial = (dt_value - EXCEL_BASE).days
        return excel_serial + LEGALSUITE_OFFSET

    @staticmethod
    def _cell_text(row: tuple[object, ...], col_idx: int) -> str | None:
        if col_idx > len(row):
            return None
        value = row[col_idx - 1]
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def _archive_closed_matters(self, report_lines: list[str]) -> None:
        if load_workbook is None:
            report_lines.append("Archive skipped: openpyxl is not installed.")
            print("Archive skipped: openpyxl is not installed.", file=sys.stderr)
            return

        api_key = self._args.api_key or os.getenv("LEGALSUITE_API_KEY") or LEGALSUITE_API_KEY
        if not api_key:
            report_lines.append("Archive skipped: missing API key (use --api-key or LEGALSUITE_API_KEY).")
            print("Archive skipped: missing API key.", file=sys.stderr)
            return

        client = LegalSuiteClient(self._args.api_base, api_key)
        logged_in_employee_id = self._args.logged_in_employee_id or LEGALSUITE_EMPLOYEE_ID
        archive_status = self._args.archive_status or LEGALSUITE_ARCHIVE_STATUS
        closed_dir_panel = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Panel L",
            "Closed_APT_LSW",
            self._date_ctx.month_year,
        )
        closed_dir_debt = os.path.join(
            self._args.cleaned_dir,
            "SBSA",
            "Debt Review",
            "Debt_Review_Close_APT_LSW",
        )
        panel_pattern = os.path.join(closed_dir_panel, f"*_{self._date_ctx.date_str}.xlsx")
        debt_pattern = os.path.join(
            closed_dir_debt,
            f"Standard_Bank_Panel_L_Close_{self._date_ctx.date_str}_DR.xlsx",
        )
        cleaned_files = sorted(glob.glob(panel_pattern)) + sorted(glob.glob(debt_pattern))
        if not cleaned_files:
            report_lines.append(
                f"No cleaned closed files found: {panel_pattern} or {debt_pattern}"
            )
            print(f"No cleaned closed files found: {panel_pattern} or {debt_pattern}")
            return

        file_refs: set[str] = set()
        missing_header_files = 0
        for cleaned_path in cleaned_files:
            download_path = self._download_path_for_cleaned(cleaned_path)
            col_info = self._find_fileref_column(download_path, cleaned_path)
            if not col_info:
                report_lines.append(f"FileRef column not found: {cleaned_path}")
                missing_header_files += 1
                continue
            col_idx, col_name = col_info
            refs = self._collect_file_refs(cleaned_path, col_idx)
            if not refs:
                report_lines.append(f"No file refs found using {col_name}: {cleaned_path}")
            file_refs.update(refs)

        if not file_refs:
            report_lines.append("No file references found to archive.")
            print("No file references found to archive.")
            return
        print(f"Found {len(file_refs)} file references to process.")

        archived = 0
        failed = 0
        for file_ref in sorted(file_refs):
            try:
                matter = client.get_matter_by_fileref(file_ref)
                if self._args.archive_dry_run:
                    archived += 1
                    print(f"Dry-run: {file_ref}")
                    print(matter)
                    continue
                archive_no = matter.get("archiveno") or matter.get("archive_no") or matter.get("archivenumber")
                payload = client.build_archive_payload(
                    matter=matter,
                    logged_in_employee_id=logged_in_employee_id,
                    archive_no=archive_no,
                    archive_status=archive_status,
                )
                if self._args.archive_verbose:
                    print(f"Archive request for {file_ref}:")
                    print(json.dumps(payload, indent=2, default=str))
                result = client.update_matter(payload)
                if self._args.archive_verbose:
                    print(f"Archive response for {file_ref}:")
                    print(json.dumps(result, indent=2, default=str))
                archived += 1
                report_lines.append(f"Archived matter: {file_ref}")
                print(f"Archived matter: {file_ref}")
            except Exception as exc:
                failed += 1
                report_lines.append(f"Archive failed for {file_ref}: {exc}")
                print(f"Archive failed for {file_ref}: {exc}", file=sys.stderr)

        report_lines.append(
            "Archive summary: archived={archived}, failed={failed}, files_missing_header={missing}".format(
                archived=archived,
                failed=failed,
                missing=missing_header_files,
            )
        )

    def _update_claim_amounts(self, report_lines: list[str]) -> None:
        if load_workbook is None:
            report_lines.append("Claim amount update skipped: openpyxl is not installed.")
            print("Claim amount update skipped: openpyxl is not installed.", file=sys.stderr)
            return

        api_key = self._args.api_key or os.getenv("LEGALSUITE_API_KEY") or LEGALSUITE_API_KEY
        if not api_key:
            report_lines.append("Claim amount update skipped: missing API key.")
            print("Claim amount update skipped: missing API key.", file=sys.stderr)
            return

        claim_files = self._resolve_claim_amount_files()
        if not claim_files:
            report_lines.append("Claim amount update skipped: no matching files found.")
            print("Claim amount update skipped: no matching files found.")
            return

        client = LegalSuiteClient(self._args.api_base, api_key)
        logged_in_employee_id = self._args.logged_in_employee_id or LEGALSUITE_EMPLOYEE_ID
        processed = 0
        updated = 0
        failed = 0
        missing_headers = 0

        for claim_path in claim_files:
            header_row = self._read_header_row(claim_path)
            if not header_row:
                report_lines.append(f"Claim amount header not found: {claim_path}")
                missing_headers += 1
                continue
            file_ref_idx, claim_amount_idx = self._find_claim_amount_columns(header_row)
            if not file_ref_idx or not claim_amount_idx:
                report_lines.append(f"Claim amount columns missing: {claim_path}")
                missing_headers += 1
                continue

            workbook = load_workbook(claim_path, read_only=False, data_only=True)
            for worksheet in workbook.worksheets:
                max_col = worksheet.max_column or max(file_ref_idx, claim_amount_idx)
                for row in worksheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
                    file_ref = self._cell_text(row, file_ref_idx)
                    claim_amount = self._normalize_claim_amount(self._cell_value(row, claim_amount_idx))
                    if not file_ref or claim_amount is None:
                        continue
                    try:
                        matter = client.get_matter_by_fileref(file_ref)
                        payload = client.build_claim_amount_payload(
                            matter=matter,
                            logged_in_employee_id=logged_in_employee_id,
                            claim_amount=claim_amount,
                        )
                        processed += 1
                        if self._args.claim_amount_dry_run:
                            print(f"Claim amount dry-run: {file_ref} -> {claim_amount}")
                            if self._args.claim_amount_verbose:
                                print(json.dumps(payload, indent=2, default=str))
                            continue
                        if self._args.claim_amount_verbose:
                            print(f"Claim amount update payload for {file_ref}:")
                            print(json.dumps(payload, indent=2, default=str))
                        result = client.update_matter(payload)
                        if self._args.claim_amount_verbose:
                            print(f"Claim amount update response for {file_ref}:")
                            print(json.dumps(result, indent=2, default=str))
                        print(f"Claim amount updated: {file_ref} -> {claim_amount}")
                        updated += 1
                    except Exception as exc:
                        failed += 1
                        report_lines.append(f"Claim amount update failed for {file_ref}: {exc}")
                        print(f"Claim amount update failed for {file_ref}: {exc}", file=sys.stderr)

        report_lines.append(
            "Claim amount summary: processed={processed}, updated={updated}, "
            "failed={failed}, files_missing_header={missing}".format(
                processed=processed,
                updated=updated,
                failed=failed,
                missing=missing_headers,
            )
        )

    def _download_path_for_cleaned(self, cleaned_path: str) -> str:
        rel_path = os.path.relpath(cleaned_path, self._args.cleaned_dir)
        return os.path.join(self._args.download_dir, rel_path)

    def _collect_file_refs(self, cleaned_path: str, col_idx: int) -> list[str]:
        refs: list[str] = []
        workbook = load_workbook(cleaned_path, read_only=False, data_only=True)
        for worksheet in workbook.worksheets:
            max_col = worksheet.max_column or col_idx
            for row in worksheet.iter_rows(min_row=1, max_col=max_col, values_only=True):
                if not row or len(row) < col_idx:
                    continue
                value = row[col_idx - 1]
                if value is None:
                    continue
                text = str(value).strip()
                if not text or self._is_header_value(text):
                    continue
                refs.append(text)
        return refs

    def _find_fileref_column(self, download_path: str, cleaned_path: str) -> tuple[int, str] | None:
        header_row = self._read_header_row(download_path) or self._read_header_row(cleaned_path)
        if not header_row:
            return None
        for idx, value in enumerate(header_row, start=1):
            key = self._normalize_header(value)
            if key in {"fileref", "filereference"}:
                return idx, "fileref"
        return None

    def _read_header_row(self, path: str) -> list[object] | None:
        if not os.path.exists(path):
            return None
        workbook = load_workbook(path, read_only=False, data_only=True)
        if not workbook.worksheets:
            return None
        worksheet = workbook.worksheets[0]
        max_col = worksheet.max_column or 1
        for row in worksheet.iter_rows(min_row=1, max_row=1, max_col=max_col, values_only=True):
            return list(row)
        return None

    def _resolve_claim_amount_files(self) -> list[str]:
        pattern = os.path.join(
            self._args.cleaned_dir,
            "Standard Bank_ClaimsAmount",
            f"Standard Bank Legal Claim Amount_Panel_L{self._date_ctx.year}_{self._date_ctx.month}_{self._date_ctx.day}_*.xlsx",
        )
        return sorted(glob.glob(pattern))

    def _find_claim_amount_columns(self, header_row: list[object]) -> tuple[int | None, int | None]:
        file_ref_idx = None
        claim_amount_idx = None
        for idx, value in enumerate(header_row, start=1):
            key = self._normalize_header(value)
            if key in {"fileref", "filereference"}:
                file_ref_idx = idx
            if key in {"claimamount", "claimamounts", "claimamt"}:
                claim_amount_idx = idx
        return file_ref_idx, claim_amount_idx

    @staticmethod
    def _cell_value(row: tuple[object, ...], col_idx: int) -> object | None:
        if col_idx > len(row):
            return None
        return row[col_idx - 1]

    @staticmethod
    def _normalize_claim_amount(value: object) -> object | None:
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            return value
        text = str(value).strip()
        if not text:
            return None
        cleaned = re.sub(r"[^0-9.\-]", "", text)
        if not cleaned or cleaned in {"-", ".", "-."}:
            return None
        try:
            if "." in cleaned:
                return float(cleaned)
            return int(cleaned)
        except ValueError:
            return None

    @staticmethod
    def _normalize_header(header: object) -> str:
        if header is None:
            return ""
        return "".join(ch for ch in str(header).strip().lower() if ch.isalnum())

    @staticmethod
    def _is_header_value(value: str) -> bool:
        return App._normalize_header(value) in {"fileref", "filereference"}

    @staticmethod
    def _resolve_date(date_arg: str | None, days_ago: int) -> DateContext:
        if date_arg:
            try:
                date_val = dt.datetime.strptime(date_arg, "%Y%m%d")
            except ValueError as exc:
                raise ValueError("Date must be in YYYYMMDD format.") from exc
        else:
            if days_ago < 0:
                raise ValueError("days_ago must be 0 or greater.")
            date_val = dt.datetime.now() - dt.timedelta(days=days_ago)
        return DateContext(
            date_str=date_val.strftime("%Y%m%d"),
            month_year=date_val.strftime("%b %Y"),
            year=date_val.strftime("%Y"),
            month=date_val.strftime("%m"),
            day=date_val.strftime("%d"),
        )

    @staticmethod
    def _build_targets(date_ctx: DateContext) -> list[tuple[str, str]]:
        items: list[tuple[str, str]] = []
        for dir_tmpl, file_tmpl in TARGETS:
            remote_dir = dir_tmpl.format(
                date=date_ctx.date_str,
                month_year=date_ctx.month_year,
                year=date_ctx.year,
                month=date_ctx.month,
                day=date_ctx.day,
            )
            filename = file_tmpl.format(
                date=date_ctx.date_str,
                month_year=date_ctx.month_year,
                year=date_ctx.year,
                month=date_ctx.month,
                day=date_ctx.day,
            )
            items.append((remote_dir, filename))
        return items

    @staticmethod
    def _ensure_local_path(base_dir: str, remote_dir: str, filename: str) -> str:
        local_dir = os.path.join(base_dir, *remote_dir.split("/"))
        os.makedirs(local_dir, exist_ok=True)
        return os.path.join(local_dir, filename)

    @staticmethod
    def _normalize_cli_args(argv: list[str]) -> list[str]:
        normalized: list[str] = []
        for arg in argv:
            match = re.fullmatch(r"--days-(\d+)", arg)
            if match:
                normalized.extend(["--days-ago", match.group(1)])
                continue
            normalized.append(arg)
        return normalized

    @staticmethod
    def parse_args() -> argparse.Namespace:
        # Example: python3 ftp_download_today.py --date 20260402 --clean-only --log-file logs/report.txt
        parser = argparse.ArgumentParser(
            description="Download LegalSuite FTP files for a specific date (default: today)."
        )
        parser.add_argument(
            "--date",
            help="Date in YYYYMMDD format; defaults to today.",
        )
        parser.add_argument(
            "--days-ago",
            type=int,
            default=0,
            help="Download files from N days ago (default: 0). Also accepts shorthand like --days-15. Ignored if --date is provided.",
        )
        parser.add_argument(
            "--download-dir",
            default="downloads",
            help="Local base directory for downloads (default: downloads).",
        )
        parser.add_argument(
            "--timeout",
            type=int,
            default=30,
            help="FTP connection timeout in seconds (default: 30).",
        )
        parser.add_argument(
            "--log-file",
            help="Write a report log to this path (default: <download-dir>/report_YYYYMMDD.txt).",
        )
        parser.add_argument(
            "--cleaned-dir",
            default="cleaned",
            help="Local base directory for cleaned files (default: cleaned).",
        )
        parser.add_argument(
            "--skip-clean",
            action="store_true",
            help="Skip cleaning downloaded files.",
        )
        parser.add_argument(
            "--clean-only",
            action="store_true",
            help="Reuse existing selected-date downloads; do not connect to FTP.",
        )
        parser.add_argument(
            "--archive-closed",
            action="store_true",
            help="Update closed matters in LegalSuite using cleaned closed files.",
        )
        parser.add_argument(
            "--archive-dry-run",
            action="store_true",
            help="Only fetch matters for closed files; do not update.",
        )
        parser.add_argument(
            "--archive-verbose",
            action="store_true",
            help="Print LegalSuite update payload and response to the console.",
        )
        parser.add_argument(
            "--update-extrascreen",
            action="store_true",
            help="Update Matter ExtraScreen data using feedback/PTP files.",
        )
        parser.add_argument(
            "--extrascreen-only",
            choices=["feedback", "ptp", "poc-summons"],
            help="Limit extrascreen updates to one file type.",
        )
        parser.add_argument(
            "--extrascreen-dry-run",
            action="store_true",
            help="Only fetch matters for extrascreen updates; do not update.",
        )
        parser.add_argument(
            "--extrascreen-verbose",
            action="store_true",
            help="Print extrascreen update payload and response to the console.",
        )
        parser.add_argument(
            "--update-claim-amount",
            action="store_true",
            help="Update LegalSuite claim amount using the claims file.",
        )
        parser.add_argument(
            "--claim-amount-dry-run",
            action="store_true",
            help="Only fetch matters for claim updates; do not update.",
        )
        parser.add_argument(
            "--claim-amount-verbose",
            action="store_true",
            help="Print claim amount update payload and response to the console.",
        )
        parser.add_argument(
            "--api-base",
            default=LEGALSUITE_API_BASE,
            help="LegalSuite API base URL (default: https://api.legalsuite.net).",
        )
        parser.add_argument(
            "--api-key",
            default=os.getenv("LEGALSUITE_API_KEY") or LEGALSUITE_API_KEY,
            help="LegalSuite API key (or set LEGALSUITE_API_KEY).",
        )
        parser.add_argument(
            "--logged-in-employee-id",
            default=LEGALSUITE_EMPLOYEE_ID,
            help="LegalSuite logged-in employee ID for updates (default: 1).",
        )
        parser.add_argument(
            "--archive-status",
            default=LEGALSUITE_ARCHIVE_STATUS,
            help="Archive status ID to send with updates (default: 2).",
        )
        return parser.parse_args(App._normalize_cli_args(sys.argv[1:]))


def main() -> int:
    try:
        args = App.parse_args()
        app = App(args)
        return app.run()
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
