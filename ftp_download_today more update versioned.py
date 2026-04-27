#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import fnmatch
import ftplib
import os
import sys
import warnings
from typing import Callable

from env_config import load_env_file

load_env_file()

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - handled at runtime
    load_workbook = None
    Workbook = None
else:
    warnings.filterwarnings(
        "ignore",
        message="Workbook contains no default style, apply openpyxl's default",
        module="openpyxl.styles.stylesheet",
    )

FTP_HOST = os.getenv("FTP_HOST", "")
FTP_USER = os.getenv("FTP_USER", "")
FTP_PASS = os.getenv("FTP_PASS", "")

TARGETS = [
    ("SBSA/Debt Review/Debt_Review_Close_APT_LSW", "Standard_Bank_Panel_L_Close_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_PTP_APT_LSW", "Standard_Bank_Panel_L_PTP_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_Feedback_APT_LSW", "Standard_Bank_Panel_L_Update_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_Reopen_APT_LSW", "Standard_Bank_Panel_L_Reopen_{date}_DR.xlsx"),
    ("SBSA/Debt Review/Debt_Review_Handover_APT_LWS", "Standard_Bank_Panel_L_Handover_{date}_DR.xlsx"),
    ("SBSA/Panel L/PTP_APT_LSW/{month_year}", "Standard_Bank_Panel_L_PTP_{date}.xlsx"),
    ("SBSA/Panel L/Feedback_APT_LSW/{month_year}", "Standard_Bank_Panel_L_PTP_{date}.xlsx"),
    ("SBSA/Panel L/Handover_APT_LSW/{month_year}", "Standard_Bank_Panel_L_PTP_{date}.xlsx"),
    ("SBSA/Panel L/Closed_APT_LSW/{month_year}", "Standard_Bank_Panel_L_PTP_{date}.xlsx"),
    ("SBSA/Panel L/Reopen_APT_LSW/{month_year}", "Standard_Bank_Panel_L_PTP_{date}.xlsx"),
    ("Standard Bank_ClaimsAmount", "Standard Bank Legal Claim Amount_Panel_L{year}_{month}_{day}_*.xlsx"),
    ("SBSA POC AND SUMMONS", "{day}{month}{year}.csv"),
]

HEADER_RULES: dict[str, Callable[[object], object]] = {}
CLAIM_AMOUNT_PATTERN = "Standard Bank Legal Claim Amount_Panel_L*.xlsx"


def parse_args() -> argparse.Namespace:
    # Example: python3 ftp_download_today.py --date 20260402 --clean-only --log-file logs/report.txt
    parser = argparse.ArgumentParser(
        description="Download LegalSuite FTP files for a specific date (default: today)."
    )
    parser.add_argument(
        "--date",
        help="Date in YYYYMMDD format; defaults to today.",
    )
    parser.add_argument(
        "--download-dir",
        default="downloads",
        help="Local base directory for downloads (default: downloads).",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="FTP connection timeout in seconds (default: 30).",
    )
    parser.add_argument(
        "--log-file",
        help="Write a report log to this path (default: <download-dir>/report_YYYYMMDD.txt).",
    )
    parser.add_argument(
        "--cleaned-dir",
        default="cleaned",
        help="Local base directory for cleaned files (default: cleaned).",
    )
    parser.add_argument(
        "--skip-clean",
        action="store_true",
        help="Skip cleaning downloaded files.",
    )
    parser.add_argument(
        "--clean-only",
        action="store_true",
        help="Only clean existing downloads; do not connect to FTP.",
    )
    return parser.parse_args()


def resolve_date(date_arg: str | None) -> tuple[str, str]:
    if date_arg:
        try:
            date_val = dt.datetime.strptime(date_arg, "%Y%m%d")
        except ValueError as exc:
            raise ValueError("Date must be in YYYYMMDD format.") from exc
    else:
        date_val = dt.datetime.now()
    date_str = date_val.strftime("%Y%m%d")
    month_year = date_val.strftime("%b %Y")
    year = date_val.strftime("%Y")
    month = date_val.strftime("%m")
    day = date_val.strftime("%d")
    return date_str, month_year, year, month, day


def build_targets(date_str: str, month_year: str, year: str, month: str, day: str) -> list[tuple[str, str]]:
    items: list[tuple[str, str]] = []
    for dir_tmpl, file_tmpl in TARGETS:
        remote_dir = dir_tmpl.format(date=date_str, month_year=month_year, year=year, month=month, day=day)
        filename = file_tmpl.format(date=date_str, month_year=month_year, year=year, month=month, day=day)
        items.append((remote_dir, filename))
    return items


def normalize_header(header: object) -> str:
    if header is None:
        return ""
    return "".join(ch for ch in str(header).strip().lower() if ch.isalnum())


def digits_only(value: object) -> object:
    if value is None:
        return None
    text = str(value)
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits


def build_header_rules() -> dict[str, Callable[[object], object]]:
    return {
        "accountnumber": digits_only,
    }


def ensure_local_path(base_dir: str, remote_dir: str, filename: str) -> str:
    local_dir = os.path.join(base_dir, *remote_dir.split("/"))
    os.makedirs(local_dir, exist_ok=True)
    return os.path.join(local_dir, filename)


def list_dir(ftp: ftplib.FTP, remote_dir: str) -> list[str] | None:
    try:
        return ftp.nlst(remote_dir)
    except ftplib.error_perm:
        pass
    current_dir = None
    try:
        current_dir = ftp.pwd()
        ftp.cwd(remote_dir)
        listing = ftp.nlst()
        return [f"{remote_dir}/{name}" for name in listing]
    except ftplib.error_perm:
        return None
    finally:
        if current_dir:
            try:
                ftp.cwd(current_dir)
            except ftplib.all_errors:
                pass


def mdtm_timestamp(ftp: ftplib.FTP, remote_path: str) -> dt.datetime | None:
    try:
        response = ftp.sendcmd(f"MDTM {remote_path}")
    except ftplib.all_errors:
        return None
    parts = response.split()
    if len(parts) != 2 or parts[0] != "213":
        return None
    try:
        return dt.datetime.strptime(parts[1], "%Y%m%d%H%M%S")
    except ValueError:
        return None


def select_newest_by_mdtm(ftp: ftplib.FTP, remote_dir: str, names: list[str]) -> str:
    candidates: list[tuple[dt.datetime, str]] = []
    for name in names:
        ts = mdtm_timestamp(ftp, f"{remote_dir}/{name}")
        if ts:
            candidates.append((ts, name))
    if candidates:
        candidates.sort()
        return candidates[-1][1]
    names.sort()
    return names[-1]


def resolve_remote_file(
    ftp: ftplib.FTP, remote_dir: str, filename_tmpl: str
) -> tuple[str | None, str | None]:
    listing = list_dir(ftp, remote_dir)
    if listing is None:
        return None, "missing_dir"
    names = [os.path.basename(item) for item in listing]
    if any(ch in filename_tmpl for ch in ["*", "?", "["]):
        matches = [name for name in names if fnmatch.fnmatch(name, filename_tmpl)]
        if not matches:
            return None, "missing_file"
        return select_newest_by_mdtm(ftp, remote_dir, matches), None
    if filename_tmpl in names:
        return filename_tmpl, None
    try:
        ftp.size(f"{remote_dir}/{filename_tmpl}")
        return filename_tmpl, None
    except ftplib.all_errors:
        return None, "missing_file"


def download_file(ftp: ftplib.FTP, remote_path: str, local_path: str) -> None:
    with open(local_path, "wb") as handle:
        ftp.retrbinary(f"RETR {remote_path}", handle.write)


def apply_cleaning(
    workbook,
    rules: dict[str, Callable[[object], object]],
    blank_headers: set[str],
    drop_header: bool,
    copy_reference: bool,
) -> tuple[int, int, int, int, int]:
    cleaned_cells = 0
    matched_columns = 0
    blanked_columns = 0
    headers_removed = 0
    copied_cells = 0
    for worksheet in workbook.worksheets:
        header_row = list(worksheet.iter_rows(min_row=1, max_row=1))
        if not header_row:
            continue
        header_cells = header_row[0]
        column_rules: dict[int, Callable[[object], object]] = {}
        blank_columns: set[int] = set()
        account_col = None
        reference_col = None
        for idx, cell in enumerate(header_cells, start=1):
            key = normalize_header(cell.value)
            if key in rules:
                column_rules[idx] = rules[key]
            if key in blank_headers:
                blank_columns.add(idx)
            if key == "accountnumber":
                account_col = idx
            if key == "reference":
                reference_col = idx
        matched_columns += len(column_rules)
        blanked_columns += len(blank_columns)
        if not column_rules and not blank_columns and not (copy_reference and account_col and reference_col):
            continue
        for row in worksheet.iter_rows(min_row=2):
            if copy_reference and account_col and reference_col:
                account_cell = row[account_col - 1]
                reference_cell = row[reference_col - 1]
                if account_cell.value != reference_cell.value:
                    account_cell.value = reference_cell.value
                    copied_cells += 1
            for col_idx, rule in column_rules.items():
                cell = row[col_idx - 1]
                new_value = rule(cell.value)
                if new_value != cell.value:
                    cell.value = new_value
                    cleaned_cells += 1
            for col_idx in blank_columns:
                cell = row[col_idx - 1]
                if cell.value not in (None, ""):
                    cell.value = None
                    cleaned_cells += 1
        if drop_header and worksheet.max_row >= 1:
            worksheet.delete_rows(1)
            headers_removed += 1
    return cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells


def clean_excel_file(
    source_path: str,
    destination_path: str,
    rules: dict[str, Callable[[object], object]],
    blank_headers: set[str],
    copy_reference: bool,
) -> tuple[int, int, int, int, int]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed")
    workbook = load_workbook(source_path)
    cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells = apply_cleaning(
        workbook, rules, blank_headers, True, copy_reference
    )
    os.makedirs(os.path.dirname(destination_path) or ".", exist_ok=True)
    workbook.save(destination_path)
    return cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells


def convert_csv_to_excel(
    source_path: str,
    destination_path: str,
    rules: dict[str, Callable[[object], object]],
    blank_headers: set[str],
    copy_reference: bool,
) -> tuple[int, int, int, int, int]:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed")
    workbook = Workbook()
    worksheet = workbook.active
    with open(source_path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        for row in reader:
            worksheet.append(row)
    cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells = apply_cleaning(
        workbook, rules, blank_headers, True, copy_reference
    )
    os.makedirs(os.path.dirname(destination_path) or ".", exist_ok=True)
    workbook.save(destination_path)
    return cleaned_cells, matched_columns, blanked_columns, headers_removed, copied_cells


def clean_downloads(
    download_dir: str,
    cleaned_dir: str,
    rules: dict[str, Callable[[object], object]],
    report_lines: list[str],
) -> None:
    if load_workbook is None:
        report_lines.append("Cleaning skipped: openpyxl is not installed.")
        print("Cleaning skipped: openpyxl is not installed.", file=sys.stderr)
        return

    cleaned_files = 0
    converted_files = 0
    skipped_files = 0
    failed_files = 0
    total_cells = 0
    matched_columns = 0
    blanked_columns = 0
    headers_removed = 0
    copied_cells = 0
    cleaned_dir_abs = os.path.abspath(cleaned_dir)
    for root, _, files in os.walk(download_dir):
        root_abs = os.path.abspath(root)
        if root_abs.startswith(cleaned_dir_abs):
            continue
        for name in files:
            lowered = name.lower()
            if not (lowered.endswith(".xlsx") or lowered.endswith(".csv")):
                continue
            if name.startswith("~$"):
                skipped_files += 1
                continue
            source_path = os.path.join(root, name)
            rel_path = os.path.relpath(source_path, download_dir)
            if lowered.endswith(".csv"):
                rel_path = os.path.splitext(rel_path)[0] + ".xlsx"
            destination_path = os.path.join(cleaned_dir, rel_path)
            blank_headers = set()
            if fnmatch.fnmatch(name, CLAIM_AMOUNT_PATTERN):
                blank_headers.add("matter")
            rel_norm = rel_path.replace(os.sep, "/")
            copy_reference = (
                rel_norm.startswith("SBSA/Debt Review/Debt_Review_Handover_APT_LWS/")
                or rel_norm.startswith("SBSA/Panel L/Handover_APT_LSW/")
            )
            try:
                if lowered.endswith(".csv"):
                    cleaned, matched, blanked, removed, copied = convert_csv_to_excel(
                        source_path, destination_path, rules, blank_headers, copy_reference
                    )
                    converted_files += 1
                    action = "Converted"
                    try:
                        os.remove(source_path)
                        report_lines.append(f"Deleted original CSV: {source_path}")
                    except OSError as exc:
                        report_lines.append(f"Failed to delete CSV: {source_path} ({exc})")
                else:
                    cleaned, matched, blanked, removed, copied = clean_excel_file(
                        source_path, destination_path, rules, blank_headers, copy_reference
                    )
                    cleaned_files += 1
                    action = "Cleaned"
                total_cells += cleaned
                matched_columns += matched
                blanked_columns += blanked
                headers_removed += removed
                copied_cells += copied
                report_lines.append(
                    f"{action}: {source_path} -> {destination_path} "
                    f"(cells updated: {cleaned})"
                )
            except Exception as exc:
                failed_files += 1
                report_lines.append(f"Clean failed: {source_path} ({exc})")
                print(f"Clean failed: {source_path}: {exc}", file=sys.stderr)

    report_lines.append(
        "Clean summary: cleaned_files={cleaned}, converted_files={converted}, "
        "skipped_files={skipped}, failed_files={failed}, "
        "columns_matched={columns}, blanked_columns={blanked}, "
        "headers_removed={headers}, copied_cells={copied}, cells_updated={cells}".format(
            cleaned=cleaned_files,
            converted=converted_files,
            skipped=skipped_files,
            failed=failed_files,
            columns=matched_columns,
            blanked=blanked_columns,
            headers=headers_removed,
            copied=copied_cells,
            cells=total_cells,
        )
    )


def main() -> int:
    args = parse_args()
    try:
        date_str, month_year, year, month, day = resolve_date(args.date)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 2

    targets = build_targets(date_str, month_year, year, month, day)
    report_lines: list[str] = []
    report_lines.append(f"Report date: {date_str}")
    report_lines.append(f"Month folder: {month_year}")

    downloaded = 0
    missing_dirs = 0
    missing_files = 0
    failed_downloads = 0
    if args.clean_only:
        report_lines.append("Download skipped: clean-only mode.")
    else:
        print("Connecting to FTP...")
        try:
            ftp = ftplib.FTP(FTP_HOST, timeout=args.timeout)
            ftp.login(FTP_USER, FTP_PASS)
            ftp.set_pasv(True)
        except ftplib.all_errors as exc:
            error_line = f"FTP connection failed: {exc}"
            print(error_line, file=sys.stderr)
            report_lines.append(error_line)
            return 1

        try:
            print("Downloading files...")
            for remote_dir, filename in targets:
                remote_path = f"{remote_dir}/{filename}"
                resolved_name, reason = resolve_remote_file(ftp, remote_dir, filename)
                if not resolved_name:
                    if reason == "missing_dir":
                        line = f"Missing directory: {remote_dir}"
                        print(line)
                        report_lines.append(line)
                        missing_dirs += 1
                    else:
                        line = f"Missing file: {remote_path}"
                        print(line)
                        report_lines.append(line)
                        missing_files += 1
                    continue

                remote_path = f"{remote_dir}/{resolved_name}"
                local_path = ensure_local_path(args.download_dir, remote_dir, resolved_name)
                try:
                    download_file(ftp, remote_path, local_path)
                    line = f"Downloaded: {remote_path} -> {local_path}"
                    print(line)
                    report_lines.append(line)
                    downloaded += 1
                except ftplib.all_errors as exc:
                    line = f"Download failed for {remote_path}: {exc}"
                    print(line, file=sys.stderr)
                    report_lines.append(line)
                    failed_downloads += 1
        finally:
            try:
                ftp.quit()
            except ftplib.all_errors:
                ftp.close()

    if not args.skip_clean:
        rules = build_header_rules()
        if rules:
            print("Cleaning downloaded files...")
            clean_downloads(args.download_dir, args.cleaned_dir, rules, report_lines)
        else:
            report_lines.append("Cleaning skipped: no header rules configured.")

    report_lines.append(f"Summary: downloaded={downloaded}, missing_dirs={missing_dirs}, "
                        f"missing_files={missing_files}, failed_downloads={failed_downloads}")

    log_path = args.log_file
    if not log_path:
        log_path = os.path.join(args.download_dir, f"report_{date_str}.txt")
    os.makedirs(os.path.dirname(log_path) or ".", exist_ok=True)
    try:
        print("Writing report log...")
        with open(log_path, "w", encoding="utf-8") as handle:
            handle.write("\n".join(report_lines) + "\n")
    except OSError as exc:
        print(f"Failed to write log file {log_path}: {exc}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
